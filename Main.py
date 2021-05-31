from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference, Series
from PIL import Image, ImageDraw, ImageFont
from PIL import Image
from datetime import date
from rich.console import Console
from rich.text import Text
from rich import print
from rich.panel import Panel
from rich.prompt import Prompt
from rich.table import Table
from rich.progress import track
from rich.prompt import Confirm
from rich.theme import Theme
import openpyxl
import glob
import requests
import random
import time
import shutil
import urllib2
import os, sys
import pickle
import binascii
import ssl
import os.path
from concurrent.futures import as_completed, ThreadPoolExecutor
import signal
from functools import partial
from threading import Event
from typing import Iterable
from urllib.request import urlopen
from rich.progress import (
    BarColumn,
    DownloadColumn,
    Progress,
    TaskID,
    TextColumn,
    TimeRemainingColumn,
    TransferSpeedColumn,
)

#Variables
prnumber = 0
panelroom = 0
equipnumber = 0
elv = 0
el = 0
imglist = []
tempsave=[]
timage=[]
prlist = []
vlist = []
olist = []
prolist = []
crashsave=[]
sdtype=0
rlsave=0
olsave=0
elsave=0
essave=0
prsave=0
prnsave=["",""]
equipname="test"
equipnumber="000-000000"
hz=0
hz1=0
updateloader=False
updateimageinjest=False
Iname1=""
Iname2=""
Iname3=""
Iname4=""
Iname5=""
Iname6=""
new_version=[]
#Classes
class eStops:
    def __init__(self):
        self.location = ""

    def gen_eStops(self):
        
        try:
            self.location = Prompt.ask("Location",default=str(self.location))
        except:
            return
class Rlockouts:
    def __init__(self):
        self.location = ""

    def gen_remoteLockouts(self):
        
        try:
            self.location = Prompt.ask("Location",default=str(self.location))  # Python 3
        except:
            return
class Olockouts:
    def __init__(self):
        self.Type = 0
        self.cType = 0
        self.sType = 0
        self.name = 0
        self.proced = 0
        self.sproced=0

    def gen_other(self):
        olheader=""
        console.print("Select Other Lockout Types", style="cyan",justify="center")
        for oroom in olist:
            olheader+=str(olist.index(oroom)) + " : " + oroom+"\n"
        console.print(Panel(olheader))
        Oltry = True
        while Oltry:
            try:
                self.Type = Prompt.ask("Enter 0-"+str(len(olist)-1),default=str(self.sType))
                self.sType=self.Type
                self.Type = int(self.Type)
            except:
                console.print("Use Only Numbers.", style="red",justify="center")
                continue
            if self.Type >= len(olist):
                console.print("Use Values Between 0 - "+len(olist)-1, style="red",justify="center")
                continue
            if self.Type < len(olist):
                self.Type = olist[self.Type]
                if self.Type == "Other":
                    self.Type = Prompt.ask("Enter Lockout Type",default=str(self.cType))
                    self.cType = self.Type
                
                Oltry = False
        
        self.name = Prompt.ask("\nLockout Name",default=str(self.name))
        console.print(Panel("Select Usage\n1 : 3A\n2 : 3B\n3 : 3C\n4 : 3D"))
        self.protry = True
        while self.protry:
            self.proced = Prompt.ask("Enter 1-4",default=str(self.sproced))
            self.sproced=self.proced
            self.proced=int(self.proced)
            if self.proced == 1 or self.proced == 2 or self.proced == 3 or self.proced == 4:
                if self.proced > 4:
                    continue
                if self.proced == 1:
                    self.protry = False
                    self.proced = "3A"
                if self.proced == 2:
                    self.protry = False
                    self.proced = "3B"
                if self.proced == 3:
                    self.protry = False
                    self.proced = "3C"
                if self.proced == 4:
                    self.protry = False
                    self.proced = "3D"
            else:
                console.print("Use Only Numbers 1-4.", style="red",justify="center")
                self.prrtry = True
class Elockouts:
    def __init__(self):
        self.volts = ""
        self.bucket = ""
        self.proced = 0
        #backup saves
        self.svolts=""
        self.sproced=""

    #def gen_elect(self,cvolts,svolts,sproced,sbucket):
    def gen_elect(self):
        console.print("\nSelect Lockout Voltage", style="cyan",justify="center")
        vheader=""
        for vroom in vlist:
            vheader+=str(vlist.index(vroom)) + " : " + vroom+"\n"
        console.print(Panel(vheader))
        Eltry = True
        while Eltry:
            try:
                self.volts = Prompt.ask("Enter 0-"+str(len(vlist)-1),default=str(self.svolts))
                self.svolts=self.volts
                self.volts = int(self.volts)
            except:
                console.print("Use Only Numbers.", style="red",justify="center")
                self.vols=""
                continue
            if self.volts >= len(vlist):
                console.print("Use Numbers between 0 - "+str(len(vlist)-1), style="red",justify="center")
            if self.volts < len(vlist):
                    self.volts = vlist[self.volts]
                    Eltry = False
                    
                    if self.volts == "Other":
                        self.volts = Prompt.ask("Enter Lockout Type",default=str(self.volts))
        self.bucket = Prompt.ask("Bucket Number",default=str(self.bucket))
        console.print("\nSelect Usage", style="cyan",justify="center")
        console.print(Panel("1 : 3A\n2 : 3B\n3 : 3C\n4 : 3D"))
        self.protry = True
        while self.protry:
            self.proced = Prompt.ask("Enter 1-4",default=str(self.sproced))
            self.sproced=self.proced
            self.proced=int(self.proced)
            if self.proced == 1 or self.proced == 2 or self.proced == 3 or self.proced == 4:
                if self.proced > 4:
                    continue
                if self.proced == 1:
                    self.protry = False
                    self.proced = "3A"
                if self.proced == 2:
                    self.protry = False
                    self.proced = "3B"
                if self.proced == 3:
                    self.protry = False
                    self.proced = "3C"
                if self.proced == 4:
                    self.protry = False
                    self.proced = "3D"
            else:
                console.print("Use Only Numbers 1-4.", style="red",justify="center")
                self.prrtry = True
#Functions
def select_PR(prnumber,prnsave):
    pr = 0
    prnumber = 0
    prheader= ""
    prsave=str(prnsave[1])
    prn2save=prnsave[0]
    console.print("Select Panel Rooms", style="cyan",justify="center")
    for proom in prlist:
        prheader+=str(prlist.index(proom)) + " : " + proom+"\n"
    console.print(Panel(prheader))
    
    invalid_input = True
    while invalid_input:
        try:
            pr = Prompt.ask("Enter 0-"+str(len(prlist)-1),default=str(prsave))
            prsave=pr
            pr = int(pr)
        except:
            console.print("Use Only Numbers.", style="red",justify="center")
            continue
        if pr >= len(prlist):
            console.print("Invalid Input Please try again.", style="red",justify="center")
        if pr < len(prlist):
            prn = " Panel Room"
            if pr == 0:
                
                panelroom = Prompt.ask("Enter Panel Room Name",default=str(prn2save))
                invalid_input = False
                return panelroom, prsave
            else:
                panelroom = prlist[pr] + prn
                invalid_input = False
                return panelroom, prsave
def elecLockouts(elsave):
    
    eltry = True
    while eltry:
        try:
            el = Prompt.ask("\nHow many Eletrical Lockouts (Enter 0-5)",default=str(elsave))
            el = int(el)
        except:
            console.print("Use Only Numbers!", style="red",justify="center")
            continue
        if el > 5:
            console.print("Invalid Input Please Enter A Value 0-5", style="red",justify="center")
        if el <= 5:
            eltry = False
            return el
def otherLockouts(olsave):
    
    oltry = True
    while oltry:
        try:
            ol = Prompt.ask("\nHow many Other Lockouts (Enter 0-5)",default=str(olsave))
            ol = int(ol)
        except:
            console.print("Use Only Numbers! other lockouts", style="red",justify="center")
            continue
        if ol > 5:
            console.print("Invalid Input Please Enter A Value 0-5", style="red",justify="center")
        if ol < 6:
            oltry = False
            return ol
def eStop_count(essave):
    estry = True
    while estry:
        try:
            es = Prompt.ask("\nHow Many eStops (Enter 0-2)",default=str(essave))
            es = int(es)
            print("")
        except:
            console.print("Use Only Numbers!", style="red",justify="center")
            continue
        if es > 2:
            console.print("Invalid Input Please Enter A Value 0-2", style="red",justify="center")
        if es < 3:
            estry = False
            return es
def remoteLockouts(rlsave):
    rltry = True
    while rltry:
        try:
            rl = Prompt.ask("\nHow Many Remote Lockouts (Enter 0-7:) ",default=str(rlsave))
            rl = int(rl)
            print("")
        except:
            console.print("Use Only Numbers!", style="red",justify="center")
            continue
        if rl > 7:
            console.print("Invalid Input Please Enter A Value 0-7", style="red",justify="center")
        if rl < 8:
            return rl
def shutdownprocedure(sdtype):
    Type = 0
    spheader=""
    console.print("Select Shutdown Procedure", style="cyan",justify="center")
    for proroom in prolist:
        spheader+=str(prolist.index(proroom)) + " : " + proroom+"\n"
    console.print(Panel(spheader))
    sdtry = True
    while sdtry:
        try:
            sdtype = Prompt.ask("Enter 0-"+str(len(prolist)-1),default=str(sdtype))
            sdtype = int(Type)
        except:
            console.print("Use Only Numbers.", style="red",justify="center")
            continue
        if sdtype >= len(prolist):
            console.print("Use values between 0 - "+str(len(prolist)-1), style="red",justify="center")
        if sdtype < len(prolist):
            sdtype = prolist[Type]
            
            if sdtype == "Other":
                sdtype = Prompt.ask("Input Shutdown Procedure",str(sdtype))
            return sdtype
def hazard(hz):
    global hz1
    hz1 = Prompt.ask("\nIs there a Hazardous energy permit? yes/no",default=str(hz1))
    if hz1 == '' or not hz1[0].lower() in ['y','n']:
        print('Please answer with yes or no!') 
    if hz1[0].lower() == 'y': #Do something 
        hz = Prompt.ask("What is the SOP Number",default=(str(hz)))
        return hz
    if hz1[0].lower() == 'n': #Do something 
        hz = None
        return
def check_exists(dir):
    isdir = os.path.isdir(dir) 
    return isdir
def resize_Image(corrected):
    try:
        file = corrected
        im = Image.open(file)
        if im.width != 318 and im.height != 228:
            console.print("Opened Image: "+file, style="green",justify="center")
            resized_im = im.resize((318,228))
            resized_im.save(file)
            console.print("Image Resized and Saved: "+file, style="green",justify="center")
        else:
            return
    except IOError as e:
        console.print("No Image", style="red",justify="center")
def generate_resources():
    try:
        with open('resources.txt') as f:
            fr="Found Resources"
            return fr
    except IOError:
        resourcedefault = ['<panelrooms>','Custom','Harvest/Trolly Floor','Cut Floor','Converting','Old Rendering','New Rendering','Scald Tub','Nippon','Cellars','CO2','Flow-Thru','Jeep Shop','Plasma','Waste Water One','Waste Water Two','Waste Water Three','Waste Water Four','Powerhouse One','Powerhouse Two','MQ Chill','Telephone','TCCS','</panelrooms>','<volts>','480v','220v','110v','24v','Other','</volts>','<otherlockouts>','Hydraulic','Water','Air','Steam','Other','</otherlockouts>','<ssProcedure>','USE STOP/START SWITCH','USE STOP/START BUTTONS','SHUT OFF PANEL SWITCH','SHUT OFF VALVE AT MOTOR (PRESSURE)','USE STOP/START BUTTON ON MAGELIS','Custom','</ssProcedure>','<template>','50 4B 03 04 14 00 06 00 08 00 00 00 21 00 64 31 2F CD A4 01 00 00 2E 07 00 00 13 00 08 02 5B 43 6F 6E 74 65 6E 74 5F 54 79 70 65 73 5D 2E 78 6D 6C 20 A2 04 02 28 A0 00 02 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 C4 55 CB 4E EB 30 10 DD 5F 89 7F 88 BC 45 8D 0B 0B 74 85 9A 76 C1 63 09 48 C0 07 18 7B 9A 58 75 6C CB 33 40 FB F7 77 6C 4A 75 85 4A 4B D4 4A 6C 92 F8 31 E7 9C 39 CE 8C 27 B3 65 EF AA 37 48 68 83 6F C4 59 3D 16 15 78 1D 8C F5 6D 23 9E 9F 6E 47 7F 45 85 A4 BC 51 2E 78 68 C4 0A 50 CC A6 27 7F 26 4F AB 08 58 71 B4 C7 46 74 44 F1 52 4A D4 1D F4 0A EB 10 C1 F3 CA 3C A4 5E 11 0F 53 2B A3 D2 0B D5 82 3C 1F 8F 2F A4 0E 9E C0 D3 88 32 86 98 4E AE 61 AE 5E 1D 55 37 4B 9E FE 50 F2 62 BD A8 AE 3E F6 65 AA 46 A8 18 9D D5 8A 58 A8 7C F3 E6 0B C9 28 CC E7 56 83 09 FA B5 67 E8 1A 63 02 65 B0 03 A0 DE D5 31 59 66 4C 8F 40 C4 89 A1 90 5B 39 A3 6F BF 70 DA 3E 6B CE F3 DB 23 12 38 1C 26 73 ED 43 CD 91 25 15 EC 6C C4 53 36 EB 1B 86 BC F2 BD 0F EB B8 7B 3E C0 64 0D 54 0F 2A D1 9D EA D9 2D B9 74 F2 3D A4 C5 4B 08 8B 7A 37 C8 50 33 8B A9 75 AF AC FF D4 BD 83 BF 6C 46 59 5E 67 47 16 92 F3 2B C0 03 75 9C FF 92 0E E2 EA 00 59 9E 87 5B 51 60 F6 24 8E B4 72 80 C7 3E FE 02 BA 8F B9 53 09 CC 23 71 DD B5 47 17 F0 3F F6 1E 1D 26 A9 F7 2C 41 AE 3F 0E F7 7D 0D 34 90 F7 F0 5F EE 67 BC 5A 39 7D D5 71 69 1E F9 D0 37 B8 BB F2 E6 E6 FB 90 42 44 EE EF 09 86 0B F8 6C 8D 39 7A 14 19 08 12 59 D8 34 C7 6D 4D 66 C3 C8 97 C3 C1 19 43 BE 7D 0C 98 2D DC B2 DC 76 D3 7F 00 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 B5 55 30 23 F4 00 00 00 4C 02 00 00 0B 00 08 02 5F 72 65 6C 73 2F 2E 72 65 6C 73 20 A2 04 02 28 A0 00 02 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 AC 92 4D 4F C3 30 0C 86 EF 48 FC 87 C8 F7 D5 DD 90 10 42 4B 77 41 48 BB 21 54 7E 80 49 DC 0F B5 8D A3 24 1B DD BF 27 1C 10 54 1A 83 03 47 7F BD 7E FC CA DB DD 3C 8D EA C8 21 F6 E2 34 AC 8B 12 14 3B 23 B6 77 AD 86 97 FA 71 75 07 2A 26 72 96 46 71 AC E1 C4 11 76 D5 F5 D5 F6 99 47 4A 79 28 76 BD 8F 2A AB B8 A8 A1 4B C9 DF 23 46 D3 F1 44 B1 10 CF 2E 57 1A 09 13 A5 1C 86 16 3D 99 81 5A C6 4D 59 DE 62 F8 AE 01 D5 42 53 ED AD 86 B0 B7 37 A0 EA 93 CF 9B 7F D7 96 A6 E9 0D 3F 88 39 4C EC D2 99 15 C8 73 62 67 D9 AE 7C C8 6C 21 F5 F9 1A 55 53 68 39 69 B0 62 9E 72 3A 22 79 5F 64 6C C0 F3 44 9B BF 13 FD 7C 2D 4E 9C C8 52 22 34 12 F8 32 CF 47 C7 25 A0 F5 7F 5A B4 34 F1 CB 9D 79 C4 37 09 C3 AB C8 F0 C9 82 8B 1F A8 DE 01 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 85 90 AC 37 8C 03 00 00 AA 08 00 00 0F 00 00 00 78 6C 2F 77 6F 72 6B 62 6F 6F 6B 2E 78 6D 6C AC 55 61 6F 9B 3C 10 FE FE 4A FB 0F 88 EF 14 9B 00 09 A8 C9 14 48 D0 5B A9 9D AA 2E 6B BF 44 AA 5C 70 8A 15 C0 CC 36 4D A2 6A FF 7D 67 08 69 DA 4E 53 D6 2D 4A EC D8 3E 3F 3C 77 F7 DC 71 FE 79 5B 16 C6 13 15 92 F1 6A 6C E2 33 64 1A B4 4A 79 C6 AA C7 B1 F9 6D 91 58 23 D3 90 8A 54 19 29 78 45 C7 E6 8E 4A F3 F3 E4 D3 7F E7 1B 2E D6 0F 9C AF 0D 00 A8 E4 D8 CC 95 AA 43 DB 96 69 4E 4B 22 CF 78 4D 2B 38 59 71 51 12 05 4B F1 68 CB 5A 50 92 C9 9C 52 55 16 B6 83 90 6F 97 84 55 66 87 10 8A 53 30 F8 6A C5 52 3A E3 69 53 D2 4A 75 20 82 16 44 01 7D 99 B3 5A F6 68 65 7A 0A 5C 49 C4 BA A9 AD 94 97 35 40 3C B0 82 A9 5D 0B 6A 1A 65 1A 5E 3C 56 5C 90 87 02 DC DE 62 CF D8 0A F8 FA F0 C3 08 06 A7 7F 12 1C BD 7B 54 C9 52 C1 25 5F A9 33 80 B6 3B D2 EF FC C7 C8 C6 F8 55 08 B6 EF 63 70 1A 92 6B 0B FA C4 74 0E 0F AC 84 FF 41 56 FE 01 CB 7F 01 C3 E8 AF D1 30 48 AB D5 4A 08 C1 FB 20 9A 77 E0 E6 98 93 F3 15 2B E8 6D 27 5D 83 D4 F5 17 52 EA 4C 15 A6 51 10 A9 E6 19 53 34 1B 9B 43 58 F2 0D 7D B5 21 9A 3A 6A 58 01 A7 CE 20 70 02 D3 9E 1C E4 7C 2D 8C A6 CE 88 A2 97 AC 5A 83 AC 49 B1 21 3B 50 55 46 57 A4 29 D4 02 D4 DD 3F 13 CA C5 F7 03 C7 D3 D7 41 2D D3 42 51 51 C1 CD 98 57 0A C4 B9 77 F6 6F 85 D8 62 C7 39 07 D9 1B 37 F4 7B C3 04 05 5A 20 3A 08 00 8C 24 0D C9 83 BC 26 2A 37 1A 51 8C CD 38 5C 7E 93 10 93 65 24 76 A4 5A 4A DE 88 94 2E 05 AD B9 5C C6 44 64 06 B0 83 72 57 CB 23 F9 92 F7 B5 F2 07 02 26 A9 0E 80 0D 11 E8 58 76 FF DF 46 03 C8 8A B0 17 E9 B5 12 06 FC BF 98 5D 42 A2 BE 92 27 48 1B 88 23 DB 57 F5 05 E4 05 0F EE AB 54 84 F8 FE 19 0D FD C0 9B F9 AE 35 F2 FC 81 E5 8E A6 AE 15 39 31 B2 06 83 78 EA 0F 63 27 08 92 F9 0F 70 46 F8 61 CA 49 A3 F2 BD 22 34 F4 D8 74 B5 86 DF 1E 5D 91 6D 7F 82 51 D8 B0 EC 85 C6 33 DA 7F 2C 3D BF 19 FA B3 1F DA 61 DD FB 6E 19 DD C8 17 ED E8 A5 B1 BD 63 55 C6 37 63 D3 C2 0E 38 B5 7B BD DC B4 87 77 2C 53 39 88 2F 40 2E 98 74 7B FF 53 F6 98 03 63 EC 39 7A 13 6A 44 33 1B 9B CF DE 60 98 A0 91 97 58 B3 79 34 B7 DC A9 0B 3D D9 73 1D 6B 16 4F 03 7F E4 0F 51 E4 B9 2D 23 FB 88 52 DB 65 81 5A 3B 1B 55 5B 19 53 E8 E4 BA F9 B6 F1 35 0D 11 6A 78 71 91 61 ED CE B1 61 74 64 08 6D EE 60 E8 B4 89 EE A1 53 52 A4 50 2D 7A 6A 11 03 8C BA 52 A2 5B 75 29 D5 E4 1C 66 D0 24 03 17 B0 8B A6 43 14 B8 16 9A 0F 3C C8 61 E0 58 23 77 E0 58 B1 3B 73 E6 DE 70 0E 9E 79 3A 87 FA 4D 12 FE 8B 7E DA 96 46 D8 BF A2 34 CB 9C 08 B5 10 24 5D C3 8B ED 86 AE 22 22 41 74 AD E7 36 F0 3C 26 1B 79 A3 08 41 5B B0 DC 04 27 96 8B 03 64 45 11 E8 CF 9B 25 03 6F 88 67 F1 DC 4B 5E C8 6A F7 57 1F EC 66 23 BB BD 4D 89 6A A0 A8 75 3D B7 EB 50 8F C9 7E F7 B0 B9 EA 36 F6 B9 7C 55 9F E1 CD 4C 67 66 7F FB 77 86 5F C1 FB 82 9E 68 9C DC 9E 68 18 7F B9 5A 5C 9D 68 7B 39 5F DC DF 25 AD 90 7E E9 6D 97 0D 3D B6 1A B2 FB 1C 4E 7E 02 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 FE 69 EA 57 0A 01 00 00 CC 03 00 00 1A 00 08 01 78 6C 2F 5F 72 65 6C 73 2F 77 6F 72 6B 62 6F 6F 6B 2E 78 6D 6C 2E 72 65 6C 73 20 A2 04 01 28 A0 00 01 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 BC 93 4F 4B C4 30 10 C5 EF 82 DF 21 CC DD A6 AD BA 88 6C BA 07 45 D8 AB AE 1F 20 A4 D3 A6 6C 9B 94 CC F8 A7 DF DE 50 B1 75 61 A9 97 E2 71 DE 90 F7 7E 3C 32 DB DD 67 D7 8A 77 0C D4 78 A7 20 4B 52 10 E8 8C 2F 1B 57 2B 78 3D 3C 5D DD 81 20 D6 AE D4 AD 77 A8 60 40 82 5D 71 79 B1 7D C6 56 73 7C 44 B6 E9 49 44 17 47 0A 2C 73 7F 2F 25 19 8B 9D A6 C4 F7 E8 E2 A6 F2 A1 D3 1C C7 50 CB 5E 9B A3 AE 51 E6 69 BA 91 E1 B7 07 14 27 9E 62 5F 2A 08 FB F2 1A C4 61 E8 63 F2 DF DE BE AA 1A 83 8F DE BC 75 E8 F8 4C 84 E4 C8 85 D1 50 87 1A 59 C1 38 7E 8B 59 12 41 41 9E 67 C8 D7 64 F8 F0 E1 48 16 91 67 8E 49 22 39 6E F2 25 98 EC 9F 61 16 9B D9 AC 09 63 74 6B 1E AC 6E DC DC CC 24 2D 35 72 BB 26 04 59 1D B0 7C E1 10 2F 80 66 90 13 79 09 E6 66 55 18 1E DA 78 70 D3 87 A5 71 FE 89 97 27 37 58 7C 01 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 2D CC 2D 95 A9 1B 00 00 7D 97 00 00 18 00 00 00 78 6C 2F 77 6F 72 6B 73 68 65 65 74 73 2F 73 68 65 65 74 31 2E 78 6D 6C B4 DD 5D 73 DB 46 9A 05 E0 FB AD DA FF A0 D5 DE 4C AA 26 96 48 51 B2 CC B2 3D 65 09 FC 26 41 5A 22 08 92 77 8C 4C DB AA 48 A2 57 A2 93 CC 6E ED 7F DF B7 89 7E BB DF E6 39 C9 D8 89 B6 6A 32 B6 1F 74 37 3E 78 D8 00 1A 1F 7C FD 8F DF EE EF 0E 7E 59 3F 3E DD 6E 1E DE 1C D6 5E 1C 1F 1E AC 1F 6E 36 1F 6E 1F 3E BD 39 2C A6 ED 1F CF 0F 0F 9E B6 AB 87 0F AB BB CD C3 FA CD E1 3F D7 4F 87 FF 78 FB EF FF F6 FA D7 CD E3 CF 4F 9F D7 EB ED 81 B4 F0 F0 F4 E6 F0 F3 76 FB A5 79 74 F4 74 F3 79 7D BF 7A 7A B1 F9 B2 7E 90 29 1F 37 8F F7 AB AD FC F3 F1 D3 D1 D3 97 C7 F5 EA C3 AE D2 FD DD 51 FD F8 F8 EC E8 7E 75 FB 70 58 B5 D0 7C FC 96 36 36 1F 3F DE DE AC B3 CD CD D7 FB F5 C3 B6 6A E4 71 7D B7 DA CA F2 3F 7D BE FD F2 A4 AD DD DF 7C 4B 73 F7 AB C7 9F BF 7E F9 F1 66 73 FF 45 9A F8 E9 F6 EE 76 FB CF 5D A3 87 07 F7 37 CD DE A7 87 CD E3 EA A7 3B 59 EF DF 6A 8D D5 CD C1 6F 8F F2 BF BA FC 77 A2 B3 D9 39 CC E9 FE F6 E6 71 F3 B4 F9 B8 7D 21 2D 1F 55 CB 8C AB FF EA E8 D5 D1 EA 26 B4 84 EB FF 4D CD D4 1A 47 8F EB 5F 6E DD 07 18 9B AA FF B9 45 AA 9D 86 B6 EA B1 B1 93 3F D9 D8 59 68 CC 6D AE C7 E6 D7 DB 0F 6F 0E FF E7 F2 E2 D5 59 D6 3E 3F FF F1 F8 B8 D5 F8 B1 D1 7A 77 FC E3 F9 E9 F9 AB 1F 6B C7 8D FA E5 F1 79 BB 7E F1 AA FE BF 87 6F 5F EF 72 32 79 7C FB FA CB EA D3 FA 7A BD 2D BE 4C 1E 0F 3E DE 6E A7 9B 89 80 64 F5 F0 E8 ED EB A3 50 EA C3 AD 04 C2 6D 84 83 C7 F5 C7 37 87 EF 6A CD E5 E9 B9 2B B2 2B 31 BB 5D FF FA 64 FE 7E B0 5D FD 74 BD BE 5B DF 6C D7 B2 4C B5 C3 83 ED E6 CB 70 FD 71 7B B9 BE BB 93 CA 27 22 FF BD D9 DC 5F DF AC EE D6 B9 8B B0 68 ED 58 BE 1C 2E F6 3F 6D 36 3F BB 06 7B 52 F5 D8 2D E9 AE 21 37 EB D5 CD F6 F6 97 75 D5 C8 55 5D 36 E1 D3 7F ED 96 46 FE DE 5C C8 BF C3 12 BB EA BA F4 76 D9 DA BB AF 8B AC E8 87 F5 C7 D5 D7 BB ED D5 E6 D7 EE FA F6 D3 E7 AD CC FE 54 36 A2 4B 61 F3 C3 3F B3 F5 D3 8D C4 5F 66 FF A2 7E EA 5A BD D9 DC C9 EA C9 FF 1F DC DF CA F7 58 BE B4 F7 AB DF 76 7F FE 7A FB 61 FB F9 CD E1 AB 17 B5 C6 F1 99 14 3E B8 F9 FA B4 DD DC 97 15 EF 36 62 A8 56 3B F3 F5 DC 5F 7C C5 FA 1F 54 3C AA 66 BB DB C2 D9 6A BB 7A FB FA 71 F3 EB 81 E4 58 36 DF D3 97 95 EB 15 6A CD BA B4 A5 CB 5F CD 3B AC D1 EF AD 90 AC 89 6B E5 9D 6B 46 16 BD 5A 43 27 17 5E CE 76 EB EC E4 12 24 03 69 55 72 2E 5B 45 36 D9 93 7C 62 BF BC 3D 7E 7D F4 8B 7C 00 37 7E 56 ED 50 C4 6D 4C D7 70 07 A4 0B D2 03 E9 83 0C 40 86 20 23 90 1C 64 0C 32 A9 44 BE 37 61 6B BC AF E8 4C 92 17 D6 F4 24 5D D3 AB 50 44 D7 F4 1A 64 0A 52 80 CC 40 4A 90 39 C8 02 64 E9 E5 64 F7 E5 90 FC 84 10 B9 AF CF 37 85 68 FB F9 F6 E6 E7 8B 8D FB 8E D0 44 9D C8 27 5E 05 CA 35 29 81 7A 19 36 D8 85 97 5D 4F B1 2B 73 09 92 81 B4 2A 39 7F 15 DA 69 83 74 40 BA 20 3D 90 3E C8 00 64 08 32 02 C9 41 C6 20 93 4A 6A C7 F1 FB F5 BE A2 B3 98 A8 2B 90 6B 90 29 48 01 32 03 29 41 E6 20 0B 90 A5 97 DD 32 1F D9 B8 C8 3E E6 39 FA 1C D7 4C 1A 11 2F 26 22 20 19 48 AB 12 1B 11 90 0E 48 17 A4 07 D2 07 19 80 0C 41 46 20 39 C8 18 64 52 49 12 91 8A 4E E4 73 88 5D CC CB BD 3E A6 2A E3 77 4D AE 37 BD 06 99 82 14 20 33 90 12 64 0E B2 00 59 7A 81 2E A6 F1 3C 99 71 CD A4 99 F1 62 32 03 92 81 B4 2A 91 CC C4 4D 5B DB DB 4F 85 22 61 3F 05 D2 05 E9 81 F4 41 06 20 43 90 11 48 0E 32 06 99 54 92 84 A8 A2 BA 1C 50 C5 10 9D EF 85 C8 97 D9 ED 18 76 5D F3 35 C8 14 A4 00 99 81 94 20 73 90 05 C8 B2 12 EC 78 E4 98 EA 39 3A 1E D7 4C 1A 22 2F 26 44 20 19 48 AB 12 DB F1 80 74 40 BA 20 3D 90 3E C8 00 64 08 32 02 C9 41 C6 20 93 4A 92 CC 54 24 87 93 1A FD 2B 5F E8 55 75 14 BB EB 67 90 A6 48 05 D2 0C A9 44 9A 23 2D 90 96 15 61 52 E4 48 F8 39 92 E2 9A 49 93 E2 C5 24 05 24 03 69 55 62 93 02 D2 01 E9 82 F4 40 FA 20 03 90 21 C8 08 24 07 19 83 4C 2A 49 92 52 51 DA BB BC DA EB 5D 7C 99 78 28 78 0D 32 05 29 40 66 20 25 C8 1C 64 01 B2 AC 04 33 F3 F2 F9 8F 82 5D 93 92 9F 78 F4 7A 51 89 3B C1 D5 EF D6 25 52 86 D4 52 92 03 EF D0 91 D7 F7 76 59 5A 66 77 FA 5B 9D 5B 21 75 91 7A 48 7D A4 01 D2 10 69 84 94 23 8D 95 E2 3E 67 A2 64 0E 92 2B B2 1D 91 2F F4 D2 EC AB 90 A6 48 05 D2 0C A9 44 9A 23 2D 90 96 15 61 A8 DC A0 1E 9E 5A BD 78 29 1D 17 1F 62 D0 13 28 57 51 A2 63 F7 DB B5 BD FD F6 85 2F 13 3B E5 4B 90 0C A4 05 D2 06 E9 80 74 2B 69 C4 28 F7 42 19 73 0C B5 77 78 DA 87 76 06 20 43 90 11 48 0E 32 06 99 54 72 1E BF 59 EF 2B A9 C9 1F F1 D8 A7 B1 D7 3B F9 32 71 BD AE 41 A6 20 05 C8 0C A4 04 99 83 2C 40 96 95 60 90 E4 50 15 83 74 F6 42 72 F4 ED 27 E5 AE 0D C9 54 EC 1E 2E BC C4 6F D3 25 48 06 D2 02 69 83 74 40 BA 95 9C C6 CF A7 E7 CB C4 73 E0 3E C8 00 64 08 32 02 C9 41 C6 20 93 4A 6C 5E 2A A9 CB 1F 31 2F 7B C3 57 57 BE 4C 3C 00 B8 AE A4 16 13 34 05 29 40 66 20 25 C8 1C 64 01 B2 AC 04 F3 52 93 9E E3 CF F5 3C BB 9A 32 C2 63 77 5B DE E4 4A 82 19 0F 74 73 78 73 68 2D 23 E5 5A C4 DA C4 3A C1 62 7F D6 F5 66 43 A3 8B 27 27 09 E1 53 AA ED ED 09 FB 5A 28 2E EE 40 E9 54 F6 FF B1 E2 5E 77 30 34 A5 74 07 3D 22 96 13 1B 07 93 9D 41 9C C3 59 DA E1 4C C2 A2 99 FD 9C AE 7A 92 BC BD 13 D2 2B 2D 64 A2 A7 6D 99 EC 21 15 48 33 A4 12 69 8E B4 40 5A 7A 22 11 74 03 8E B0 F3 7B E5 FA 2C 18 9E FE F6 4E AC 56 0D 63 D6 8F 4D 37 16 CC 74 64 C4 32 62 2D 62 6D 62 9D 60 B1 BB EA 7A 4B E2 E9 C7 8A 65 20 24 66 60 6F 64 B8 EF EB C5 8F 6D E0 A5 76 6A 23 50 3B 4D B3 33 34 A5 62 3A FD 0C A5 A6 5A 4E CA 8D 83 C5 FC 4C D4 EC C0 A4 B7 86 FD 9A 9C EC 7D BF AE 74 63 D8 2C FA E5 B0 59 04 2A 74 8E B1 D4 0C A9 44 9A 23 2D 90 96 FA 81 C0 F8 53 CD 8D 66 3E C7 85 12 3F 94 FB 47 BD CF C5 6E 66 EE 32 96 E9 2C 7D BD E4 82 41 6D AF F7 C9 B4 A2 94 D2 CF B2 45 AC 4D AC E3 AD EE 36 C1 EF F6 3E 5D 5F 2A 49 AC 5F 34 99 A4 33 ED EB 0C 6A F1 0B 35 08 16 57 6B 18 2C 8E 15 8C 88 E5 A4 BD 31 69 6F 12 B6 9C ED 1C AB E5 3B 49 0E 85 F7 8F 85 AF 74 F5 6D 22 FD 8A D9 44 02 15 3A 4B 9B 48 28 55 62 A9 39 D2 02 69 A9 1B 1C 13 E9 86 4A 9F 23 91 7E E4 F8 8F 3A 9C 0B F9 1C 77 87 7E 26 90 BE DA 59 52 6F AF C3 C9 7C BD 9A B9 4E D1 22 D6 0E 16 3F B7 8E B7 BA C9 55 D7 5B 12 3F BF 20 35 73 18 A8 CD C9 95 58 8D E4 20 58 3C 3C 18 06 8B 83 0B 23 62 39 69 6F 4C DA 9B A8 25 FD A1 1F D2 B6 1D 73 7D 7F 04 55 D7 D5 A6 CF AF 97 4D 1F 50 A1 73 B4 E9 83 52 25 96 9A 23 2D 90 96 BA BD 31 7D 6E 8C F5 77 4E 4C BF F7 DA B1 1F 75 36 9F F3 85 7C 98 2E 6F 35 99 7F 18 E7 08 16 FB 8F 2C 58 EC 3F 5A C4 DA A4 BD 8E DA 59 6C AF EB 2D C9 17 2E 5E 9F 34 37 20 8B 37 24 8B 32 22 96 93 F6 C6 A4 BD 89 5A 92 2F 3F FE 9D E4 6B 7F 0C CD 57 AC DB 7C F9 F5 B2 F9 02 2A 74 8E 36 5F 50 AA C4 52 73 A4 05 D2 52 B7 37 E6 CB 8D CC 3E 53 BE FC B8 AF CC 4B B3 74 E1 6E D1 D8 E5 2B 76 10 97 C1 62 07 91 05 8B 1D 44 8B 58 9B B4 D7 51 3B B3 E7 23 7E C0 D9 9C C4 86 AA B6 FF C2 C5 1B 90 C5 1B 92 45 19 11 CB C9 E2 8D 49 7B 13 B5 24 5F D5 B2 24 67 B5 B5 1A 04 CC 97 B2 01 D3 F1 F6 B0 D9 A7 DA 7E 4C 53 81 34 43 2A 91 E6 48 0B A4 A5 27 72 72 E1 86 71 9F 29 60 7E 60 3B E9 C0 D4 6C 07 A6 66 3B 30 35 DB 81 A1 B5 DD 7D 3E 7B 1D 62 47 2D E9 C0 FC E8 B4 0D 18 2E 5E 9F 34 37 08 66 8F CF 70 51 46 A1 5C 5C E4 9C B4 37 26 ED 4D D4 92 80 F9 21 F6 A4 03 DB 3B EB BD D2 8A 66 9C 0D 69 8A 54 20 CD 90 4A A4 39 D2 02 69 E9 89 E4 CB 8D E8 3E 53 BE FC 78 71 D2 81 A9 D9 0E 4C CD 76 60 6A B6 03 43 6B D7 B0 BD 8E 5A D2 81 F9 81 6A 9B 2F 5C BC 3E 69 6E 10 CC 1E 80 E1 A2 8C 42 B9 B8 C8 39 69 6F 4C DA 9B A8 25 F9 AA E6 91 74 60 F5 FD FB 20 B4 A2 CD 97 5F 36 BB 83 04 2A B0 E2 0C A9 44 9A 23 2D 90 96 9E 48 BE DC 40 EF 33 E5 CB 8F 22 27 FD 97 9A ED BF BC C9 C5 07 DD 91 66 6E 70 DC F5 4B 2F 6D FF 85 D6 F6 E5 EA C7 71 07 D1 D1 BA 49 FF E5 C7 AF 6D BE 70 F1 FA 5A D5 1C 1F 0E 82 D9 FE 4B EB DA F3 4B B4 9C B4 37 26 ED 4D D4 92 7C 55 ED BD B4 F7 D9 D4 EA FB E3 BE BE A6 BD F0 84 34 45 2A 90 66 48 25 D2 1C 69 81 B4 F4 44 02 E6 46 86 9F 29 60 7E D8 39 E9 C0 D4 6C 07 E6 CD 8C A9 67 32 2E BE 0B D8 79 EC 35 5A C4 DA DE EA C7 F1 EB DA D1 72 49 07 E6 07 BC 6D C0 70 F1 FA 5A 55 8E 1F E2 19 A4 96 B3 1D 98 9A 3D 83 44 CB 49 7B E3 60 B1 BD 89 5A 12 B0 AA BD EA AE E4 DD 65 D9 2B 2D 66 AE C5 5C 13 9B 12 2B 88 CD 88 95 C4 E6 C4 16 C4 96 DE 30 55 EE 52 FF 33 A5 6A D7 94 3B 47 34 03 B9 C1 4C B7 A5 66 BB AD 60 A6 DB 22 D6 F6 96 74 5B 5A CE 76 5B DE EC 79 23 59 BC 3E 59 BC 41 30 D3 6D 05 33 DD 16 B1 9C B4 37 26 ED 4D D4 92 1B 48 FD E5 17 7B 07 69 45 35 9B 2A AD 6A 6C 4A AC 20 36 23 56 12 9B 13 5B 10 5B EA 56 86 B3 C5 3A BD 52 70 BA BB 4C FE 9D A3 11 BB A6 5C AA CC D9 62 30 D3 57 A9 D9 BE 2A 98 E9 AB 88 B5 BD 25 7D 95 96 B3 7D 95 B7 24 55 7E 48 DC 2C 5E 9F 2C DE 20 98 E9 AB 82 99 BE 8A 58 4E DA 1B 93 F6 26 6A 49 AA FC D5 15 9B 2A 1D C5 8F DF D3 6B AD 9A A4 0A CB 15 A4 DC 8C 58 49 6C 4E 6C 41 6C A9 5B 19 53 C5 C7 FC FF 54 AA 70 8C FC 62 F7 D4 C9 DE 18 97 5A D2 57 F9 BA F6 10 2B 94 8B 1D 44 DB 5B DA 57 E9 65 03 33 C6 E5 CB 25 A9 22 43 F8 64 F1 06 C1 6C 5F A5 75 6D 5F 85 96 93 F6 C6 A4 BD 89 5A 92 AA AA BD BA 4D 95 8E B1 DB 54 A1 4D B5 39 93 B4 82 D8 8C 58 49 6C 4E CC 3D 2F B4 3B 46 31 F3 58 EA 56 C6 54 F1 71 7B 9E AA EF 78 60 42 87 C1 E3 01 F6 45 5D 2D 1E 13 5D AA BD 8C E5 32 6F AF 22 B5 90 DA 9E EA E6 B2 54 47 ED 38 76 29 5D 6F A7 71 17 DC 0B 0B 12 E7 D0 27 0B 37 50 93 8E 5E 8F BA 86 C1 62 4F 36 22 96 93 F6 C6 A4 BD 89 5A 92 2F 3F 46 6F F3 55 91 3D 5E F7 35 0D 4D 91 0A A4 19 52 89 34 47 5A 20 2D 75 EB 62 AE 9E 6F 44 5E BE 67 D5 31 B7 BD B6 53 DF BB B6 73 11 4A C5 AD 76 49 2C 23 D6 22 D6 26 D6 21 D6 0D 16 F7 C6 3D 52 AE 4F 6C 40 6C 48 6C 44 2C 27 36 26 36 F1 66 6F 0C F2 E4 8E 5F C2 A5 D5 FD 1B C9 7C 99 53 73 8B 34 D2 14 A9 40 9A 21 95 48 73 A4 05 D2 52 09 03 F7 3D 43 F4 DF D1 91 F9 61 E3 F3 B8 3F B9 70 4F 43 BA 51 86 73 7B 00 86 96 91 72 2D 62 6D 62 1D 62 5D 62 3D 62 7D 62 03 62 43 62 23 62 39 B1 31 B1 89 B7 24 6A 7E 20 3E 79 60 63 EF AB 7B E5 EB 25 59 F3 57 26 62 FC A6 58 AA 40 9A E9 72 99 9D 5F 49 6C 4E 6C 41 6C A9 B3 C0 BC FD C5 11 7B F3 70 A1 1F D5 3E 4F 6E B4 AA ED 8D 09 5E B8 E7 61 77 A9 33 77 EB 13 CB 88 B5 88 B5 89 75 88 75 89 F5 88 F5 89 0D 88 0D 89 8D 88 E5 C4 C6 6A AF 62 27 3F F1 96 A4 CE 8F E1 DB 1D A8 DF 7A 27 E6 92 90 B6 66 6C 4A AC 20 36 23 56 12 9B 13 5B 10 5B 7A 23 43 14 7F 71 E4 DE E4 CC 0F 17 9B 07 37 2E EA 3A 84 6C 6E EE 21 96 11 6B 11 6B 13 EB 10 EB 12 EB 11 EB 13 1B 10 1B 12 1B 11 CB 89 8D 83 C5 EE 7C E2 2D 09 95 1F B8 8F 5B EA 2A D4 34 F7 EF 13 9B 12 2B 88 CD 88 95 C4 E6 C4 16 C4 96 DE 48 A8 FE E2 70 BD 09 95 1F BE 4E 42 A5 F7 81 DB 50 A1 65 72 FD BF 3A 51 89 E5 5A C4 DA C4 3A C4 BA C4 7A C4 FA C4 06 C4 86 C4 46 C4 72 62 E3 60 36 54 D5 FA 26 A1 AA C8 9C 61 5C 85 9A 36 54 BA A5 A2 4D 49 B9 82 D8 8C 58 49 6C 4E 6C 41 6C E9 8D 84 EA 2F 0E D1 9B 50 E9 28 7B 3C 53 BC 70 17 CC DC EE 4F 9E 21 09 F7 E0 10 CB 88 B5 88 B5 89 75 88 75 89 F5 88 F5 89 0D 88 0D 89 8D 88 E5 C4 C6 C1 6C A8 AA ED 92 84 CA 8F D0 C7 83 A7 2B 5F F3 3C EE 11 AF 91 A6 48 05 D2 0C A9 44 9A 23 2D 90 96 9E 30 4C EE FE C8 BF 32 32 1F C3 B4 6B C9 1D AD 9B 30 A9 D9 30 11 CB 88 B5 88 B5 89 75 88 75 89 F5 88 F5 89 0D 88 0D 89 8D 88 E5 C4 C6 C1 4C 98 BC D9 30 79 B2 3D 94 96 32 61 42 9A 22 15 48 33 A4 12 69 8E B4 40 5A 7A 22 61 7A BE 01 79 F7 86 9F DD 51 B8 4D 93 0E 15 9B AE 49 CB 99 84 65 C4 5A C4 DA C4 3A C4 BA C4 7A C4 FA C4 06 C4 86 C4 46 C4 72 62 E3 60 36 4D D5 76 49 D2 A4 8F 3B D8 87 92 F6 EF AE 09 6D 85 9E FE 1A 69 8A 54 20 CD 90 4A A4 39 D2 02 69 E9 89 E4 EB F9 86 E6 DD FB 31 20 5F 3A A0 6B F3 85 96 69 5D 93 B9 16 B1 36 B1 0E B1 2E B1 1E B1 3E B1 01 B1 21 B1 11 B1 9C D8 38 98 CD 57 B5 0D 92 7C 55 E4 3E A9 F8 D0 DB FE 9B 8C 7C 5B 72 20 13 5E 65 14 9A 8F 77 07 7A 3A 89 33 2C B0 D4 0C A9 44 9A 23 2D 90 96 9E 48 BE FE 7F 06 E9 4F FC 80 7C B2 67 D4 DB B5 6D D6 D0 32 AD 9B 64 0D CB B5 49 B9 0E B1 2E B1 1E B1 3E B1 01 B1 21 B1 11 B1 9C D8 38 98 CD 5A B5 6E 49 D6 2A 4A B2 56 AB EF 3F E7 E6 1B 3B B5 3B CB AA A2 A1 29 96 2A 90 66 48 25 D2 1C 69 A1 14 AF 4E 2D 3D ED AE 94 A4 EF 43 E2 23 F7 E4 29 B7 3F 7E E4 FB 44 87 ED ED 9E 52 6F D6 B6 E9 42 CB B4 6E 92 2E 2C D7 26 E5 3A C4 BA C4 7A C4 FA C4 06 C4 86 C4 46 C4 72 62 E3 60 36 5D D5 BA 25 E9 AA 48 2E 8C 99 E7 9F 20 5C A1 50 EC C9 80 A6 7E 8E E6 1A 5B 81 34 43 2A 91 E6 48 0B A4 A5 D2 EE A3 4F D3 F5 7C 77 D2 4B CF 8C BB 4A BD CB DB 06 0C 2D D3 BA 49 C0 B0 5C 9B 94 EB 10 EB 12 EB 11 EB 13 1B 10 1B 12 1B 11 CB 89 8D 83 D9 80 55 EB 96 04 0C E8 CA D7 34 A5 AE 91 A6 48 05 D2 0C A9 44 9A 23 2D 90 96 4A BB 13 DA 34 4E DF 33 0A FF ED 57 7D 4E 74 AC DD F6 5D DE 92 73 46 B4 4C EB 26 D1 C2 72 6D 52 AE 43 AC 4B AC 47 AC 4F 6C 40 6C 48 6C 44 2C 27 36 0E 66 A3 55 AD 5B 12 2D 4F 71 77 73 E5 6B 9E 9B FB 23 90 A6 48 05 D2 0C A9 44 9A 23 2D 90 96 4A F8 1A C9 13 3E F0 FE FD FB 41 1D 62 37 57 10 77 6D BB C1 AC 78 A3 C2 25 B1 8C 58 8B 58 9B 58 87 58 97 58 8F 58 9F D8 80 D8 90 D8 88 58 4E 6C 1C CC 3C C1 ED 2D C9 52 B5 F9 CE 64 4C 2B 5E AC DE BF 9D D9 D7 7B 19 7B FD 6B A4 29 52 81 34 43 2A 91 E6 48 0B A4 A5 12 E9 B8 F8 08 FC F7 A7 0B 6F 47 BF 90 EB 59 D5 D3 39 E6 75 B7 C4 32 62 AD 60 71 A4 B9 AD 26 9F 81 1E 66 74 D4 CC 26 EF 12 EB 11 EB 07 8B 03 94 83 60 B1 67 19 06 33 F7 72 11 CB 89 8D 83 C5 6F D8 C4 5B 92 2E 72 B7 FC FE E5 69 5F CD DE 7B 83 34 45 2A 90 66 48 25 D2 1C 69 81 B4 54 DA 9D 59 A4 7B C5 BF 38 12 2F 17 BD F5 C5 4D 27 7A F7 B8 B9 91 4B 4D 6E 55 0C 43 F1 C4 32 62 AD D0 5E CC 52 5B 4D 9E 59 88 F9 F2 F3 35 5B BD 1B CA 99 7B B9 88 F5 83 C5 73 AE 41 B0 B8 1E 43 6F E6 46 B3 11 52 8E 34 0E 8D C5 95 98 78 4B C2 55 AD 43 F2 28 06 8C 75 F9 32 E6 62 8F 6F C9 DE D8 85 54 20 CD 90 4A A4 39 D2 02 69 A9 84 E1 6A 3C DF 3D F3 BB A6 DC 3D A7 26 5C 6A 36 5C C4 32 62 AD D0 9E 09 97 9A 0D 57 B0 B8 D5 BB C4 7A C4 FA C1 4C B8 82 99 70 79 B3 E1 42 CA 91 C6 A1 31 13 2E 6F 36 5C 9E 92 70 ED 3F 46 A6 65 4C B8 90 A6 48 05 D2 0C A9 44 9A 23 2D 90 96 4A 24 5C CF 37 52 DF D0 7B D3 6D B8 BC 25 E1 42 CB B4 AE 29 D7 0A ED D9 70 F9 BA 49 B8 D4 6C B8 D0 7A DA 9E F9 9E F7 83 D9 70 69 5D 1B AE CA 92 70 01 E5 BE 35 53 6A 1C 66 60 C3 85 C3 F4 BE 58 12 2E 78 CD 6E 55 CD EE 16 B5 9A B9 4A 8D 54 20 CD 90 4A A4 39 D2 02 69 A9 44 C2 F5 7C C3 F4 0D BD A5 DC 86 CB 5B 12 2E B4 4C EB 26 E1 D2 F6 6C B8 BC 25 E1 52 B3 E1 42 EB E9 3C 92 70 69 39 1B 2E 35 1B AE CA 92 70 01 E5 7E 0E 49 B8 B4 31 1B 2E 1C A3 F7 35 93 70 ED 3F C0 AF 65 6C CF 55 B5 64 77 8B 58 AA 40 9A 21 95 48 73 A4 05 D2 52 89 84 8B 8E D1 EF 5E 69 F8 17 5E 0F D6 20 F7 D1 AB 25 39 F3 E5 8C 65 A4 5C 2B B4 67 73 E6 EB 26 39 53 B3 39 43 EB 69 7B 49 CE B4 9C CD 99 9A CD 59 65 49 CE 80 72 3F 87 24 67 DA 98 CD 19 8E CF FB 9A EE E5 A4 F1 0D 52 FB 67 8E 5A 28 9E 62 5C 23 4D 91 0A A4 19 52 89 34 47 5A 20 2D 95 70 7C BE F1 4C E3 F3 BB 76 F6 0F BE FC 18 7B 12 2D B4 4C EB 26 5D 98 2F 27 0F 5B EA 51 7C 5B CB 25 D1 F2 E5 EC 91 7D 28 67 8E EC 89 F5 83 D9 68 69 7B 36 5A 95 25 D1 02 CA 7D 6B 49 B4 B4 31 1B 2D 1C 9C F7 35 E5 06 70 F3 FA E2 FD CB 8C B1 50 18 9C 47 9A 22 15 48 33 A4 12 69 8E B4 40 5A 7A 92 31 7A 59 AE E4 BC B1 C1 07 E7 BF 7B 50 62 D7 CE 7E B4 FC E8 7A 12 2D B4 4C EB 26 D1 F2 E5 92 68 79 4B A2 A5 66 7B 2D B4 9E CE 23 E9 B5 B4 9C 8D 96 9A 8D 56 65 49 B4 80 72 3F 87 24 5A DA 98 8D 16 0E CB FB 9A E6 79 83 2B A4 6B A4 29 52 81 34 43 2A 91 E6 48 0B A4 65 F8 A8 30 48 7C 58 FE FB 83 A4 2F 7C B1 87 59 DE 92 20 A1 65 EE 95 8F EE 4E 8A 24 48 DA 9E ED A3 BC 25 41 52 B3 41 42 EB E9 3C 92 20 69 39 1B 24 35 1B A4 CA 92 20 01 E5 7E 0E 49 90 B4 31 1B 24 1C 84 F7 35 65 94 48 BB 9F 2B 4F 67 76 67 57 55 3C 8B C3 64 53 2D 15 A9 40 9A 21 95 48 73 A4 85 52 5C AE 65 F8 A8 30 48 CF 34 08 DF D0 37 BB D8 20 79 4B 82 84 96 69 DD 24 48 DA 9E 0D 92 B7 24 48 6A 36 48 68 3D 9D 47 12 24 2D 67 83 A4 66 83 54 59 12 24 A0 DC CF 21 09 92 36 66 83 54 59 32 D2 50 51 12 24 3F 28 6F 83 E4 C9 06 09 A8 F0 4B 61 E2 36 43 2A 91 E6 48 0B 25 1B 24 FD F8 30 48 CF 34 DE DE D0 B1 75 1B 24 6F 49 90 D0 32 AD 9B 04 49 DB B3 41 F2 96 04 49 CD 06 09 AD A7 F3 48 82 A4 E5 6C 90 D4 6C 90 2A 4B 82 04 94 FB 39 24 41 D2 C6 6C 90 2A 4B 82 E4 49 3E A4 F8 A3 0F FB 0F 1E FA E6 E5 61 A8 78 D4 14 EA 29 4D B1 54 81 34 43 2A 91 E6 48 0B A4 65 F8 F0 30 5A 7C B4 FD FB 77 76 3A D4 1E 3F A6 0B F9 85 82 EA 52 8E B9 50 48 2C 23 D6 0A 16 2F A9 B4 D5 E4 3D A9 61 A8 3D 58 9C 6F 97 58 8F 58 3F 58 9C C7 40 CD 7C F4 C3 60 E6 B1 69 62 39 B1 71 B0 78 79 60 E2 2D 49 97 BF 11 3E 6E A9 2B 2D 65 B3 04 A5 A6 58 AA 40 9A 21 95 48 73 A4 05 D2 52 D7 A8 FA 61 4F FB 63 74 F2 1A 84 67 79 8D EE AE 1D F7 36 FB F8 B9 5C 78 AB 99 CE F7 92 58 46 AC 15 DA 33 6F 78 08 E5 A2 75 48 DD 2E B1 1E B1 3E 99 C7 80 CC 63 48 EA 8E 88 E5 C4 C6 C1 62 B7 37 F1 66 B3 E4 C9 EE F2 3C D9 63 27 25 B3 CB 43 2A 90 66 48 25 D2 1C 69 A1 64 76 79 BA 46 24 4B FB 63 E9 7F 7C BF 9E 3C 7E BB EB 6A 6C 62 3C 99 F7 29 5C 6A 31 63 19 B1 16 B1 36 B1 0E B1 2E B1 1E B1 3E B1 01 B1 21 B1 11 B1 9C D8 D8 9B 79 1F FD C4 53 12 98 6A 4B 25 81 A9 28 09 8C 27 1B 18 A0 C2 37 6F 8F 91 90 4A A4 39 D2 42 C9 06 46 AF 77 C0 8E EC 74 7F 7C FC 5F 04 C6 0F F7 9A FB 5A 76 2D B8 57 C5 C7 F9 5D 12 CB 88 B5 88 B5 89 75 88 75 89 F5 88 F5 89 0D 88 0D 89 8D 88 E5 C4 C6 DE CC FB B8 26 9E 92 C0 54 1B 4F EE C2 0D 67 67 5A D1 1C 54 2B D9 C0 54 15 4D 3A 0A 2C 35 43 2A 91 E6 48 0B 25 1B 18 BD 86 81 81 A1 63 DE D5 8B AE BE FD 66 3B 79 35 4B F5 D2 C6 B0 29 2E 94 CC EB F3 2E 89 65 C4 5A C4 DA C4 3A C4 BA C4 7A C4 FA C4 06 C4 86 C4 46 C4 72 62 63 6F E6 0D 5C 13 4F 49 8E FC 4D E8 36 47 15 25 1D 8F 27 9B 23 A0 42 E7 18 4B CD 90 4A A4 39 D2 42 C9 E6 48 AF 51 60 8E 9E 69 48 5B EE A8 DF 25 C9 0D 4C C5 33 8C BD 21 FF 8B 50 CA DC 1E 15 2C AE 7B A6 E6 56 45 5A 5B CB 3D 31 1F DF 5E 8E F3 CB 77 D3 56 2E FF FD ED 70 74 F8 F7 AB 5E A7 3B FD 5B 76 7A F6 F7 61 2B 77 7F FE F0 63 ED 87 1F 5E 1F 7D 74 3F 16 FE 9F B3 77 C3 A2 F5 1F E9 CF 9E B4 B4 5D B9 AB 35 0C 94 87 79 C5 63 F0 8E 9A 7B 77 53 5C 9B BD 5B C0 BB A1 94 19 3A 0F 16 5B EB 87 B9 36 DC 7A 1C DC AC A4 9B 3E 7C 3B 1D 67 EF 16 7F D3 E5 6D 34 4E 1A 7B ED 0F C8 D2 0E 89 8D 88 E5 C4 C6 C1 E2 D6 9F A8 D9 97 20 79 4B 3A C8 EA E3 4D 82 ED C9 06 1B A8 F0 6D 25 7B 54 28 55 62 A9 39 D2 42 C9 06 5B AF 90 60 B0 9F 69 40 FD D4 8F 1D CB 49 A8 46 E6 22 58 3C E9 BE 54 93 83 42 2D 97 A9 C9 BD 88 6A AD 60 B1 BD 76 A8 1B 63 D9 51 93 1B C2 B4 6E 37 58 FC 00 7B C1 E2 3C FA 64 BE 03 32 DF 21 B1 11 B1 9C D8 98 D8 44 2D 09 53 B5 FD 92 30 55 94 84 C9 93 0D 13 50 E1 DB 4F C2 04 A5 4A 2C 35 47 5A 28 D9 30 F9 8F 9A 1C CF 3F D3 A0 BA FB C9 30 37 30 EE 2E 31 C6 0B A3 7B 57 E0 2F B4 94 3C A1 1D 6E F2 53 93 07 6D 63 BC 7C 6B 72 C3 58 8C 97 CE 21 96 6B 6B 5D 39 22 0D 23 0F 61 1E 32 C0 13 7B B8 BD 5F 67 EA 92 25 E9 91 25 E9 07 8B 4B 32 08 EB 6A 5E E1 46 6C 44 2C 27 36 26 36 51 4B 02 57 6D 81 E4 7C C0 8F B4 DB C3 3B 1C 7C F7 8D 25 87 77 50 6A 86 A5 4A A4 39 D2 42 C9 06 CE 7F 58 24 70 6E 34 F5 5B DF B9 FC 2F 8F F7 E2 3D A5 A7 7E 94 56 2E 44 C6 DE CC 9B 3C 03 19 E3 A6 16 77 6D 59 A8 6B 5E 6A 1A 2C D6 6D AB C9 F1 6C 8C 9B B6 17 EB 76 B5 9C 99 6F 2F 58 9C 6F 9F CC 77 40 E6 3B 24 36 22 96 13 1B 13 9B A8 1D C7 D8 BC 57 33 EF 77 40 BA 46 9A 22 15 48 33 A4 12 69 8E B4 40 5A 7A 92 3F E4 53 48 AE 35 9F BA 41 5D 1B AE 7F 71 B2 59 8D 01 DB C4 EC CB E5 AE C9 37 87 B1 4C 06 D2 02 69 83 74 40 BA 20 3D 90 3E C8 00 64 08 32 02 C9 41 C6 20 13 90 F7 20 57 20 D7 20 53 90 02 64 06 52 82 CC AD 54 9F F2 D1 D3 E7 F5 7A 9B AD B6 AB B7 AF 3F C8 FF 5F 6E 1E 9E 36 77 B7 F2 D7 B5 24 E1 7E FD F8 69 7D B9 BE BB 7B 3A B8 D9 7C 7D 90 BE BF 76 EC DE D8 11 FC E0 71 FD F1 CD A1 BC A3 AC E9 9E 6B B2 35 AA 29 F2 66 B1 A6 7B 1A 09 A7 BC 3B 79 D9 74 CF C8 E0 14 79 4A 46 5A 63 53 E4 D9 16 69 8D 4D 79 57 3B 6E BA DF 4A 24 F3 39 3D 6E BA 61 51 32 1F 99 E2 06 33 49 9D DA 59 F3 42 7E 43 81 4D 39 97 29 BB 1D D7 DE 36 78 57 3F 6E 5E 54 57 8C 60 4A 5D A6 EC FA A8 BD 29 BD 5A A3 D9 AF DE 15 BF 37 45 7E E5 40 D6 87 CD A7 23 F3 E9 D0 F9 74 EA 75 99 C2 E6 73 29 AD B9 57 DD E3 FA 5C 4A 6B EE C5 D2 6C 4A 5D A6 B0 D6 E4 37 FF 64 D9 D8 94 77 B2 3E EE 47 92 B0 B5 AB D3 B3 A6 DB A9 91 29 92 10 F7 A6 35 32 E5 65 53 7E F3 9A B5 25 4D 31 AF BD 6A BA D7 CA B3 79 1C CB 3C D8 27 7D 55 AF C9 14 B6 FE 57 B2 35 DD 2B 7A 59 6B 27 32 85 AD FF 7B D9 CE EE D7 12 58 1D 99 52 BD C8 60 EF 93 96 97 93 C9 14 B6 9E F2 8E 44 99 42 D7 F4 44 D6 A7 3A B8 DB 6F ED 54 96 BA DA 97 C2 14 59 6A 39 45 66 DB B3 21 53 76 BD 31 D4 91 25 90 63 5F AC 23 A7 A9 4D 77 54 CF A6 C8 37 BE DA 8F EC 7F 0F 24 03 97 BF 33 A5 21 53 58 6B 72 A2 DE 74 C7 08 38 1F 39 ED 97 29 AC 8E FC E6 6E D3 FD 52 2E D6 91 DF CA 6D BA 5F 24 C5 29 F2 4B A8 52 87 4D E9 49 DA FB 34 ED F2 D3 94 52 87 6D 51 F9 91 49 99 0F FF 26 CA 52 FF CE 37 BE 21 DF 2A B6 3E F2 8B 4D 32 85 6D 83 4B 99 E2 7E 83 87 7C 7B 65 09 DC EF 75 B2 29 27 32 85 2D 9B 5C C4 91 4F 8E F6 87 27 E7 4D 77 99 12 5B 93 67 CE 9A EE E2 22 D9 A2 32 C5 5D 12 C4 29 EF EA 92 1D 9A EB 77 D2 17 B4 E8 DE A2 55 93 8E 9F B4 D5 92 A6 E8 DE 45 BA 4F 36 EF 9E 2C 14 5D A6 5A 33 63 DB B1 25 3B 1C E6 93 5A D3 1D F2 E1 CA 5D 9C 4A 90 E8 D7 4C 7E D9 4A 42 C1 EA F4 A4 DB 70 3F 1D C2 62 79 2E 11 63 53 E4 77 66 A4 35 BA 8A D2 A5 BB 97 FB 93 D6 64 8A 7B 4D 3F F9 62 C8 14 F7 B2 7E 9C D2 93 6E D0 BD 74 9B B5 26 F1 A7 53 E4 05 ED D2 1A AD 23 3B 08 F7 B3 72 EC AB 29 9F 24 9D D2 93 ED E6 7E 19 8A 6D 9D 33 69 8D 4D 91 31 B9 A6 1B EC C0 3A 1D E9 50 DC 89 03 4E 79 5F 6B 2E 69 77 DF 90 1D 94 DC F9 84 35 AE 1A D2 75 CB DD 2D 64 8A 7C 95 DC 45 40 36 45 76 38 D5 B9 F5 5E F7 F8 5E 0E 6E DC 83 A8 64 C9 E4 AB E4 9E 22 64 53 5E C9 14 96 82 F7 8D E3 A6 7B 7E 87 D4 69 C8 8A CA C3 17 6C 4A 5D A6 B0 6D F0 BE 71 22 53 E8 2E EF A4 21 4B C0 77 EC B2 3B AE 46 25 F6 D7 B4 21 3B 1C 19 FB 65 4B 20 BB C9 AA D3 38 8A C7 9E 6F 5F 7F 59 7D 5A 8F 56 8F 9F 6E 1F 9E 0E EE D6 1F E5 38 F4 F8 85 7B E3 EF E3 ED A7 CF E1 1F DB CD 17 E7 EE 3E F2 9F 36 DB ED E6 5E FF F5 79 BD FA B0 7E 74 FF 92 91 94 8F 9B CD 56 FF 21 0B E0 5A BE 5E 6F BF 7E 39 78 BA 59 DD AD DF 1C BA 43 E4 CD E3 ED FA 61 BB DA DE 6E 1E DE 1C DE AD 1E 3E C8 B4 2F 6B 99 5D F3 F6 C3 9B C3 C7 DE 87 DD D6 FB F0 B8 FA F5 F6 E1 53 D4 EA CC FC D7 CD E3 CF BB 83 EA B7 FF 07 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 26 2B A3 14 87 1B 00 00 9E 9C 00 00 18 00 00 00 78 6C 2F 77 6F 72 6B 73 68 65 65 74 73 2F 73 68 65 65 74 32 2E 78 6D 6C B4 9D 6B 6F DB 48 96 86 BF 2F B0 FF C1 E3 05 06 DB 40 27 36 65 C9 96 35 71 06 BA 92 D4 85 22 29 F1 22 7E 53 DB 72 62 B4 6D 79 65 A5 D3 83 C5 FE F7 2D 8A 55 AC 53 F5 9E 64 9D 8E 16 98 9E 24 8F 4E 55 1D 52 6F 1D 92 F5 52 E4 87 7F FE F9 F4 78 F2 C7 66 F7 FA B0 7D BE 39 75 DE 9F 9F 9E 6C 9E 6F B7 77 0F CF 9F 6E 4E 93 E5 E8 5D FB F4 E4 75 BF 7E BE 5B 3F 6E 9F 37 37 A7 FF DA BC 9E FE F3 E3 BF FF DB 87 AF DB DD EF AF 9F 37 9B FD 89 E8 E1 F9 F5 E6 F4 F3 7E FF D2 39 3B 7B BD FD BC 79 5A BF BE DF BE 6C 9E C5 27 F7 DB DD D3 7A 2F FE B9 FB 74 F6 FA B2 DB AC EF 0E 8D 9E 1E CF 1A E7 E7 97 67 4F EB 87 E7 D3 AA 87 CE EE 2D 7D 6C EF EF 1F 6E 37 83 ED ED 97 A7 CD F3 BE EA 64 B7 79 5C EF 45 FE AF 9F 1F 5E 5E 55 6F 4F B7 6F E9 EE 69 BD FB FD CB CB BB DB ED D3 8B E8 E2 B7 87 C7 87 FD BF 0E 9D 9E 9E 3C DD 76 FC 4F CF DB DD FA B7 47 B1 DD 7F 3A CD F5 ED C9 9F 3B F1 BF 86 F8 EF 42 0D 73 E0 30 D2 D3 C3 ED 6E FB BA BD DF BF 17 3D 9F 55 39 E3 E6 5F 9F 5D 9F AD 6F EB 9E 70 FB DF D4 8D D3 3C DB 6D FE 78 28 BF 40 DD 55 E3 AF A5 E4 B4 EA BE 1A BA B3 8B BF D8 D9 65 DD 59 B9 BB 76 9D 2F 0F 77 37 A7 FF 7D D5 6F 5F 0E BB 17 FD 77 4E FF DA 79 D7 BC BA EE BE EB 9E 37 2E DF 5D 76 9B 17 CD EB E1 B0 35 B8 1C FD CF E9 C7 0F 07 9D 84 BB 8F 1F 5E D6 9F 36 8B CD 3E 79 09 77 27 F7 0F FB E5 36 14 40 68 F5 F4 EC E3 87 B3 3A EA EE 41 08 A2 DC 09 27 BB CD FD CD 69 D7 E9 74 7B AD 76 19 73 08 49 1F 36 5F 5F C9 DF 4F 4A FD FE B6 DD FE 5E 7E E0 8B BC CE CB 21 37 8F 9B DB 52 49 27 6B F1 C7 1F 9B FE E6 F1 F1 E6 34 76 C4 A4 78 FD AF 43 B7 E2 EF 9D D8 21 43 97 CD 55 1A 74 8C D1 41 F7 22 E3 BB CD FD FA CB E3 3E DE 7E F5 36 0F 9F 3E EF 45 E2 2D B1 37 4A 39 75 EE FE 35 D8 BC DE 0A 1D 8B E1 DF 37 5A 65 B2 B7 DB 47 91 A6 F8 FF 93 A7 87 72 42 5E 0A 21 AE FF AC FE F2 F5 E1 6E FF F9 E6 F4 FA BD D3 3C BF 14 E1 27 B7 5F 5E F7 DB A7 AC C2 87 A4 74 C3 6B D5 50 FC 45 36 6C 7C A7 E1 59 35 F0 61 5F 0D D6 FB F5 C7 0F BB ED D7 13 21 49 47 6C FB CB BA 9C E0 4E A7 21 8A 81 DA 82 6A EC 7A 9B BE B5 49 62 5B CA 5E BA 65 37 22 F5 6A 1B 4B D2 93 E4 F2 B0 D5 25 E9 03 19 54 A4 71 51 6D 5B 19 34 54 48 48 4A EC B6 57 F1 AD FD F1 D1 69 34 3E 9C FD 21 BE 86 5B 39 DC A8 8A 6A 1F BE FF 43 02 2E 10 0F 88 0F 64 0C 64 02 64 0A 64 06 24 00 32 57 DB D1 AC B7 3F AC 90 DE D6 C8 06 B1 0D 16 B2 97 4B DD 68 29 51 4B FC 59 EF A0 E6 95 B9 7F 12 1D 54 6A AE DC B1 29 A2 0C 51 8E 68 85 A8 50 A8 51 6F 5C 57 2A E0 42 13 A9 80 C3 0E 38 13 6A AB 25 27 4A CF 5F 96 5C 53 08 A2 52 5C D9 8B 50 DC 55 3D 60 4F 12 2D 8A 3E 90 41 45 9C AB EB BA D9 B0 42 8D 0B AD DD 51 85 DA 3A C8 05 E2 01 F1 81 8C 81 4C 80 4C 81 CC 80 04 40 E6 2A 69 3D BD C2 0A 55 35 F1 B0 8B 22 20 71 45 B4 26 17 B2 9F 4B FD 4D 2E 25 6A E9 AF 32 41 94 22 CA 10 E5 88 56 88 0A 85 74 5E DD 6E C5 A8 9C 48 EE 86 9C 44 A5 78 9B 9C F6 9F 1F 6E 7F EF 6D CB 12 CD 96 B3 66 AB 16 57 D9 A7 29 2E 49 88 B8 80 0C 2A 62 88 AB 42 8D 0B 2D D3 91 42 A4 78 21 F2 10 F9 88 C6 88 26 88 A6 88 66 88 02 44 73 85 F4 44 08 2B 44 55 06 24 AE 08 51 99 EC 87 AA 4C A2 96 38 CC E9 32 D6 B6 CA 98 0E AA CB 18 A2 0C 51 8E 68 85 A8 50 48 4F A2 6E B7 62 54 77 64 6B 0C DD 35 8F 73 E4 2C BB 31 A5 26 09 91 1A 90 41 45 0C A9 55 A8 D1 14 67 33 E4 C8 79 61 1D 39 55 94 3E A2 B8 88 3C 44 3E A2 31 A2 09 A2 29 A2 19 A2 00 D1 5C 21 5D 9B C2 0A 69 65 45 36 88 6D B0 90 BD 50 ED 49 24 B4 A7 64 95 20 4A 11 65 88 72 44 2B 44 85 42 54 69 76 A6 DD 1E 21 86 D2 C4 1C 61 2A DC FB AB FA 14 F1 CD A7 69 65 4F A6 D8 24 21 62 03 32 A8 88 21 B6 0A 35 9A 7A A2 8C 14 D2 DF 8E 8B C8 43 E4 23 1A 23 9A 20 9A 22 9A 21 0A 10 CD 15 D2 02 08 2B 24 26 8F 92 44 04 24 06 B2 90 1D 51 71 49 D4 BA A2 85 ED DA 2A 6C 3A A8 2E 6C 88 32 44 39 A2 15 A2 42 21 FD AD 76 BB 32 79 72 7E 46 88 21 37 71 4D F2 66 B9 FD 5F C7 D4 8B FA 90 5A F6 6A 4A 4F 12 22 3D 20 83 8A 18 D2 AB 50 A3 A9 E7 D2 48 21 7D 94 75 11 79 88 7C 44 63 44 13 44 53 44 33 44 01 A2 B9 42 7A AB C3 0A 91 6A 14 01 89 81 2C 64 47 54 7A 12 B5 F4 E1 3A 41 94 22 CA 10 E5 88 56 88 0A 95 84 9E 36 DD 6E C5 C8 B9 5C 8F 10 43 68 62 8A FC 3F 08 AD EC 55 08 4D EF 84 5E 45 9C 73 9D 65 1F D1 A0 42 8D 0B 7D B8 19 4A D4 14 D7 D8 FA 24 C5 B1 8E A8 3A 48 CD 65 17 91 87 C8 47 34 46 34 41 34 45 34 43 14 20 9A 23 0A 11 45 88 62 89 5A 7A 0F 2E 24 BA D4 47 80 65 85 F4 37 9F D8 20 B5 41 66 83 DC 06 2B 1B 14 15 D0 A5 BB DB B5 43 BA F2 1B C7 4B CF 72 B5 13 57 3B DE 7C 24 D5 D5 AC EC E8 E6 54 6F 7B AF 02 8D 4B 43 2A D6 B2 45 5F 06 39 BA DD 40 23 22 31 EB A4 6D A8 82 F4 AE 1D 21 72 11 79 88 7C 44 63 8D 48 0E 56 EE 13 6C 37 65 DB 35 CD E9 31 C3 76 01 A2 39 A2 10 51 84 28 96 E8 8A 2E 87 38 97 66 0E 0B 19 74 A9 F7 DF B2 42 44 AB 15 70 2E C4 9F F5 5C 6F 9D 9B 1D A5 3A 48 CD F5 0C 51 8E 68 85 A8 50 49 51 25 DB 59 75 A5 AE 50 C9 42 67 A8 E4 CB F7 E2 98 FE 96 95 3B 2D E4 B2 1F 43 C8 15 68 5C 91 62 29 91 58 D8 54 5B 3D 40 34 44 34 42 E4 22 F2 10 F9 88 C6 88 26 88 A6 88 66 88 02 44 73 44 21 A2 08 51 AC F6 97 3E 6C 2C 24 A2 72 AB 10 91 5B 05 9C 6B 71 75 A9 E5 66 1D 5A 52 1D 54 CB 0D 51 8E 68 85 A8 50 49 D1 4B 10 3B AB 6E 8F 10 E3 50 7D 58 23 3F 4A E5 3C F4 64 28 4E 12 C7 11 2B 3C E5 BE D8 EF C4 49 E3 FD C7 E5 7F 76 FF D6 75 9A BF 7C 38 BB 17 8B C0 A5 19 A0 96 7F FB 75 0B 52 48 15 53 15 58 F6 E2 8F 44 37 7D A7 F9 F7 C7 FD 3F FE FE 69 FF 8F D3 D3 5F FB F3 A0 DF 5D 0E 03 F1 DF E1 33 E7 57 01 63 DF F5 CA 11 45 E8 AF D3 61 50 FD ED 97 77 E7 BF FC 22 3E 65 72 18 AA F1 C8 44 19 31 CC 65 98 C7 30 BF 66 FA CA 7C 2C 59 A3 A1 4F 6C 27 0C 9B 2A 26 EC 04 25 94 19 C3 02 86 CD 19 16 32 2C 62 58 E5 9C DC 9C 36 A4 B7 51 AE 33 2F 54 1C 55 BF 64 44 FE 6A 6B 4D FD 5B C7 9C 94 44 D5 13 80 61 39 C3 56 0C 2B EA DC E8 24 80 E4 BA 3D 8A CC 69 50 AE 7C 1F 67 1A 54 6B E8 E4 0C 42 D8 4E 65 29 76 1C BD 9B FA 35 23 A5 57 31 72 E6 35 AC 19 39 43 60 98 CB 30 8F 61 7E CD F4 B8 63 C9 4C 31 4A 27 80 08 74 AA E2 0C 31 CA 38 C2 02 26 6E CE B0 90 61 11 C3 4A DB AE DC 7F A6 18 25 33 C4 58 31 2A 46 15 25 4E 24 75 31 B6 4E C2 52 D5 FF A5 9E 8C 19 C3 72 86 AD 18 56 D4 8C 5E A5 4B 48 AF 9E 28 32 C5 58 AE 88 1F 47 8C F6 6A 7B AF AC C5 87 9D 29 AE DD A0 26 5F B2 35 59 B5 D0 BB 67 20 7B 71 F8 9A 7C F9 9D 9A 7C 61 D6 E4 CB BA 26 5F 7E B7 26 57 39 38 46 4D 46 E6 AA BC 48 9C C7 30 BF 66 B4 26 EB ED B4 F7 8C CF 1E AD 26 F5 BE D4 7B 66 CA B0 19 C3 02 86 CD 19 16 32 2C 62 58 5C 33 3D B1 17 8A 19 93 84 B8 2D 07 33 29 A9 A3 8C 49 62 9D ED A7 24 4A 57 6C B9 C7 C8 C4 C9 99 B8 15 C3 8A 9A 19 93 C4 4E 4E 54 EC 6F B9 43 E2 5A EB 58 93 C4 B6 06 7A 87 BE CB 8A 43 4D 6E 69 2A 90 8A 38 90 71 8E 51 B1 AB 38 87 EC F4 51 1D 47 96 50 19 E6 31 CC AF 19 AD D8 98 CB A4 CE 99 8A 11 E3 66 4C 5C C0 B0 39 C3 42 86 45 0C 8B 6B 46 C5 28 73 31 C4 48 2C 18 29 46 15 65 88 B1 65 5D AD A9 FE 8D 8A AD 5B 2A 81 E6 4C DC 8A 61 45 CD 0C 31 DA C9 09 31 7E CB 32 12 87 F6 63 AD E4 1F BA 32 4F A3 A5 C9 60 A8 51 31 5A 94 2B F6 8D A2 DC FE 4E 51 6E 99 45 B9 5D 17 E5 F6 77 8B B2 1C CF 28 CA C8 5C B9 45 B4 78 7B 0C 13 45 B6 3A 4F 12 97 FC EA 1B 1C 4B 56 9E 9B 60 51 E6 0E 57 13 D2 42 F5 32 55 3D 5F EB 6B DF 19 C3 02 86 CD 19 16 32 2C 62 58 5C E7 42 E7 81 FC E6 8C 79 60 5B 46 89 6C A9 EB 45 0A 24 03 92 03 59 01 29 24 A1 8B 13 10 24 94 FE 2D CB AA BC 09 EA 48 9E D5 A1 2B 53 E9 D2 DF 30 94 AE 18 55 7A C5 CC BA AB 18 3D 53 46 E6 CA 51 69 7D F6 18 E6 D7 8C D6 5D CC 65 22 E3 E8 D9 F3 54 B5 35 F4 26 73 21 2C 60 E2 E6 0C 0B 19 16 31 2C AE 73 A1 7A 93 39 1B 7A AB 98 B8 C7 86 38 CA F6 E2 5A 22 7B BB D6 45 31 45 94 21 CA 11 AD 10 15 6A 03 C8 2D 3B E2 66 23 74 B3 28 32 4F 93 4B FB E1 27 4E 93 C5 C5 6D 7D 97 5B E5 64 D0 8B 36 E9 80 08 29 E2 79 72 9B 3D 4F 56 2D A8 50 2B C6 97 E4 C6 F9 77 4A F2 95 51 92 1B E7 AA 24 37 CE BF 5B 92 E5 78 46 49 46 E6 96 B7 27 96 97 A4 C6 79 32 32 BF 8E A3 25 59 6F 27 96 64 6E CF 4C 64 2F E6 14 91 A3 19 53 04 59 A0 32 20 71 73 86 85 0C 8B 18 16 D7 B9 D0 29 22 B7 C8 98 22 C4 DB 92 A7 26 36 49 65 5F BA E0 64 40 72 20 2B 20 85 24 46 49 B6 07 13 25 99 20 73 1E 94 36 C9 91 E6 41 E5 B8 D0 79 20 ED 20 A3 24 2B 46 95 5E 31 B3 24 2B 46 4B 32 32 D7 41 E6 31 CC AF 19 2D C9 98 CB 44 C6 99 7A 93 63 18 7A 43 16 A8 31 0C BD 61 5C C8 C4 45 0C 8B EB 5C A8 DE 64 CE 86 DE 2A 46 17 2F 6C 92 CA BE A8 DE EC 98 1C 62 56 40 0A 49 0C BD D9 1D 09 BD 11 64 EA AD 34 33 8E A4 B7 CA 17 A1 7A 93 EE 09 57 77 45 ED E3 D6 8C 55 0B AA 46 69 CB B0 EB 13 8D C6 77 EA EE B5 59 77 1B 75 DD 6D 7C B7 EE CA F1 8C BA 8B CC 75 90 79 0C F3 6B 46 EB AE DE 4E A8 BB EC 9E 99 C8 5E CC 79 20 33 30 E6 01 B2 40 65 60 CC 03 8C 0B 99 B8 88 61 71 9D 0B 9D 07 CA 29 23 F6 9D 8C A3 F3 C0 F6 CE 52 88 C9 80 E4 40 56 40 0A 49 8C 79 80 46 1D 6D 67 CE 83 D2 53 79 EB 3C 78 FB FD 34 4E 65 D5 D0 39 21 2D 1E 71 1B BF BA A4 E9 CB A8 46 43 DF 0E 31 90 CC AC C1 D2 31 32 96 23 90 B9 75 5B BD E7 3D 86 F9 35 A3 35 58 E5 A7 73 99 30 F9 4D 15 13 95 45 BB 19 B2 2D 61 01 13 37 67 58 C8 B0 88 61 B1 62 17 E4 2E 07 C5 8C 1A 0C 66 9E 8C A2 97 61 76 4C 06 31 39 90 15 90 42 12 43 7B E8 DA D1 76 E6 AD F6 E5 4F 5B 8E AF BD 46 D9 AD 71 49 26 09 BD F7 BE 8A D1 64 00 31 43 20 23 20 2E 10 0F 88 0F 64 0C 64 02 64 0A 64 06 24 00 32 07 12 02 89 80 C4 40 16 92 34 E8 0D 34 92 E9 E5 C4 04 48 0A 24 03 92 03 59 01 29 24 D1 73 4C DC 73 5F 7D 65 D4 7C A0 C8 54 D6 4F 3A 61 E4 AA AA A1 CC 21 51 54 F5 6D 56 D6 62 5E AF 8E D2 D5 A3 CF B0 01 C3 86 0C 1B 31 CC 65 98 C7 30 9F 61 63 86 4D 18 36 65 D8 8C 61 01 C3 E6 0C 0B 19 16 31 2C 66 D8 42 31 F1 2D AB 4A BB 64 E2 12 86 A5 0C CB 18 96 33 6C C5 B0 A2 CE 45 9F CB 08 45 DA 5E 5D 57 E9 E0 20 52 53 91 3F 62 87 BD FD 38 DB 50 7E 8F BE C1 A3 57 33 7D F4 ED 33 6C C0 B0 21 C3 46 0C 73 19 E6 31 CC 67 D8 98 61 13 86 4D 19 36 63 58 C0 B0 39 C3 42 86 45 0C 8B 19 B6 60 D8 92 61 09 C3 52 86 65 0C CB 19 B6 62 58 51 33 52 0A BB 12 1A D5 F1 5B AE 53 E3 67 5D 27 BD E6 74 E8 EA E6 B4 49 E5 27 FD 05 3D 53 FA 2A 4A A3 01 A2 21 A2 11 22 17 91 87 C8 47 34 46 34 41 34 45 34 43 14 20 9A 4B 44 EE 9C 06 12 01 89 81 2C 80 2C D5 60 FA F7 20 40 52 20 19 90 1C C8 0A 48 A1 88 71 E0 45 0F 89 36 34 CB DC 4F 7A 48 E2 77 8E F5 4F 28 E5 EA BD F8 43 1F 77 AD 7B 27 7B 8D 2A E8 82 1E 76 01 0D 30 6A 88 68 84 C8 45 E4 21 F2 11 8D 11 4D 10 4D 11 CD 10 05 12 E9 A9 3D 07 12 02 89 80 C4 40 16 40 96 40 12 20 29 90 0C 48 0E 64 05 A4 50 C4 D0 9A 6D 22 89 43 EA B7 5C 9C F2 47 E8 3F 75 87 09 D1 9A 5C B1 17 7F E8 5B 6C 40 6B 55 90 A1 35 40 83 43 56 E2 EA 83 FE 86 17 A2 46 18 E5 22 F2 10 F9 88 C6 88 26 12 E9 62 34 05 32 03 12 48 42 85 66 3B 19 21 C4 44 40 62 20 0B 20 4B 20 09 90 14 48 06 24 07 B2 02 52 28 62 08 0D 3D 1A DA D0 2C 6A E5 A2 F5 CF DC CA 44 84 26 D7 BF F5 3E EE 95 8B 84 A5 8D 21 FE 2C 6F 2C ED FE 2D 16 17 17 E5 4D A5 1F AD FB BB FB 24 B2 BE D5 99 61 43 86 8D 18 E6 32 CC 63 98 CF B0 31 C3 26 0C 9B 32 6C C6 B0 80 61 73 86 85 0C 8B 18 16 33 6C 21 19 59 A8 03 92 00 49 81 64 40 72 20 2B 20 85 22 86 0A D1 21 A1 0D 4D 15 96 4B D9 C7 59 B1 2E 6F 44 28 25 A7 2F E5 7B 12 D1 25 92 2A 88 2E 91 D8 64 08 AD 46 40 5C 20 1E 10 1F C8 18 C8 04 C8 14 C8 0C 48 20 09 AD 69 B6 4B 10 42 4C 04 24 06 B2 00 B2 04 92 00 49 81 64 40 72 20 2B 20 85 22 86 9A D0 FF A0 0D 4D 35 FD A4 FF 41 4F D4 E4 72 B3 F1 63 01 EB 99 1F BD F2 E9 2D 42 72 54 5F 36 19 40 CC 10 C8 08 88 0B C4 03 E2 03 19 03 99 00 99 02 99 01 09 24 A1 FA B2 57 DF 43 88 89 80 C4 40 16 40 96 40 12 20 29 90 0C 48 0E 64 05 A4 50 C4 D0 17 FA 0A B4 A1 A9 AF 1F F1 15 F8 07 12 D5 17 02 E0 25 08 BF A0 14 53 4B 54 CF FA 74 ED CA BA 23 BA 2F 83 2E 68 50 CB 7A 86 C3 40 06 89 25 6A 75 44 1D 22 1A 21 72 11 79 3A 2B D5 97 CF 26 6A DD 95 3A E6 12 BD B2 16 18 27 3A A8 BE 21 0C 53 98 21 0A 10 CD 31 D1 90 4B F4 DA FA C1 43 C4 25 7A 6D ED F6 58 05 69 E1 2C 30 85 25 A2 04 51 8A 28 43 94 4B A4 A7 E0 0A 48 A1 88 21 67 B4 2A 68 43 43 CE C2 75 39 D6 C1 F7 D0 95 61 4F 48 22 E4 5C 3B 63 2A 48 54 74 7D 41 62 FD 4E 7F 20 83 A8 78 11 8D 10 B9 88 3C CC C1 47 34 E6 D2 BA B2 AE 93 26 3A A8 96 2A 0E 38 43 14 20 9A 63 0E 21 A2 48 8D 48 A7 FA B5 35 CD 62 4C 6B 81 03 2E 11 25 88 52 44 19 A2 5C 22 22 4C 20 85 22 54 98 10 D5 ED 51 64 0A F3 78 4E 87 78 D0 99 E5 9B 49 62 08 53 06 51 61 5E 5A 17 2A 03 D9 CE 10 A6 FC FD 90 2E B4 23 8C 72 11 79 98 83 8F 68 AC 72 A7 69 5D 59 E7 20 13 1D A4 85 09 69 CD 30 87 00 D1 1C 73 08 11 45 5C 5A D7 56 69 8F 31 AD 05 0E B8 44 94 20 4A 11 65 88 72 89 A8 30 6D 2B A3 50 31 86 30 D1 F0 A0 5D 99 C2 FC C9 DF FF 88 AB 71 75 02 20 1E 7A 60 0B B3 22 86 30 65 90 21 4C EB 27 AB 03 D9 93 21 4C F9 0B 1C 2A 4C 40 2E 36 F4 24 22 39 F8 88 C6 2A 77 43 98 D6 59 C8 44 07 69 61 42 0E 33 CC 21 40 34 C7 1C 42 44 11 97 D6 B5 55 C8 63 4C 6B 81 03 2E 11 25 88 52 44 19 A2 5C 22 2A 4C FB D7 34 85 8A 31 84 89 BF B9 A1 5D 99 C2 FC 49 F7 83 0A 13 7E 74 73 51 11 43 98 32 88 DC D9 25 A3 0C 19 56 51 04 8D 30 CA 45 E4 E1 88 3E A2 B1 44 E2 49 6F E4 8C D9 3A 9F 98 E8 20 2D 43 48 6B 86 39 04 88 E6 98 43 88 28 E2 D2 BA B6 CA 76 8C 69 2D 70 C0 25 A2 04 51 8A 28 43 94 4B 44 65 68 7B 20 85 8A 31 64 88 4E 09 ED CA 94 21 EF 94 BC ED 09 09 54 83 D2 EE D0 4F 9A BC A8 88 A1 41 19 64 14 47 EB F4 7E 20 DB 19 AA 94 BF 66 A1 C5 11 90 8B 0D 3D CC C1 47 34 96 C8 50 65 DB 3A 99 98 E8 20 AD 4A C8 61 86 39 04 88 E6 98 43 88 28 E2 D2 BA B6 6A 76 8C 69 2D 70 C0 25 A2 04 51 8A 28 43 94 4B 44 55 69 BB 25 85 8A 31 54 89 9E 0A ED CA 54 25 EF A9 FC B8 2A A5 E5 41 54 59 11 43 95 32 C8 50 A5 75 3D 39 10 0B 4D 87 B5 4A 72 85 8E 68 84 C8 45 E4 49 64 1C B2 21 AD B1 8C 32 55 69 9D 49 4C 74 90 56 25 64 3A C3 1C 02 44 73 4C 2B 44 14 71 69 5D 5B 25 3C C6 B4 16 38 E0 12 51 82 28 45 94 21 CA 25 A2 AA B4 AD 95 42 C5 18 AA 44 03 86 76 65 AA F2 27 0D 18 5A 2E AB 15 77 72 63 6A B9 40 54 2D 26 E9 AB 6F 19 64 08 D3 BA 9E 1C C8 76 46 B9 94 56 0E 2D 97 80 5C 6C E8 61 0E 3E A2 B1 44 E4 71 AD 13 44 53 EC 7E 86 28 40 34 C7 BE 42 44 91 46 E4 67 57 E7 56 D1 8E B1 E1 02 47 5C 22 4A 10 A5 88 32 44 B9 44 54 87 B6 B9 52 A8 18 43 87 68 C1 D0 AE 4C 1D 1E CF 82 29 AB 8B 79 93 AA 24 46 81 94 41 86 0E AD CB C7 81 6C 67 E8 B0 6A 67 9C 4C 02 72 B1 A1 87 39 F8 1A E9 EF BB 6D 9D 3A 8C D5 D6 D0 05 98 B6 55 C9 27 3A 48 97 4C C8 6A 86 59 05 88 E6 98 68 88 28 52 23 D2 FD E7 9C 5B A5 3C D6 51 2A AF 05 8E B8 44 94 20 4A 11 65 88 72 89 A8 56 AB 1D A1 49 A1 62 0C AD DA 51 62 61 88 20 53 AB 3F 69 F0 D0 9A 59 AD FB D3 9A 59 11 43 AB 32 C8 D0 AA 75 45 39 28 9F 5F 56 1A 8F F4 60 0E 68 84 51 2E 22 4F 22 E3 60 0E 69 8D 65 54 B9 9B EA 85 D4 B6 55 CA 27 3A 48 2B 13 D2 9A 61 0E 01 A2 39 A6 15 22 8A 70 C4 18 D1 02 BB 5F 22 4A 10 A5 88 32 44 B9 44 54 86 B6 C3 53 A8 18 43 86 E8 03 D1 AE 4C 19 1E CF 07 12 B7 D4 D8 25 B3 F6 81 F4 A1 5B 06 D1 AB 6D F9 4B 11 2A 3A 40 23 D9 3B 91 A6 8B C8 93 C8 10 1D 24 31 56 99 1A A2 B3 1D 1D 1D A4 45 07 69 CD 30 87 00 D1 1C D3 0A 35 A2 C7 69 DB D2 51 49 D0 C2 ED 9C DB 9E 0E A6 BA C0 24 96 88 12 44 29 A2 0C 51 2E 11 95 A6 ED D6 14 2A C6 90 26 7A 3A B4 2B 43 9A E5 B3 C8 8F 74 43 C5 A1 2B C3 D3 91 84 56 48 15 64 54 48 6B B1 63 20 83 68 85 44 34 42 E4 22 F2 30 07 1F D1 98 4B AB 6D 7B 3A 3A A8 16 2B 0E 38 43 14 20 9A 63 0E 21 A2 88 4B CB 39 B7 4D 1D CC 6B 81 23 2E 11 25 88 52 44 19 A2 5C 22 A2 4C 20 85 22 54 99 10 D5 ED 51 64 2A F3 78 A6 4E 13 4C 1D 49 0C 65 72 A6 8E 6D 95 CB 76 86 32 D1 D4 C1 28 17 91 87 39 F8 88 C6 2A 77 BD 1B 27 88 A6 D8 FD 0C 51 80 68 8E 23 86 88 22 3D 22 AD A3 B6 87 83 79 2D 70 C4 25 A2 04 51 8A 28 43 94 4B 44 75 08 1E 8E 8A 31 74 88 1E 0E ED CA D4 E1 F1 3C 1C F1 0B 01 EB E0 2D 89 A1 43 CE C3 B1 5D 6F D9 CE D0 21 7A 38 18 E5 22 F2 30 07 1F D1 58 E5 6E 1C CE 6D 73 51 07 E9 0A 89 1E 0E E6 10 20 9A 63 0E 21 A2 88 4B CB 39 B7 4D 1C CC 6B 81 23 2E 11 25 88 52 44 19 A2 5C 22 AA 4C 30 71 54 8C A1 4C 34 71 68 57 A6 32 59 13 E7 CD 4F 18 26 17 37 E2 25 11 B6 30 D1 C3 51 41 E4 AC 52 22 43 86 E8 E1 60 94 8B C8 93 88 9E 55 22 1A EB 24 C8 A5 8C 6D 25 62 A6 53 1C 70 86 28 40 34 C7 1C 42 44 11 97 96 73 6E 9B 38 98 D7 02 47 5C 22 4A 10 A5 88 32 44 B9 44 54 86 60 E2 A8 18 43 86 68 E2 D0 AE 4C 19 1E C9 C4 69 82 89 23 89 51 1D 19 13 E7 CA BE F5 42 B6 33 64 89 26 0E 46 B9 88 3C CC C1 47 34 56 B9 D3 A3 74 9D A9 AE 85 68 D9 E0 88 01 A2 39 8E 18 22 8A 74 12 F4 28 6D 7B 36 98 EA 02 47 5C 22 4A 10 A5 88 32 44 B9 44 54 84 E0 D9 A8 18 43 84 E8 D9 D0 AE 4C 11 1E C9 B3 11 6F 18 B3 2B 21 7A 36 2A 88 1E 0B AF EC DB 2C 64 90 21 42 30 47 46 18 E5 22 F2 24 32 6A 23 7A 36 5C 5A 6D DB DF D6 41 5A 96 E8 D9 60 0E 01 A2 39 A6 15 22 8A 70 C4 18 D1 02 BB 5F 22 4A 10 A5 88 32 44 B9 44 54 83 E0 D0 A8 18 43 83 E8 D0 D0 AE 4C 0D F2 0E CD 0F FB 86 E2 BD 39 B6 06 D1 9E 51 41 86 06 6D 37 5B 06 19 1A 44 7B 06 A3 5C 44 9E 44 86 06 21 AD 31 97 D6 B5 ED 66 EB 20 AD 41 48 6B 86 39 04 88 E6 98 56 A8 11 AD 83 D6 44 88 54 12 C6 AA 8F 63 7B 38 98 EA 02 93 58 22 4A 10 A5 88 32 44 B9 44 54 A9 E0 E1 A8 18 43 A9 E8 E1 D0 AE 4C A5 96 0B E6 47 78 A7 6C B3 5A 78 27 8B E2 92 10 C5 F5 11 0D 10 0D 11 8D 10 B9 88 3C 89 0C 5D 56 69 11 34 56 99 D2 E9 72 6D FB D9 3A 48 EB 12 CD 19 CC 21 40 34 C7 B4 42 44 11 97 96 E3 D8 E6 0C E6 B5 C0 11 97 88 12 44 29 A2 0C 51 2E 11 15 21 98 33 2A C6 10 21 9A 33 B4 2B 53 84 BC 39 F3 E3 E5 12 9C 99 26 F8 16 7D 44 03 44 43 44 23 44 2E 22 0F 91 8F 68 8C 68 82 68 8A 68 86 28 90 88 68 7C 2E 91 B0 79 94 7C 43 8C 8A 74 14 A9 8E 8E 75 E4 88 B1 AF 05 26 B1 44 94 20 4A 11 65 88 72 89 A8 E4 C0 88 51 31 86 E4 D0 88 A1 5D 99 92 E3 8D 98 1F 97 1C B8 30 E5 0B C8 4C 53 AF 8F 68 80 68 88 68 84 C8 45 E4 21 F2 11 8D 11 4D 10 4D 11 CD 10 05 12 19 92 03 A7 29 C4 A8 48 22 C3 7F 74 1C DB 60 D1 51 B5 05 8D 49 2C 11 25 88 52 44 19 A2 5C 22 2A 39 30 58 54 8C 21 39 34 58 68 57 86 E4 C4 0B D9 8E 72 A8 3D F4 63 B8 2B 92 D0 43 2D A2 01 A2 21 A2 11 22 17 91 87 C8 47 34 46 34 41 34 45 34 43 14 48 44 25 A7 76 04 A9 72 18 15 E9 28 5A E5 6C E7 04 FB 5A 60 12 4B 44 09 A2 14 51 86 28 97 88 48 0E 48 A1 08 95 1C 44 75 7B 14 99 92 E3 9D 93 1F AE 72 E5 0B EC CD DB 73 24 31 24 07 F6 C7 00 A3 86 88 46 88 5C 44 1E 22 1F D1 18 D1 04 D1 14 D1 0C 51 20 91 21 B9 DA 1A AA 0F AC 18 15 A9 DD 45 4F 31 1D C7 36 49 74 54 5D E5 30 89 25 A2 04 51 8A 28 43 94 4B 44 25 07 26 89 8A 31 24 87 26 09 ED CA 94 1C 6F 92 FC B8 E4 C0 21 69 81 7F D0 47 34 40 34 44 34 42 E4 22 F2 10 F9 88 C6 88 26 88 A6 88 66 88 02 44 73 44 21 A2 08 51 8C 68 81 68 89 28 41 94 4A A4 95 93 01 C9 81 AC 80 14 8A 18 EA 42 A3 83 36 34 D5 75 BC 5F AB B4 C0 E9 90 44 6F 64 1F C8 00 C8 10 C8 08 88 0B C4 03 E2 03 19 03 99 00 99 02 99 01 09 80 CC 81 84 40 22 20 31 90 05 90 25 90 04 48 0A 24 03 92 03 59 01 29 14 31 C4 84 76 05 6D 68 8A E9 48 76 45 AB 5A 9D 6E B4 C5 0A 61 7D 57 5D C3 5A 3F EA D5 51 E4 77 CD 35 23 4E 9A 64 4E 79 A3 8E E8 6D 23 1F 83 33 68 89 53 96 C3 73 70 FE 23 ED 4E 93 E1 DF CC F7 E7 0C 55 33 71 17 85 3A 9C 8C 14 13 C5 5A 31 57 0D 29 5E D1 40 92 B5 D6 19 BC 3A 31 DD 9B 5F 33 F2 00 6B C5 AE 1B 65 9A 27 B7 6B 71 37 E3 A9 78 68 CF B8 CE B6 D9 BC 68 DA AB 2B 4C AE 53 86 CD 18 16 30 6C CE 6C 67 28 99 9E CC 11 90 18 C8 02 C8 12 48 02 24 05 92 01 C9 81 AC 80 14 8A 18 82 46 EB 83 36 34 05 7D 24 EB 43 BC DE BC 3C DD 6B B4 F5 F3 49 7A 35 23 CF 96 53 8C 3C 50 7C 20 99 43 D6 1E 86 35 D3 FD 8D 14 23 AF 52 77 EB FE B4 C2 3C 15 27 2E AB 94 86 FD BA AD 56 E7 98 19 77 C2 8C 3B 65 D8 8C 61 01 C3 E6 4C CE A1 64 54 67 B6 51 10 43 CC 02 C8 12 48 02 24 05 92 01 C9 81 AC 80 14 8A 18 3A 43 7B 83 36 34 75 76 24 7B 43 9C 59 1F 74 76 25 8A 76 5D 38 1D 6B 79 BE 57 47 D1 03 B3 6A A9 1F B6 3D 90 71 65 E1 AC 9E 1C 26 08 FF E8 B0 A1 0A 35 8A A5 F4 1D 8C 62 A9 86 31 8A A5 B5 42 E6 A9 04 DB B4 58 CA 96 6D 5A 2C E5 08 75 82 E3 6F 25 38 61 12 9C 32 6C C6 B0 80 61 73 C5 C8 C6 85 92 51 E5 DA C6 41 0C 31 0B 20 4B 20 09 90 14 48 06 24 07 B2 02 52 28 62 28 17 ED 0E DA D0 54 2E 6F 77 F0 2F 95 7F FB 73 87 5B D5 62 77 E3 8A 3C 42 AC 66 3A D7 7E CD C8 63 C4 24 73 C4 6D 67 AA C0 0D 6B A6 1F 24 3B 52 4C 9C 1B EB 83 B9 1C B7 AD 99 A7 C6 68 6B 2F C6 AF 99 FE B2 C7 CC B8 13 66 DC 29 C3 66 0C 0B 18 36 67 72 0E 25 D3 7B 2A 02 12 03 59 00 59 02 49 80 A4 40 32 20 39 90 15 90 42 12 B1 9A AC 76 7D B7 AB 98 46 3D 8A 4C D9 D9 06 C7 F7 1F F1 D4 6D E1 4A 36 45 55 DF 67 AF 9F 37 9B FD 60 BD 5F 7F FC F0 B4 D9 7D DA F4 37 8F 8F AF 27 B7 DB 2F CF A2 98 8A E2 4D F0 C9 6E 73 7F 73 DA 6D 5D 76 4A 0D 8A 8D A8 1B 54 9F 88 52 D9 29 15 86 9F B8 E2 93 52 3F F8 89 A8 5E 9D F2 1B C6 4F BA AD A6 18 E7 50 9A 61 9C A6 18 87 FB 44 9C 98 8A 71 B8 4F C4 B9 A4 18 87 FB 64 EA B4 3B E5 DB 21 98 DC C4 27 E5 FB 2A F0 13 F1 CE A5 4E BF 7A 33 85 95 DB 58 7C 52 BE 67 08 DB 88 97 0D 8A 71 D8 DC C4 27 E5 EB E2 B8 71 C4 3E 60 3F 11 6F 00 17 BD 1D 66 BB 95 41 AF DD E9 73 39 C7 ED 4E CC 6E A5 D8 48 8E 8B 37 D6 8A 11 0E F3 1F B6 B1 21 F2 E5 3E E9 89 AC CA 57 58 73 FB F2 5C B4 E1 3E E9 3A 9D 01 B7 BF 86 4E 67 CE 8D 31 14 5F E4 61 DE 5B 59 0D AF 3A 31 D7 CF D2 E9 14 87 CA 66 C5 2F 2F 3A C5 A1 92 D9 BC D5 29 B8 FE 17 4E A7 7C 85 03 6E 5A D2 EE 14 87 39 6D 75 94 5C 77 0A 6E C7 26 62 27 95 2F 26 67 7A 72 44 AE EC 4E 12 2F E3 15 9F 70 9B 91 38 62 3B D8 2F 63 E9 88 2D 61 C5 23 DE 3F D9 29 DF 09 C8 08 AE 71 D9 29 9F F3 C9 4C EE 76 C7 E7 B6 66 20 B6 A6 7C 7B 38 D3 42 E4 5C BE A5 96 FB 44 4C 53 76 FC 81 98 3E E5 EB BC B8 36 22 01 76 32 4E C5 FC E5 32 13 EF 12 10 DF 3C B7 CF C4 CB DE C4 F6 73 A3 F4 C4 94 2F 5F A1 83 E3 8B 17 15 89 36 FC 44 B9 14 13 85 EB 2D 16 FB A6 7C 4B 39 37 B1 1B 62 1C 6E DF 88 17 36 8B 36 87 4F CE 74 41 FE F8 E1 65 FD 69 33 5B EF 3E 3D 3C BF 9E 3C 6E EE 45 71 3E 7F 5F 3E 5D 77 F7 F0 E9 73 FD 8F FD F6 A5 E4 57 82 FF B6 DD EF B7 4F EA 5F 9F 37 EB BB CD AE FC 97 38 4B BE DF 6E F7 EA 1F 22 B5 B2 E7 C5 66 FF E5 E5 E4 F5 76 FD B8 11 55 5F 1C 38 EE 1F F6 CB 6D F6 70 B7 FF 2C 1A 9D 9E 6C 77 0F 9B E7 FD 7A FF B0 7D BE 39 7D 5C 3F DF 89 D0 97 8D 18 BD F3 70 77 73 BA F3 EF 0E 1B 79 B7 5B 7F 7D 78 FE A4 E9 61 E7 9F 7D DD EE 7E 3F 1C 66 3E FE 2F 00 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 C1 17 10 BE 4E 07 00 00 C6 20 00 00 13 00 00 00 78 6C 2F 74 68 65 6D 65 2F 74 68 65 6D 65 31 2E 78 6D 6C EC 59 CD 8B 1B 37 14 BF 17 FA 3F 0C 73 77 FC 35 E3 8F 25 DE E0 CF 6C 93 DD 24 64 9D 94 1C B5 B6 EC 51 56 33 32 92 BC 1B 13 02 25 39 F5 52 28 A4 A5 97 42 6F 3D 94 D2 40 03 0D BD F4 8F 09 24 B4 E9 1F D1 27 CD D8 23 AD E5 24 9B 6C 4A 5A 76 0D 8B 47 FE BD A7 A7 F7 9E 7E 7A F3 74 F1 D2 BD 98 7A 47 98 0B C2 92 96 5F BE 50 F2 3D 9C 8C D8 98 24 D3 96 7F 6B 38 28 34 7C 4F 48 94 8C 11 65 09 6E F9 0B 2C FC 4B DB 9F 7E 72 11 6D C9 08 C7 D8 03 F9 44 6C A1 96 1F 49 39 DB 2A 16 C5 08 86 91 B8 C0 66 38 81 DF 26 8C C7 48 C2 23 9F 16 C7 1C 1D 83 DE 98 16 2B A5 52 AD 18 23 92 F8 5E 82 62 50 7B 7D 32 21 23 EC 0D 95 4A 7F 7B A9 BC 4F E1 31 91 42 0D 8C 28 DF 57 AA B1 25 A1 B1 E3 C3 B2 42 88 85 E8 52 EE 1D 21 DA F2 61 9E 31 3B 1E E2 7B D2 F7 28 12 12 7E 68 F9 25 FD E7 17 B7 2F 16 D1 56 26 44 E5 06 59 43 6E A0 FF 32 B9 4C 60 7C 58 D1 73 F2 E9 C1 6A D2 20 08 83 5A 7B A5 5F 03 A8 5C C7 F5 EB FD 5A BF B6 D2 A7 01 68 34 82 95 A6 B6 D8 3A EB 95 6E 90 61 0D 50 FA D5 A1 BB 57 EF 55 CB 16 DE D0 5F 5D B3 B9 1D AA 8F 85 D7 A0 54 7F B0 86 1F 0C BA E0 45 0B AF 41 29 3E 5C C3 87 9D 66 A7 67 EB D7 A0 14 5F 5B C3 D7 4B ED 5E 50 B7 F4 6B 50 44 49 72 B8 86 2E 85 B5 6A 77 B9 DA 15 64 C2 E8 8E 13 DE 0C 83 41 BD 92 29 CF 51 90 0D AB EC 52 53 4C 58 22 37 E5 5A 8C EE 32 3E 00 80 02 52 24 49 E2 C9 C5 0C 4F D0 08 B2 B8 8B 28 39 E0 C4 DB 25 D3 08 12 6F 86 12 26 60 B8 54 29 0D 4A 55 F8 AF 3E 81 FE A6 23 8A B6 30 32 A4 95 5D 60 89 58 1B 52 F6 78 62 C4 C9 4C B6 FC 2B A0 D5 37 20 2F 9E 3D 7B FE F0 E9 F3 87 BF 3D 7F F4 E8 F9 C3 5F B2 B9 B5 2A 4B 6E 07 25 53 53 EE D5 8F 5F FF FD FD 17 DE 5F BF FE F0 EA F1 37 E9 D4 27 F1 C2 C4 BF FC F9 CB 97 BF FF F1 3A F5 B0 E2 DC 15 2F BE 7D F2 F2 E9 93 17 DF 7D F5 E7 4F 8F 1D DA DB 1C 1D 98 F0 21 89 B1 F0 AE E1 63 EF 26 8B 61 81 0E FB F1 01 3F 9D C4 30 42 C4 92 40 11 E8 76 A8 EE CB C8 02 5E 5B 20 EA C2 75 B0 ED C2 DB 1C 58 C6 05 BC 3C BF 6B D9 BA 1F F1 B9 24 8E 99 AF 46 B1 05 DC 63 8C 76 18 77 3A E0 AA 9A CB F0 F0 70 9E 4C DD 93 F3 B9 89 BB 89 D0 91 6B EE 2E 4A AC 00 F7 E7 33 A0 57 E2 52 D9 8D B0 65 E6 0D 8A 12 89 A6 38 C1 D2 53 BF B1 43 8C 1D AB BB 43 88 E5 D7 3D 32 E2 4C B0 89 F4 EE 10 AF 83 88 D3 25 43 72 60 25 52 2E B4 43 62 88 CB C2 65 20 84 DA F2 CD DE 6D AF C3 A8 6B D5 3D 7C 64 23 61 5B 20 EA 30 7E 88 A9 E5 C6 CB 68 2E 51 EC 52 39 44 31 35 1D BE 8B 64 E4 32 72 7F C1 47 26 AE 2F 24 44 7A 8A 29 F3 FA 63 2C 84 4B E6 3A 87 F5 1A 41 BF 0A 0C E3 0E FB 1E 5D C4 36 92 4B 72 E8 D2 B9 8B 18 33 91 3D 76 D8 8D 50 3C 73 DA 4C 92 C8 C4 7E 26 0E 21 45 91 77 83 49 17 7C 8F D9 3B 44 3D 43 1C 50 B2 31 DC B7 09 B6 C2 FD 66 22 B8 05 E4 6A 9A 94 27 88 FA 65 CE 1D B1 BC 8C 99 BD 1F 17 74 82 B0 8B 65 DA 3C B6 D8 B5 CD 89 33 3B 3A F3 A9 95 DA BB 18 53 74 8C C6 18 7B B7 3E 73 58 D0 61 33 CB E7 B9 D1 57 22 60 95 1D EC 4A AC 2B C8 CE 55 F5 9C 60 01 65 92 AA 6B D6 29 72 97 08 2B 65 F7 F1 94 6D B0 67 6F 71 82 78 16 28 89 11 DF A4 F9 1A 44 DD 4A 5D 38 E5 9C 54 7A 9D 8E 0E 4D E0 35 02 E5 1F E4 8B D3 29 D7 05 E8 30 92 BB BF 49 EB 8D 08 59 67 97 7A 16 EE 7C 5D 70 2B 7E 6F B3 C7 60 5F DE 3D ED BE 04 19 7C 6A 19 20 F6 B7 F6 CD 10 51 6B 82 3C 61 86 08 0A 0C 17 DD 82 88 15 FE 5C 44 9D AB 5A 6C EE 94 9B D8 9B 36 0F 03 14 46 56 BD 13 93 E4 8D C5 CF 89 B2 27 FC 77 CA 1E 77 01 73 06 05 8F 5B F1 FB 94 3A 9B 28 65 E7 44 81 B3 09 F7 1F 2C 6B 7A 68 9E DC C0 70 92 AC 73 D6 79 55 73 5E D5 F8 FF FB AA 66 D3 5E 3E AF 65 CE 6B 99 F3 5A C6 F5 F6 F5 41 6A 99 BC 7C 81 CA 26 EF F2 E8 9E 4F BC B1 E5 33 21 94 EE CB 05 C5 BB 42 77 7D 04 BC D1 8C 07 30 A8 DB 51 BA 27 B9 6A 01 CE 22 F8 9A 35 98 2C DC 94 23 2D E3 71 26 3F 27 32 DA 8F D0 0C 5A 43 65 DD C0 9C 8A 4C F5 54 78 33 26 A0 63 A4 87 75 2B 15 9F D0 AD FB 4E F3 78 8F 8D D3 4E 67 B9 AC BA 9A A9 0B 05 92 F9 78 29 5C 8D 43 97 4A A6 E8 5A 3D EF DE AD D4 EB 7E E8 54 77 59 97 06 28 D9 D3 18 61 4C 66 1B 51 75 18 51 5F 0E 42 14 5E 67 84 5E D9 99 58 D1 74 58 D1 50 EA 97 A1 5A 46 71 E5 0A 30 6D 15 15 78 E5 F6 E0 45 BD E5 87 41 DA 41 86 66 1C 94 E7 63 15 A7 B4 99 BC 8C AE 0A CE 99 46 7A 93 33 A9 99 01 50 62 2F 33 20 8F 74 53 D9 BA 71 79 6A 75 69 AA BD 45 A4 2D 23 8C 74 B3 8D 30 D2 30 82 17 E1 2C 3B CD 96 FB 59 C6 BA 99 87 D4 32 4F B9 62 B9 1B 72 33 EA 8D 0F 11 6B 45 22 27 B8 81 26 26 53 D0 C4 3B 6E F9 B5 6A 08 B7 2A 23 34 6B F9 13 E8 18 C3 D7 78 06 B9 23 D4 5B 17 A2 53 B8 76 19 49 9E 6E F8 77 61 96 19 17 B2 87 44 94 3A 5C 93 4E CA 06 31 91 98 7B 94 C4 2D 5F 2D 7F 95 0D 34 D1 1C A2 6D 2B 57 80 10 3E 5A E3 9A 40 2B 1F 9B 71 10 74 3B C8 78 32 C1 23 69 86 DD 18 51 9E 4E 1F 81 E1 53 AE 70 FE AA C5 DF 1D AC 24 D9 1C C2 BD 1F 8D 8F BD 03 3A E7 37 11 A4 58 58 2F 2B 07 8E 89 80 8B 83 72 EA CD 31 81 9B B0 15 91 E5 F9 77 E2 60 CA 68 D7 BC 8A D2 39 94 8E 23 3A 8B 50 76 A2 98 64 9E C2 35 89 AE CC D1 4F 2B 1F 18 4F D9 9A C1 A1 EB 2E 3C 98 AA 03 F6 BD 4F DD 37 1F D5 CA 73 06 69 E6 67 A6 C5 2A EA D4 74 93 E9 87 3B E4 0D AB F2 43 D4 B2 2A A5 6E FD 4E 2D 72 AE 6B 2E B9 0E 12 D5 79 4A BC E1 D4 7D 8B 03 C1 30 2D 9F CC 32 4D 59 BC 4E C3 8A B3 B3 51 DB B4 33 2C 08 0C 4F D4 36 F8 6D 75 46 38 3D F1 AE 27 3F C8 9D CC 5A 75 40 2C EB 4A 9D F8 FA CA DC BC D5 66 07 77 81 3C 7A 70 7F 38 A7 52 E8 50 42 6F 97 23 28 FA D2 1B C8 94 36 60 8B DC 93 59 8D 08 DF BC 39 27 2D FF 7E 29 6C 07 DD 4A D8 2D 94 1A 61 BF 10 54 83 52 A1 11 B6 AB 85 76 18 56 CB FD B0 5C EA 75 2A 0F E0 60 91 51 5C 0E D3 EB FA 01 5C 61 D0 45 76 69 AF C7 D7 2E EE E3 E5 2D CD 85 11 8B 8B 4C 5F CC 17 B5 E1 FA E2 BE 5C D9 7C 71 EF 11 20 9D FB B5 CA A0 59 6D 76 6A 85 66 B5 3D 28 04 BD 4E A3 D0 EC D6 3A 85 5E AD 5B EF 0D 7A DD B0 D1 1C 3C F0 BD 23 0D 0E DA D5 6E 50 EB 37 0A B5 72 B7 5B 08 6A 25 65 7E A3 59 A8 07 95 4A 3B A8 B7 1B FD A0 FD 20 2B 63 60 E5 29 7D 64 BE 00 F7 6A BB B6 FF 01 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 9E AC D5 76 F3 0C 00 00 ED D1 00 00 0D 00 00 00 78 6C 2F 73 74 79 6C 65 73 2E 78 6D 6C DC 5D 5B 6F E2 48 16 7E 5F 69 FE 03 F2 3B 31 36 90 09 11 30 EA 74 1A A9 A5 D9 D1 48 E9 95 E6 D5 01 13 AC F1 85 B1 4D 2F 99 D5 FE F7 AD 32 37 BB CD C1 75 3B 55 95 CD 43 37 71 A0 EA AB 53 E7 7E 4E 15 D3 5F F6 49 DC FB 1E E6 45 94 A5 33 C7 BB 1B 38 BD 30 5D 66 AB 28 7D 9B 39 FF FA B6 E8 3F 38 BD A2 0C D2 55 10 67 69 38 73 DE C3 C2 F9 65 FE D3 3F A6 45 F9 1E 87 2F 9B 30 2C 7B 64 88 B4 98 39 9B B2 DC 3E BA 6E B1 DC 84 49 50 DC 65 DB 30 25 7F 59 67 79 12 94 E4 D7 FC CD 2D B6 79 18 AC 0A FA A1 24 76 FD C1 E0 DE 4D 82 28 75 0E 23 3C 26 4B 96 41 92 20 FF 73 B7 ED 2F B3 64 1B 94 D1 6B 14 47 E5 7B 35 96 D3 4B 96 8F 5F DF D2 2C 0F 5E 63 02 75 EF 8D 82 65 6F EF DD E7 7E 6F 9F 9F 26 A9 9E B6 E6 49 A2 65 9E 15 D9 BA BC 23 E3 BA D9 7A 1D 2D C3 36 DC 89 3B 71 83 E5 65 24 32 B2 D8 48 DE D8 1D F8 8D B5 EF 73 C1 91 46 6E 1E 7E 8F E8 F6 39 F3 E9 3A 4B CB A2 B7 CC 76 69 39 73 FC 11 41 4A 69 F0 F8 67 9A FD 3B 5D D0 BF 91 2D 3E BE 6D 3E 2D FE EE 7D 0F 62 F2 C4 73 DC F9 74 99 C5 59 DE 2B C9 DE 11 D2 55 4F D2 20 09 0F EF F8 1C C4 D1 6B 1E D1 B7 AD 83 24 8A DF 0F 8F 7D FA A0 DA EE E3 FB 92 88 10 9F 3E 74 29 90 03 9C F9 F4 95 BE 4B D3 5C E7 79 2A 6C 78 6B 3A CF 33 C2 A5 DD 79 9E 7B 8D 7B 84 BC A6 3A 3F F8 0F FA D6 E5 6B A4 A1 3F D0 B8 AE 6A AE 8B B4 7E CA A3 20 BE 2A AB B0 58 22 8B 4B 43 05 54 73 71 C3 6D 4A B6 DC 6A 6B 0A 2F 7F 7B 9D 39 0B F2 33 20 3F 94 68 2A B5 5E 54 D7 7C 72 CB D6 C9 50 1E B2 A0 34 95 9A 2A 82 9F 46 1D D6 D0 A3 EE EE 69 42 5F A7 F0 54 BA 19 C5 32 D7 56 71 83 6C 02 CA 65 A8 91 3E 68 E6 A4 52 9D 05 71 69 A2 38 3E 7B 58 43 EA 4B 91 07 F3 29 71 45 CB 30 4F 17 E4 97 DE F1 F5 B7 F7 2D F1 A4 52 E2 35 1F 3C A2 EA 7D 1D EF 7E CB 83 77 CF 1F F3 7D 60 70 DF F8 84 5B 41 24 6E 57 96 AF 88 5F 7F F2 06 C7 D4 F1 3B 3C 9B 4F E3 70 5D 12 DD 94 47 6F 1B FA 7F 99 6D C9 BF AF 59 59 12 E7 77 3E 5D 45 C1 5B 96 06 31 75 E4 4E 9F A8 7F 92 04 04 C4 F7 9F 39 49 B8 8A 76 09 19 F6 E0 67 45 E9 2A DC 87 AB 99 73 5F 31 A9 4B 27 69 CC C1 F8 39 82 86 03 4C 73 19 A8 53 30 0E 5E 51 B5 22 2A 0A 1A 51 EA F3 EC F0 89 A8 DC 2B D6 C1 46 E7 39 18 D1 1D F8 BA 9B AD 01 81 40 9A 45 9C B4 C8 80 CA 0D 89 C8 21 99 BE F0 B6 0C 3B F1 4E A1 78 C5 A7 E1 6E C2 A8 69 2F A6 F7 B7 A5 FE E6 F0 35 25 C7 34 3C 2B 13 F3 AE ED A0 FD 99 3E A5 42 2F B3 D2 84 C1 0A FD A8 A3 18 99 0A 61 A9 62 34 17 51 95 AC 1C FB 83 4D 57 CA 61 1D 6A F2 26 44 5E 55 CC 29 79 2D CB 80 02 A6 C3 ED B9 A2 22 51 E8 CF 6D 40 78 39 1F 91 6B 8C 42 61 24 5C CB 7D C5 E1 2E 51 34 D6 E8 31 2B D8 8A 53 51 FC 1F 88 42 87 16 E2 F5 17 44 63 39 5E 27 C7 18 6C 41 0E E9 C0 CB E7 13 F1 3A 0A 56 98 3F 64 D0 8C FA 4F 0F A1 99 4D 3B 3F 6A A5 D6 94 DF 68 7C 68 56 E2 DE 16 A3 A1 17 BB 2A FD 48 56 8B 57 0D 00 21 88 62 FD CF 1B C2 A3 71 12 43 C0 2A 46 41 4B 34 0D 1A A1 6D F1 AA 19 E9 2C 96 00 43 D7 D7 1F 94 4B 38 73 36 EC CE D8 C7 B2 B6 4C 68 AF 64 0C 0D 2B 53 A4 8C 1B 23 2F 63 65 44 B9 15 81 25 BB C7 EF 76 88 27 B2 3E 8A 15 17 29 52 34 33 E2 FC 2A 47 11 6D 7E 4C 6F 6B 08 4D 6C 41 8E 58 24 40 8F 24 24 68 78 AC B5 93 26 83 65 18 C7 2F B4 C6 FE C7 FA 5C BF A7 E5 FB FD BA 97 EE 92 45 52 7E 25 55 76 D2 A9 4B BB 2C 4F 2F 49 7B C1 F1 E5 A1 54 7F F8 85 70 91 5B 1F ED 30 76 6D 58 7F 42 9B 18 F8 07 EE ED D7 E7 19 78 3F ED 93 BE D4 E3 A7 7B C1 76 1B BF 3F 55 AD 0A C7 B6 53 DE D1 86 4A 47 A3 3D B3 EA B0 91 3D 52 38 DA 58 E9 68 3F 2B 1D 8D B4 89 EB A1 DB 81 63 3E C5 D1 5B 9A 84 B4 CD B9 6A E6 E3 66 C1 89 5A 1E 94 DE 68 22 13 17 B9 6D 6D 0D ED 2B A2 2B 15 13 97 C6 D8 2D 26 12 18 9B 00 A1 4D 46 2D 7D 43 3A BD 7E E0 02 DA 74 CE 09 1C 1A 9C B4 0A CB 0F 4E 95 E6 35 E4 FE 0D 2D A2 86 E3 6E 28 16 35 13 DC C3 1C CD 37 01 39 0F 70 9D 48 AD 25 08 EC 2E E1 90 AB 1B D0 12 20 11 CE 91 06 7E 85 50 F3 29 39 FC 70 D0 35 BD 4D 96 47 7F 13 8E A6 A7 26 68 79 D0 E9 D1 9E B7 B3 12 72 F7 6B D8 94 CA 53 95 73 17 21 4A 33 EF 22 DF 7C D2 1B AB 66 3A CD AB 63 56 78 9C 9B 07 E9 29 A4 E5 79 D0 7C CC 62 A9 85 59 44 4C 95 B4 18 10 17 1A F2 2E 3C 96 C1 FF DA 65 65 F8 7B 1E AE A3 FD C5 12 42 DA 8D 57 01 51 FF FE A6 D6 C1 D0 02 6C 2A 71 58 75 31 0B A9 44 A2 C7 9B CE A4 2A 72 71 E8 EB 9A 44 34 FC A7 B6 44 A8 F3 CD 5A 4E 88 08 C3 13 15 A1 DF C2 2A 31 DF 98 BE 01 B4 9D AD 38 40 80 E4 D2 14 57 E3 00 7A 62 11 1D 84 5E A5 69 23 A7 97 3A 93 15 2D A1 FF 6D 97 BC 86 F9 A2 3A 3F 7C 2D 86 E0 A3 1A B4 4C 15 4E AD F4 D8 9C DE 01 D9 E8 AB 22 AE 62 2D E6 6D 1A 3D 8E 5E 46 4B EA 64 93 9A 47 A7 15 81 24 7B A8 42 B4 A1 C1 47 2A C2 68 68 F0 B6 E4 01 26 46 58 6F 88 59 34 45 01 08 92 4F 09 B9 B0 ED 64 0A 7F 4E 42 9E C7 38 69 47 3C 81 AB 12 DE B1 18 C8 78 A9 31 30 37 EC 8B 05 AC 71 D3 2B ED F2 93 89 38 9F E8 DD 90 8D 8E 9C 13 07 BD 21 78 55 43 41 A7 92 23 38 E4 9C 3A 35 1C 30 BA 91 36 55 33 C3 8D 1C B7 9A 09 6E 24 AA F9 26 A8 19 E1 06 CB 74 F8 4D 0A 58 66 49 52 4E E1 E1 9E 8D 5B E1 15 04 B0 1D 65 34 14 A2 79 80 EA 94 1C 24 74 92 14 C4 D7 0A 8C 00 C1 74 36 5F B6 59 92 4C 20 8A 96 4B 7F 29 F4 71 24 3A 64 69 C1 A7 52 24 69 01 BA CF 68 39 0D E1 14 90 4C 80 27 65 6C 41 1A 75 68 26 21 EE A9 F2 F4 F6 A5 C9 84 11 C9 78 CF 38 BB D6 0E 85 00 07 5B 5B 72 B3 55 0A 33 9E 6E 65 0E 42 B4 D1 88 B9 9A AA 1A 11 D9 9C AB AE B6 36 BF A9 96 5F 60 34 2D 10 64 6D 9E 94 3A C8 1D EA 43 9D F3 A7 0E 72 87 7E B1 11 B2 36 0F 56 1D 95 B5 F9 B4 FC 90 AB EC F1 D5 AE 0D 63 4E 15 89 83 0D 29 31 26 7F 06 82 87 2F FD 4C F0 6A D9 8C 8E 4C A4 EA C8 94 09 1E B9 D0 0B E8 12 42 A8 4B 1E 4D 90 80 54 48 97 CB F8 63 30 7E 94 D0 56 77 48 AE 49 A5 0E 41 B6 D8 41 81 20 77 14 84 6C A4 B2 C5 0E 0A 44 65 8B AD 3D 94 9A B1 23 33 C3 D5 B1 88 94 A4 B2 22 53 D6 62 7A 13 89 32 BE 06 52 A4 ED E0 AB 31 49 82 A0 5D BF 57 CB 2A 3E 46 BC 2E 6E E9 41 98 18 89 0E F5 30 C7 18 39 2B 61 98 50 30 6F B1 47 02 41 B6 D8 BC 43 90 3B 38 D6 A4 47 E2 41 45 56 0F 23 2F 2D CC BF 20 4C AB A4 0C 44 69 95 66 85 14 2B AD 50 28 EF 40 65 4C 36 82 98 10 95 7D 57 35 02 34 40 88 92 21 8A C9 93 E9 91 97 74 28 40 3A D9 C8 4F 18 15 00 59 1E B7 91 9F 10 35 96 28 8F 5B C8 E2 ED 6E 4B F9 2A A0 2C 37 61 54 26 25 31 21 1A E9 2E 66 82 BA 94 3D 63 05 5C 8C C6 79 A9 1A 37 04 08 83 B9 99 52 D1 E0 9E 19 2B 71 83 88 8C 95 B8 41 44 C6 4A DC 20 22 63 2D 37 20 22 63 2D 37 20 22 0C DB 26 25 6B 43 EB F4 A3 7D A2 A6 B7 2B 89 BA F8 A7 BB 4F C0 70 A9 B5 6B D7 CF 13 E9 6A 0D 82 70 1A EB E7 02 63 15 AC C3 97 C2 01 9D 31 5B 07 F2 96 31 01 D4 9A B0 60 D2 9A 20 22 63 96 05 42 64 4E 8F 43 88 AC D3 E3 ED 6B 50 74 A9 47 C8 1D B0 CF CF 35 97 19 D4 19 2E 31 86 B9 3A 03 26 49 48 B4 AB C7 54 4E 57 6B D0 24 49 27 9A B9 B7 2E F7 6D 2C 28 00 CD 9B 75 7E 9B B9 34 0E 44 23 63 59 13 70 D3 8C 65 04 C0 1B D3 0C E6 03 41 4C 06 13 DE E0 E5 6F 18 0A 80 51 51 82 98 2C 4C 79 93 2F BF B5 4E 79 0F 0D 1A 14 A8 47 76 68 50 EE 20 4C E4 6B 97 8D ED 1D 88 C9 A0 2E 00 31 19 E4 A7 DA 3D 01 84 83 2E 37 76 62 EE 1D 7F D3 35 88 D2 AA E6 0B 10 25 A2 5E 15 68 B3 07 5A 04 4D B6 88 80 92 61 B0 7C 0E 62 32 68 B9 41 4C 06 0B C3 06 4E 6D 74 25 5E 21 48 06 FD 1B B0 6F BD 15 C4 9B E8 94 6E 37 04 9A 40 C1 79 19 0B 52 C3 76 3B EC 34 41 0B AD 77 3C 80 CC 69 45 07 FD AD 6B 00 F5 DD BA E1 F1 DD 14 24 C9 9C C6 7B BA BB 54 AC F1 43 EF A2 00 B5 1D C7 12 05 D8 E1 5E E8 BB CD 08 D2 0A 23 2B 0E F7 8C AC 38 F1 35 6C ED 96 11 F3 AD 95 16 E0 DD DA 6D 45 AD EC 9A 63 BE 1B 34 E1 3B 8B AC AB 73 FB C6 4E FC 43 E2 3D B4 43 B0 5A 91 1F A6 60 41 91 7B 5B BC 15 70 B4 F0 41 19 30 BF 80 C1 43 CA 51 0E 11 63 79 FE 2C 08 75 E7 AE 9E 94 44 A4 65 E7 15 59 10 26 83 85 5E 90 4E 06 33 33 20 26 83 F9 66 10 93 C9 DA 13 C4 4F 16 E6 9B 31 9B 19 14 E6 9B 11 29 C7 8F 12 AA 0A 8F 0C CA 01 98 4A B0 C3 51 6E ED 9F 09 77 5D 6B 26 01 2C 8A 9B 34 74 80 ED C5 74 11 3A 8D 6F 2D A3 D1 A8 84 F9 06 B3 C6 B4 0A 77 D5 4B 69 47 BF EA 1C 53 61 4A D1 6F 04 33 D5 68 05 56 46 0C 62 02 13 28 5A 63 1B 9A CF BE 7A B5 1B FE FD 10 6C 57 D5 42 F8 3A BC 5E 05 89 30 39 7C F8 99 44 39 7C 1D 1E BA 71 FA E1 DF 9C 21 47 3F FC 5B B0 E4 F0 E1 5F 23 2A 75 5E C2 A4 31 07 4F 4C 18 AC DD 83 3D 9C 06 2D 14 18 41 58 78 BD C3 C8 60 C6 01 A4 93 41 7E 02 F3 B7 36 5C 42 A7 35 CE 02 AF 37 B5 2A 2F 0A A2 B4 2A 2F 0A A2 44 E4 74 85 57 C5 22 F6 41 29 44 89 98 9F 51 88 D2 AE 2B 03 A1 00 CF AA 8C 1C 78 E1 9A 4D F7 2C 43 20 11 6D BE D4 35 CB 8D 6C 8C 55 CA 12 A2 A4 55 BA 12 02 89 78 60 40 DD 76 5B A5 CF 21 4A 5A A5 CE 21 90 56 69 73 08 A4 55 CA 1C 4C DF A9 B9 D2 B7 0A AE F9 25 05 44 A5 C6 E5 16 44 55 3B 0F 4C CC C8 E5 18 09 7E 0E 8F FF 70 70 03 20 7E 92 4C 12 A0 35 FD 8A D0 16 E3 A7 A1 3E 0A 05 F9 85 D9 38 4D D5 41 D6 C6 A7 35 C8 2C DF A6 6D 13 89 65 F0 E2 67 A3 DB 1D 5F 32 78 3B 4C 39 CA BD D8 32 DF AE DE CE 15 B1 7F 47 B9 BA 93 28 0D DB D4 0E 23 8C 43 62 FF 0E 7A D4 93 19 0D 32 B5 0D 90 36 32 81 39 7D EB 6E 8A A0 5F 39 A6 BC 36 2F 55 1F EA EA 7C D6 55 9F 84 F6 B0 8B 62 DA EA A7 50 0F C8 D0 64 5B 28 08 CA E4 BD 16 20 28 93 BD 4E 20 28 83 F5 AC 9A 53 D4 EC 75 C2 C0 24 1F 52 36 30 A2 7C 79 85 6A 8C 18 52 A0 18 23 46 D6 4A 31 44 B4 D3 3A 52 61 4F 93 1B D1 EC AA 42 8C 06 0D 05 A4 69 30 BF 82 46 F8 DE 78 83 B6 0B A2 13 CA 1D 9F 92 17 EA 19 3C D2 00 B2 13 62 35 43 94 9D 50 EE 1E 15 3E 96 65 42 10 15 AA 30 AB 4A 6A A0 FB 64 55 09 15 DC 71 AB 4A 41 20 4A AB 6A 41 20 4A DD 47 6D 98 6E 8A 6F B1 A1 9E 9B E2 DD 65 18 C7 7F AC 8B F9 94 BE 78 29 DF E3 B0 E8 2D B3 5D 5A CE 1C CF A9 3D ED A5 41 12 CE 9C DF B2 3C 09 E2 5A 5E E2 75 17 C5 65 94 7E 5D CD 9C 81 43 D5 EE 65 9C F9 74 B5 5F 9F 47 AB FE 5A 06 AF 71 D8 9C 85 58 81 55 B8 0E 76 71 F9 ED FC C7 99 73 79 FD CF 70 15 ED 12 42 9E E3 BB 7E 8F BE 67 65 35 C4 CC B9 BC FE 35 7A DB 94 E4 E8 05 41 10 EE CB 5F 8B B2 FA BF B7 CB A3 99 F3 9F 2F 4F 3F 4F 9E BF 2C FC FE C3 E0 E9 A1 3F 1A 86 E3 FE 64 FC F4 DC 1F 8F 3E 3F 3D 3F 2F 26 03 7F F0 F9 BF 64 4D 49 9C 16 8F 7B 6F 34 73 36 65 B9 7D 74 DD 62 B9 09 93 A0 B8 4B A2 65 9E 15 D9 BA BC 5B 66 89 9B AD D7 D1 32 74 8B 6D 1E 06 AB 62 13 86 65 12 BB FE 60 30 71 27 6E 12 44 29 21 1B 19 E4 B1 88 C9 BB F2 E3 62 8F E0 5F 2E CF 66 4E ED 97 03 FC 8A 7E 04 7E 1D FB C4 BF 1F 7C 1A 7B 83 FE 62 38 F0 FA A3 FB E0 A1 FF 70 3F 1C F7 17 63 CF 7F BE 1F 3D 7D 19 2F C6 35 EC 63 31 EC DE C0 F5 BC 0B F8 F1 63 19 25 61 1C A5 A7 BD 3A ED 50 FD 29 D9 24 F2 EB 8D 45 B8 A7 9D 70 0B CA 57 2F 94 52 F3 FF 01 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 B4 56 B5 DE 31 0B 00 00 9E 26 00 00 14 00 00 00 78 6C 2F 73 68 61 72 65 64 53 74 72 69 6E 67 73 2E 78 6D 6C D4 5A 5D 6F E2 58 12 7D 5F 69 FF 43 89 A7 6E 29 81 4C 67 A6 77 15 A5 19 19 30 89 B5 0E 66 6C 93 4C E6 CD 31 97 60 B5 F1 65 6C 93 4C E6 D7 EF A9 EB 6B 1B 6C 48 D3 99 EC 48 1B B5 3A C1 BE DF 55 75 EA D4 B9 5C FE FC C7 2A A6 27 91 66 91 4C BE 74 7E E8 9E 75 48 24 A1 9C 47 C9 E3 97 CE CC 1F 9F FE BB 43 59 1E 24 F3 20 96 89 F8 D2 79 11 59 E7 E7 FE 3F FF 71 99 65 39 A1 6F 92 7D E9 2C F3 7C 7D D1 EB 65 E1 52 AC 82 AC 2B D7 22 C1 9B 85 4C 57 41 8E 8F E9 63 2F 5B A7 22 98 67 4B 21 F2 55 DC FB 74 76 F6 B9 B7 0A A2 A4 43 A1 DC 24 39 E6 3D FF A9 43 9B 24 FA 7D 23 86 FA C9 A7 1F 3B FD CB 2C EA 5F E6 7D F3 97 99 35 BD 31 27 3E 79 53 73 68 8D AD 21 DD 98 C3 6B 63 62 0D 0D FB B2 97 F7 2F 7B DC B0 68 6C 3B C3 FF 38 33 BF E7 1B 57 F8 45 53 D7 19 9A A3 99 6B 36 DB 19 33 FF DA 71 AD DF 0C DF 72 26 17 E4 4C EC 7B 32 6F A6 B6 73 6F 9A 1E DD 5D 3B 74 6D DC 9A E4 CD 86 43 D3 F3 C6 33 1B EF 87 0E 1A 98 BE 39 22 FF DA 24 DF 35 AC 89 35 B9 E2 39 AE 5C E3 06 63 90 E1 93 6D 1A 9E 4F 06 FE 9E 4C 66 86 4D 03 C3 B3 3C 32 5C 93 CA 29 CD 51 73 31 23 F3 D4 9C 98 EE 95 F5 9B 1E AF 58 F3 45 B3 DD C4 F1 CD 0B 9A CA 67 91 D2 6A 03 0B 3C 08 8A 65 F8 55 CC 49 6E 72 92 09 E5 4B 41 38 EC 2C DB A4 82 B2 68 2E 7A 71 94 08 92 0B 12 BF 6F A2 F5 4A 24 79 B7 31 2A 9B F1 22 5B 07 21 CC CB 5D 45 FA 24 3A 7D FA D4 A5 89 CC A3 C5 8B 1A 33 58 2C 44 98 63 1E B1 5A C7 F2 45 88 0C 8F 83 5C BD 13 89 48 1F 5F 60 CB 24 4F 65 8C E9 65 28 E6 3C 7F 94 61 81 F0 24 4A C5 4A 3E 89 F9 91 33 9F 77 69 B8 14 E1 57 35 F8 B3 4C BF 52 00 F7 A1 5C C2 35 D5 BE 82 38 DE 5A C7 32 78 12 98 47 24 94 05 0B 11 BF D0 5A 66 51 0E 7F E6 53 49 E9 B8 DD FE D8 A5 5B 91 96 BB 85 0B A7 41 CE 2B D7 9B CA 78 09 14 25 E4 8C C7 3C 68 22 36 79 1A 60 AF 7A AA 23 77 F6 53 97 5C 75 14 6A 6B 79 F0 48 88 2C 65 41 5A A4 72 A5 9E CE A3 0C 93 26 38 ED 23 07 FD DC A5 BB 25 76 5F 19 98 10 88 E1 D7 4C B9 04 8F 0F 33 84 B1 08 52 F6 82 B5 90 EB 58 9C E0 24 61 B3 E8 4F 51 18 F0 3B 5D E3 5F 5D F2 37 69 C2 FE 56 4F CA 33 A9 89 09 91 CF 4E 80 33 A4 54 E6 38 47 B4 E3 B7 FA 54 65 D2 DC 58 DF 3D D2 FF 55 90 DC 93 E7 CC 5C 44 65 33 3A 7C 87 06 26 82 74 E2 59 23 D3 35 47 AD F0 01 36 A8 70 87 15 9B 7D 2D CF B1 F1 0E D1 3C 32 6F AD 3D 83 CF BC 16 84 38 40 01 97 34 E0 B4 56 63 DA E6 D0 77 19 A2 0E 37 A9 80 6D 32 BB 19 98 6E 73 51 15 78 D1 B7 DE 8F 0C BF B5 BC 1A 36 27 C6 4D F3 ED 81 A0 27 F3 D4 CB E5 9A 6C 19 2A BB 1D 17 3C CA A9 73 C1 BD BE 32 0E 1D E8 BD 05 E4 23 D3 1B BA D6 94 B1 F7 C8 29 68 FB 67 11 A5 2B 04 7A 26 02 40 52 97 3C 11 02 6E F0 60 11 64 39 22 21 28 42 8A 7D 2E D0 9E C8 2F A3 38 D6 40 C9 91 97 CB B7 CC CC F8 CA 88 CB D8 30 17 4F 51 08 94 95 2A 88 96 98 2E 16 14 06 49 22 15 30 E7 08 91 A3 31 6F 67 7B F3 28 45 F8 73 D8 20 62 17 B1 7C EE D2 6B 7B 2C 62 AD BD C3 37 EC 0F 18 FB B7 EC 70 FC 3D 86 5A A6 72 F3 B8 54 0B 03 A3 E0 93 5F CA 58 64 6F D8 1E 00 9C B7 A7 4D C5 FE F1 90 06 48 9F 48 9D E9 01 B3 D6 70 FC 86 F9 6A 67 28 F1 F6 D8 1C A8 77 1B 2E 91 BD BB 9C FA 90 E9 AB DC B7 56 04 00 B0 8E C5 89 E4 31 78 E4 20 20 0B BE 22 63 78 0B 9F 50 BE 8C 1A 07 94 82 4A 1D 8A 7A E5 7D E7 06 E9 64 89 A6 FC 6F 8A FF B2 3F E9 29 88 C1 D0 3E 75 7A 78 34 46 8E 2F 1E 18 69 14 C4 FC 6C 11 AC 22 F8 9E 6A A5 1A F5 54 C7 BC 3F D5 B9 B1 91 D5 CA D3 D7 AE C6 09 B5 91 45 31 6D 45 E8 5E 59 35 52 76 61 92 F7 59 EE 28 CA B2 68 0D 40 01 DF 95 69 C1 1D C0 86 A2 F9 06 79 5E 53 9C 87 17 9D 53 8E 5C 22 F2 FD 7B 2E D1 87 EF AE 02 E5 13 94 2D E5 26 86 FF 0A C0 D0 69 ED 5D 44 F7 72 53 70 C3 20 CF 41 D6 40 D1 24 05 40 93 27 6C AD 5A CC 11 E7 FB C3 FB AE 5D 73 C9 BF C6 23 83 F5 3A 8E 2A 4C 3D CE 08 E0 B1 EF 69 03 6F 89 04 37 97 CF 05 92 68 6B A4 2F 94 70 B9 13 B3 EF AC D7 1C 80 15 0B 3E A1 E7 65 14 2E C1 C2 BE CF 77 C0 82 DF 73 DD 3A 37 F7 90 FA 38 45 6B 87 CE E4 26 0D C1 E5 F7 3A F6 21 B8 80 67 98 35 11 9F 8B 87 34 CA 4E E0 66 32 C6 2F 45 F4 00 B2 29 45 70 BF 0C D1 5E B2 73 5D 04 A8 16 8C C2 CC D6 73 24 F1 65 24 E2 79 D6 E2 84 3B 05 1A 5D 90 AA B9 DA A4 AA 55 E8 D1 10 B5 97 2A E7 40 06 A7 A6 3B 76 DC 1B 54 6C 83 7B FA 05 E5 18 AA 47 7C D8 1A 06 75 22 B7 F2 C0 1A 5B 4B B8 75 6C BC 6E 11 AB 2D 1A 49 9A 37 A2 F4 3B C0 1B 41 47 7D D7 B1 69 EA DC 15 64 51 31 D0 26 D3 3B 50 00 1E 97 72 60 8F 32 BA 50 18 BD 1A 61 31 78 45 1A 85 70 54 26 31 70 84 D2 1F F6 14 6D BB C1 56 57 D8 03 13 47 8A A3 E5 CA DA 1C 31 65 E6 6A 78 EC D8 B6 73 C7 9F 3C 63 6C FA F7 CD 1D 56 6C F4 66 86 F2 18 A6 B9 73 5C AE 91 8B 5A 1A 83 C1 08 CD 3E BE E3 D8 28 9D 6F 0D CB 36 06 36 0A 68 B4 06 0F 1F 91 E1 D1 C4 E4 BA DC 70 5B 13 15 D3 D3 95 6D 78 5E BB 46 B0 26 DE 0C 44 9F 4B 78 1E BC B5 CA 9A B4 5F D9 CE 2D 94 00 57 35 C6 8E E9 F3 D9 19 B1 47 B4 3A DD CC 6C DF 3A BD 81 30 D0 DA 81 E1 0E F5 08 43 9B CB 85 56 DF BA C1 B5 E3 34 75 81 C3 95 F9 5E 14 42 14 AB 5C 7B 3C 14 55 46 3D 30 13 30 A8 81 1B AA F4 AE BD A8 85 1F B5 9B 6C C5 D8 C8 F2 10 06 B0 98 EF F5 BC 3B CB 1F 5E D3 C0 F9 B5 7D 18 3B 8C 41 93 34 AD 69 D4 34 0C E0 0B B2 56 B0 87 A6 F9 76 C8 04 91 03 25 AA 50 44 82 44 C4 45 47 C6 27 2D 23 00 84 9A 03 AC 55 43 30 2B 1E 69 93 15 4C 8A 15 15 55 CD AE 44 2E D2 16 4A A8 E2 5B E9 2E 25 2D 4B 50 00 6C 67 65 36 4C F6 1C E5 9C B9 1F F7 AD 99 7B D7 3B 3C D9 FA 5B 2D BF 52 56 78 82 66 7F 2D C2 14 96 A0 68 51 AB 26 E8 96 45 59 0E 21 60 41 2C A3 F0 6E 8A E3 6C 8E B1 92 A0 3C AC F3 A5 D8 20 73 E1 72 C6 30 4A C3 4D 94 83 5D 32 F7 00 66 14 D5 3C 06 DC 73 78 3B CB AC F4 29 68 50 09 64 24 A4 ED 52 9F D2 D0 43 48 45 AD C3 04 89 DD 1D 06 B6 88 12 AC 4A CB 5C B0 BB 54 D9 A5 12 1E 4E 5A 07 02 0B 2B D5 08 5D D9 12 85 04 05 35 82 15 C9 7A 15 6A CF CD BE 87 8F E0 04 22 96 22 86 BB EB 6B 6D 60 18 CB AC 50 55 9A 2E AB CA 51 10 98 97 AD FA B4 E5 7F CA D1 F6 14 AB 07 6A 20 9C 4D 73 88 63 0B 9C D6 CA CF 8D F6 A3 41 EB D1 D4 9E 5D 9D EA 68 E6 4C 5A 4A 18 CD 65 4C E3 CD E3 A9 16 B2 58 36 2C 25 26 A5 46 B1 63 CE 05 93 EC 5A BD 6C 0E A0 3D 39 94 A9 52 B0 EA D3 44 2F 3E 4A 15 71 98 84 07 DC 24 18 6D 9F 4B 8A 3F C2 18 61 0C 0A 52 7A 55 E1 BA 95 7C 08 41 4C F9 85 8E F4 E6 22 D8 8F 8A D2 0A 9E 04 8E C3 EA 2A CF 59 8A AF 51 26 63 D6 20 08 D1 AD 2A D4 E6 00 1B 38 03 A6 84 0C 51 2C 95 9B 6F 69 07 BC 91 4C A9 17 7A 88 A0 39 00 1F 90 DA EF 41 25 A3 6D B5 61 EB 91 6D 5D 5D 2B 75 8B 53 68 49 49 0E D0 1B 95 5B 18 FF 0A 2D 59 21 DF 03 04 D8 AF 38 E2 02 C5 18 21 0A F1 A3 05 24 81 42 19 55 A9 2A 8E A7 50 8F 4D A4 4B EF 3D E5 5E 9D 32 EA D8 C1 1D 04 E2 88 35 64 46 FB D6 76 DE 22 F8 34 97 AA 13 CB B6 98 03 5B 7D 58 06 D9 FA 63 6B C2 01 1C 48 A4 39 EE 2C 70 2A 1A 71 9F 64 0C 04 63 35 81 65 6E CE 34 4A F0 E1 D4 D1 9C EB 41 00 7B B8 B0 04 C0 2A 4F 93 05 42 21 2F 14 49 A6 34 2E A9 14 03 A0 DC 17 D5 EB 48 40 6E DA 96 F2 D9 10 55 62 D1 26 69 BB C3 A8 8D B2 63 B2 26 23 EB D6 1A F1 0D 45 C1 5C 4F 68 0C 6E 45 20 3A 3A 45 8F 1C D0 1F DC 38 34 97 62 FE 6A 79 7E 4B 5B 7D 2D 77 D7 09 5B E5 D5 AA DC AF 33 74 73 0E 95 E3 6B 45 B9 3A F3 4A F9 60 3F AA 4E 6F CB 51 F7 A7 E8 2D D1 FD 19 DA 5A 56 AB ED 9C 4B B8 6E C3 F5 04 52 4C 80 82 5A 05 DC 87 25 4B 8B 69 F1 E9 63 73 75 6A 61 85 46 AF 89 08 3D 42 A3 4A F8 D6 AC E8 D2 0B E3 68 0D E0 D8 CA 06 38 04 DC 22 24 F3 C6 60 FB 99 57 81 09 B5 A2 C9 3B DC 87 04 A1 C4 75 0C 58 09 E0 50 49 9A 2A 66 F7 41 48 5B AF 54 47 0C 81 4A E7 48 56 E7 59 AE 24 5C 2F 15 D7 27 FB F0 B4 41 CA B6 0D 3B 39 78 19 D2 3F 37 BF E1 81 EC 79 1E 29 C7 42 21 E1 92 69 80 22 DE 38 BE E3 B6 FC AC E2 74 95 8B AC 82 04 12 0D CE 0A 95 CD 5A 1D 40 D3 A8 1F B2 8F D8 D7 8C E9 5C CB 94 3B DC 8E AD 19 15 A5 ED 0E A1 93 8B C5 89 BE DC 28 2C 01 47 E1 0C 86 60 DC 17 A8 3B 74 8A 4F 74 83 BA 8C 3B 2E 90 0E 8A C5 EC 3A 86 02 F8 D2 E6 CD 11 B5 A8 BD E3 02 95 E4 BB D7 01 76 05 EA E6 78 5A AF 3E 60 FF 66 EB F3 F1 37 6C 57 A0 05 CC E7 5F E3 F6 13 65 1C AA 3A 1B F5 EE B5 51 D9 F4 08 23 32 76 16 42 21 24 37 26 06 3A D3 1C 56 09 B7 D2 46 75 59 A8 22 58 1B 46 8D A6 47 D9 0B B2 95 AD 8F 8A C8 0A 79 E0 0B AD C0 DE B1 5F 79 0B B1 63 B0 E6 B1 7E 8F 01 CB 1B 86 6D 8B B5 33 4C 19 C0 CD 37 CA 18 53 93 CA BA 1B 15 30 6E D1 B9 02 9E BA 16 62 0D B7 66 88 BE A6 91 0F 14 84 BB 83 7F 4B 52 2E 5A 6F C9 B3 0F D0 8B DF AE 28 1F 2A 52 07 AD 79 DE 7D 8E B6 14 F7 D7 B6 D2 1F B0 59 6E 0D 1B 12 43 2D 0B FE 75 D1 FD 90 C0 BF 2F 3B 83 F9 02 F2 E7 51 B8 89 71 37 FC 01 19 F5 E3 4E 59 7D A4 C6 5D DC 5C BD B3 A9 FB E7 C3 FF 43 A3 42 32 1A 98 BD 2B 28 46 7F A3 69 BD 30 15 CF 0A EC CA 1B 48 86 FF 67 DC 0C 69 9B E2 1B 36 51 5C 9D E6 31 37 2B FF 23 93 6A 79 EB 5D AE 6B 0E 39 FA 7B FB E1 6B B2 F9 7E D5 FC 35 50 AC 55 D6 B9 04 0B 4D A0 4E 64 9B 10 02 7C B6 00 45 78 29 45 59 65 4D 7D B5 B0 E7 16 A0 0D 3D A1 8C 51 09 A4 8F 0F 5F 3A 63 FC 9C E1 E7 BB AF EB C0 F4 AD F1 3D DD E3 BB 15 F8 D6 13 D4 D8 5B 68 DB AE CE 0B DB 58 B0 F7 0B 59 B5 DA D7 CC 3F 87 BE 90 55 EB A9 4A D7 1D 98 90 8F EB BE 3D 7C BF AC FF 5F 00 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 3A 3A C6 E7 C2 02 00 00 A5 05 00 00 18 00 00 00 78 6C 2F 64 72 61 77 69 6E 67 73 2F 64 72 61 77 69 6E 67 31 2E 78 6D 6C 9C 94 C9 6E DB 30 10 86 EF 05 FA 0E 02 EF 8A A8 C5 9B 60 39 B0 65 A9 0D 90 B6 41 91 9E 0B 9A A2 62 02 12 29 90 F4 12 14 79 F7 0E 29 C9 46 DA 06 68 7B F2 98 33 9C E5 9B 9F 5A DE 9E DB C6 3B 32 A5 B9 14 19 0A 6F 30 F2 98 A0 B2 E2 E2 29 43 DF 1E 4B 7F 8E 3C 6D 88 A8 48 23 05 CB D0 33 D3 E8 76 F5 FE DD F2 5C A9 F4 A4 B7 CA 83 04 42 A7 F0 37 43 7B 63 BA 34 08 34 DD B3 96 E8 1B D9 31 01 DE 5A AA 96 18 F8 AB 9E 82 4A 91 13 A4 6E 9B 20 C2 78 1A E8 4E 31 52 E9 3D 63 66 DB 7B D0 90 8F FC 47 B6 96 70 81 56 AE 33 73 92 39 6B 9A B5 A0 7B A9 3C 56 71 B3 D6 19 82 09 EC E9 10 53 2B D9 F6 D1 54 36 2B BC 0C EC 48 D6 74 19 C0 F8 52 D7 AB 28 99 4D 27 57 9F 3D 72 6E 25 4F AB B0 BF 62 CD F1 CC FA C3 28 9E 47 93 8B CF 5D 71 B9 AF 05 8D BC 16 8E FF 5C 18 EA 46 6F 15 BE 26 FF B5 70 B8 18 7A 85 A6 AE 85 C7 72 1D A7 FD 05 71 7C E0 F4 41 0D 4D 7C 3E 3E 28 8F 57 19 8A 90 27 48 0B 5B A6 A6 C1 F8 3B E0 32 BC FA 08 1B DA 09 A1 C0 59 31 4D 61 CD 8F CF 5A 0A EF 4E 18 45 20 02 68 92 94 9D CD BD 36 83 E5 1D 14 CF D0 8F B2 8C 36 93 A2 4C FC 12 2C 3F C1 9B C4 DF 14 C9 C2 2F 81 4F 11 CD CA 3C 8A A7 2F F6 76 38 4D 29 E8 C0 80 04 EF AA 71 FF E1 F4 37 05 B4 9C 2A A9 65 6D 6E A8 6C 03 59 D7 9C B2 51 51 A0 A7 30 09 9C 02 DC 28 3F E6 C5 26 29 8A 38 F1 E3 BC 28 FC A4 CC B1 BF 49 70 E2 47 78 13 AD 37 93 68 01 CD BD A0 60 B5 0C 5C F7 E3 AF 9B A2 57 82 C5 72 25 64 79 D9 68 2B 92 57 F8 76 0D EF 4A DE 80 6C 48 6A ED 61 80 BF 7A 0E FD 0C 5B 49 0F 2D 13 A6 7F 13 8A 35 0E 85 DE F3 4E 23 4F A5 AC DD 31 58 8E BA AB C2 37 59 47 F3 35 C6 30 92 9F 4F 70 0E AC 67 85 BF 5E 24 33 7F 86 8B 19 4C 3D 0F F3 30 EF 59 27 E9 41 B3 7B 49 49 B3 ED F8 05 76 F2 CF B0 F1 00 FB 48 9A 0C E1 B7 40 F6 48 2C 1A AD E8 57 46 0D 20 04 DB 28 66 E8 DE 9A 35 90 1B CE 61 11 17 87 C3 7C 25 6B A1 EB 0E 54 BA 3B 7D 92 15 08 94 1C 8C 74 30 CE B5 82 47 4C 52 20 E9 9D 41 C1 EE B9 22 EF 39 43 71 98 C4 D1 C4 36 E6 16 EC 51 EB C6 13 C0 01 DF 38 0A 01 B3 69 18 4F 5D 00 54 1E F3 74 4A 9B 0F 4C B6 9E 35 00 3A B4 EC EA 90 23 28 A3 57 CB 18 62 13 0B 69 57 EF 6A 34 E2 D5 01 E4 B4 27 6E 10 DB FA 60 5E 9E 20 6D 38 AC 7C 4B 0C 19 55 F5 EA 9B 35 84 DB 2F EC EA 27 00 00 00 FF FF 03 00 50 4B 03 04 0A 00 00 00 00 00 00 00 21 00 CA 02 BC 85 BF 19 00 00 BF 19 00 00 13 00 00 00 78 6C 2F 6D 65 64 69 61 2F 69 6D 61 67 65 31 2E 70 6E 67 89 50 4E 47 0D 0A 1A 0A 00 00 00 0D 49 48 44 52 00 00 00 B4 00 00 00 40 08 06 00 00 00 B3 CE 5D 3D 00 00 00 01 73 52 47 42 00 AE CE 1C E9 00 00 19 79 49 44 41 54 78 DA ED 5D 0B 98 5D 55 75 3E 37 2D 56 8B C5 6A 69 69 29 62 92 39 E7 CE E4 01 6A 53 DB 5A B5 A4 56 68 41 12 32 73 CE 99 9B 99 84 10 DE 2F A1 95 2A A2 95 30 50 2A 56 0A 4A AD 8A 3C 7C D4 1A 8A 96 62 85 00 16 CB E3 43 0B 05 23 B4 50 50 0B 3E 50 30 05 21 0D 6D 5E 93 4C A6 FB 5F 6B AF FD 38 E7 DC 7B CF BD 73 27 19 CC D9 DF B7 BF 3B 73 EF 79 EE F3 EF B5 D7 E3 5F EB 04 41 CF DA D8 AC 60 D1 29 FB E4 BE 9E BD FA A5 41 7F DA 1F D4 E3 A3 54 7F 47 50 4F 3E A8 3E AF 09 C2 E4 FA 20 4C 6F 0C A2 E4 1F 55 5F AB FA 95 EA FB 0B 83 70 E8 44 F5 DB E2 A0 6F F9 AB D5 DE B3 9A 9C A3 16 54 AD 6A D3 D3 0A 80 3C 6F E8 D7 82 28 8E 15 30 FF 46 7D DE AF C0 BA 31 88 D2 C9 A0 7F 79 FB 5E 6F 4C AA ED 77 A9 FE B4 EA 5F 55 FD A2 A0 2F 7D 6B B0 20 7D B9 77 8E 0A D8 55 EB 71 AB 05 8B 17 FF AC F7 4D DF D0 EF 2B 00 7E 42 F5 EF 06 F5 61 05 D0 11 0D D2 61 DD 1B 1A B8 23 05 1D DF 37 9C 6D 87 ED B6 51 AC 00 1E AF 57 13 64 2C 08 E3 05 DE 64 0A D2 9F A9 1E 45 D5 A6 D8 00 A2 31 AB 0E 84 E9 52 25 81 6F 53 40 DE 41 00 24 29 9B 4E 7A A0 66 A9 BB 41 F5 7F 57 A0 BC 43 A9 1E 5F 52 FD 0B EA FF 2F AA 7E 93 EA F7 A8 FE 98 95 E6 32 19 D2 49 2B DD 01 EE E4 39 D5 AF 55 C7 78 BD 39 3F 4D AC B1 59 D5 73 A9 5A E7 CD 95 CA F5 F4 B7 14 B8 6E 36 80 03 F8 EA A9 95 CA 90 D4 51 FC 79 D2 9D C3 F4 B0 60 60 F9 EC 20 5C B1 5F A1 54 85 0A B1 20 7D 95 DA 76 80 F4 ED 28 FE 33 25 91 6F 51 C7 F8 09 4D 90 EC 24 89 92 FF 53 FD 8A 60 76 FA AB 85 D7 56 B5 AA 95 56 31 C2 23 7F 4E 81 E9 CF 55 DF AE C1 65 55 84 28 F9 1F 32 F2 EA 43 83 41 7D 64 FF 29 9F B5 1E CF 55 C7 3B 95 A4 7A 94 4C F0 39 44 6A D3 B9 7F 10 44 43 A3 76 62 2C DA A7 7A 54 55 6B 6F F8 C9 92 DE 37 B8 50 49 DB AF 05 03 A3 0C 64 74 06 D6 4F 14 E8 2E 0B A2 C1 79 C5 FB 43 2A 67 BB 1C B7 E8 B7 02 83 8F 75 74 A8 28 3B 79 15 50 2A 8E 48 EF 30 B9 CA 18 8E 95 A4 AE 5A 6B 7D 59 00 15 2F 23 1D 96 C1 BC 4D 4B 64 18 6D D7 06 73 D2 FE BC 6A 42 93 A0 D6 A5 37 A2 66 C0 9E 05 68 3D 7E 9B 3A EF 9D 56 CD 51 C0 1E 58 81 6B F9 B7 60 EE 60 DD 01 75 E5 05 A9 5A 13 30 87 F1 E9 AC 5A 34 18 CC 0C 6A 65 C4 A5 6F F7 B6 9F 3E 20 D5 7C F7 A0 3A 57 18 9F 43 BA 34 4F AC AD FA 9A 7E 18 84 43 BF 53 81 BA 6A 05 6A 82 01 F3 39 BC B4 43 77 C5 72 4F 2A C6 75 41 7D C9 FE 9E 51 B7 BB 74 F9 05 E9 4B AC B4 1E 7A 03 79 4E 18 CC 5B 09 DC 61 F2 BC 92 DA 6F A9 D4 8F AA 65 96 7B 2D 99 C5 C3 20 12 1A 11 3D 0F C8 7B C0 65 46 40 D5 2B C8 41 E9 AB C8 F5 E7 AB 42 CF 05 73 E3 DF DC CD 93 AD 6A 33 B2 89 54 63 9D 59 24 B3 76 C5 C5 67 EF 01 A9 5C 42 0D 51 93 2A 4A FE 2E E8 17 50 8F B0 DB 70 CE D0 6B 2A 49 5D 81 59 7B 33 B0 74 37 58 CD A0 4F 07 CC 33 09 20 EE C4 82 CB 50 D4 0F 7C D6 D3 3B AC 8A 52 45 15 F7 4E 23 10 7E E6 30 F9 BA 27 ED 5C 35 63 26 02 43 40 8D CF 28 F9 67 2D A1 C5 50 FC 4B 5F 9D AA DA DE A5 37 47 F1 C5 46 1F E5 CF B5 CD 25 73 33 3F F2 EE E8 19 70 0A A8 5F 03 72 54 F2 5F 36 E4 AE 54 A6 30 FE C3 4A F5 D8 1B 55 0D 0E 67 6F E7 80 C9 72 76 CD 0D 0C FE D2 8B C6 B8 92 6B EC 4F DF 4A 91 45 7B 1F 0F 06 07 2E F9 F9 4A F5 D8 2B 9A E3 A5 88 D2 75 26 9C 4D 86 60 7C 54 4B 30 43 3F 3D F4 D8 7D 09 2C 8B 32 FD D0 23 F6 6D DA A7 B2 2D 6D AF CE 99 16 02 D3 06 72 C2 E4 03 FA 5E 98 34 15 26 EF A9 A4 F4 DE D0 04 AC 61 72 8C 71 CF 11 00 E2 6B 0A 41 EF 2E F5 00 09 96 77 44 E9 EA C9 37 A8 83 EA 89 CF 30 79 40 1D EF 7E BF C7 F7 D3 F7 B2 AD 74 DA 36 F6 B7 C5 77 B2 AD 1C D3 9C 23 79 84 22 86 F9 6B B3 80 9D 9B BE 42 ED FF 1F 0E 59 EA 87 EA EF 03 3B 07 B5 84 E8 A7 BB 57 AD C7 D2 39 F9 8A F3 F0 9F 35 61 E4 BC 74 B6 52 10 54 CE F9 AB 1D 9E F3 72 FB 37 F4 EF 66 3D CB 87 6E BA DD A8 7F 4C E9 F3 56 C2 37 3E DA 54 85 30 93 74 E8 58 7D 3F 7C 8C 30 5E C3 BB 54 6A C7 4F B7 74 AE 27 7F 40 4B B3 D0 3F A3 E4 D2 16 FA A6 0B E8 BF 36 1E 05 18 91 22 DD 4D E6 49 BC 43 EB E4 D9 3E 9E E9 45 DB EC 30 40 E4 6B DA A1 CF C1 D7 19 A6 49 F3 6B 94 EF 94 24 8E 92 7F 35 09 04 AE 4D 50 5E 4A E3 5E 67 59 C3 99 84 80 CB 53 E9 B4 BB 52 D9 39 76 D5 7A D7 A2 F8 93 0C 44 52 39 36 52 1E 60 F3 87 6E 01 5D 4F 3E A6 3D 21 DB 35 88 B7 92 AA 12 A6 C7 92 67 21 1A FA 3D 75 CC 37 53 A7 BF 07 5F 4B CC 38 06 E8 4E 92 9E A1 FA 8C 92 B3 F4 6F 8B 83 FE A1 37 D1 F6 F8 3B 4A 8E A6 68 65 98 DC CA 92 36 E5 89 42 0C BB 78 B8 A5 91 27 93 35 4A 8E 37 FB D2 F9 E2 95 6D 01 6D D5 96 83 89 F3 CD EA D3 DD 3A 11 E1 2E BF C7 77 F3 A7 FA 3D 8C BF E6 75 DE E7 AE 26 FD 1E DE 87 8E BD 2E 18 58 36 BB 32 5A 7B D1 A0 57 46 F1 F7 4C 8A 14 22 6E 59 E0 B6 02 B4 35 22 B7 29 BD 76 59 DB F3 D5 93 3F 35 C6 1A AB 03 E3 C4 C9 68 D7 60 E4 89 E4 67 CA 68 DA 1A 00 FA FB 83 47 5F 49 7A BE 4D 0E F8 62 13 DB A0 40 65 49 E7 D3 64 85 8A 43 E3 A3 D5 9D 79 AB FC 3E 1F 9F C7 E6 55 29 7C 37 7F 55 7E 7B 1C 43 56 32 52 9F D4 18 F4 35 16 BE 68 BC 49 33 BA 01 18 9E 04 53 C6 61 6B 09 96 07 34 E7 FE 7D DC 07 84 E3 33 C6 B1 24 62 17 25 E7 FB 80 56 6A 09 54 1E 34 64 89 67 FD CD B2 DF EC C5 2F 25 F7 1B 01 61 B8 BD 84 76 01 1B 25 1F 76 92 70 9F A6 C4 81 56 F7 68 DD 7F FD 6A 5C 36 E8 89 A7 F9 D7 C9 8F D5 35 DF AB BE BF 5D FD 7D 2B D9 1E 9C 46 76 9F 56 8B 26 74 DF 41 C6 32 FF F6 15 DE 56 ED 43 FB E2 1A 1A 7A 32 D3 B1 FF 3B 88 1A F3 2A 2F 4C 6F D4 8D 8F 1B 90 44 C9 13 46 C7 6C 6E 79 3B 3A 74 FC 11 AD 72 8C AB 89 70 84 71 E5 B5 92 98 48 AF CA 4A 68 A8 17 4D A5 93 93 0C 8B 44 59 DA B7 9D 0E 2D A7 D4 BF 31 87 DA 4E DA 7A B2 BC 94 BA 82 94 B0 28 79 C6 72 BF 93 2B 48 0D A1 89 97 B9 AF B9 50 AB 00 E2 54 54 A3 9D A4 66 F9 63 59 A3 7D 59 95 F9 30 DB 08 0D 36 C2 B1 1A 54 80 9E 62 3B 28 7D 99 02 C9 37 8C 17 21 8C FF B6 D0 3D D7 DC D7 7B 55 70 C8 C9 93 94 00 1B 0E 1D D4 FA 81 98 48 64 87 80 76 8E 89 84 5C 90 F9 B1 8C 47 F1 48 69 09 BD 80 18 79 8E DA A1 26 62 2B B5 C3 93 D0 4A 9A 72 02 C1 CD 2D 27 6A 9F D2 FD DB 03 3A BB 3A 7E 29 18 58 89 55 F1 C7 15 A0 7B 22 9D 07 E7 51 1E A0 E4 04 86 E9 E9 CE 03 68 05 E8 C0 78 47 EA F1 25 E4 42 33 0F A2 99 64 EF 56 42 3B C7 EC 1B FC 15 35 E9 DE A9 B6 BF C0 B8 15 5B 7A 08 3C 5E F7 F5 8E C7 E4 2E 3B 11 5A B8 FD 06 1A 75 32 92 09 D0 3A 77 91 22 8E 8E 97 02 DC 17 9E 6C 87 65 00 3D 41 69 63 F4 1B 6D 63 B3 70 44 C2 87 43 89 06 F4 F3 26 7D 6D CF 03 FA 45 EC 71 81 17 A1 9E 8A BA 31 41 DE 85 F2 96 76 AD B3 C1 98 0A A0 BB 39 9F FC 2E DC EE F4 3C C7 A5 68 83 2C AD 00 3D 37 8E D4 B6 2F F0 B5 2A F5 A3 68 7B EB FA 7C 4B 4E 42 9B FB CA 26 EF EA 6B E2 E3 4F 90 E1 29 35 47 B0 6A 02 D4 6E A7 FD 33 C1 AD EC 36 66 DB 66 1C 75 C9 28 CA D8 36 DE B6 46 BD 6B 0F 6A 39 9F C9 0F D5 C7 A3 6B 6D B3 6A E2 1A 4B 5D B7 7B CD 6E 4A 5E 33 A1 19 C5 7F EC 19 4B C2 1F 2E EF 3A AA B5 91 E6 BD 04 74 90 F1 05 97 6C 72 DE 64 C8 14 B1 81 47 46 12 00 9A FB DA 59 1A 83 17 02 E9 6B AF AD D6 39 A0 0B 83 53 0C 0A 78 78 24 6D AC B7 D2 B3 E6 81 BF 1D 38 CB 9E CB 4D AE 68 7B CC DC 73 2A 7B FD B3 DA AF 54 45 D7 80 00 8A 5D 86 BF 69 4B 6E 75 02 E8 B2 89 B0 3D 91 D0 1D 9C 2F 73 5E 21 5E 59 7F F6 D1 6D EE B5 DC 8A D0 1D A0 ED F1 F0 1B D9 32 E9 2F 53 19 35 D4 1A C9 76 7C 6F C8 55 FA 9A A1 7E 15 6D 8F EF C2 23 F7 CB AB 8E C8 C3 1C FA 0D F6 6A 25 27 A9 89 D4 20 DE BB 51 C9 88 36 9C 92 4A D7 1F 1F 5A A8 3E BA 20 03 56 20 14 EA E9 A0 BA DF D5 EA 98 AB 88 FB 43 B6 80 B3 5F 16 98 D8 AF E8 BA FB 97 1E 48 1C 1D 77 8C B1 6F 5F BA 48 9D A3 A1 8E 7D 8A 8E 6F BC BE 39 A8 A3 F8 D3 4E C9 AD DB BD 9B 2F D5 C6 66 F9 33 31 FB 7F 8F 8C C2 29 E9 78 E2 85 A0 E5 DD 16 AE A9 C7 C7 95 B8 57 BB B4 37 6B 5D 01 5A 8F C5 EC 65 BF A8 1E D2 1D EC 0A 4C 1F 55 FD 71 EE F1 E3 E4 71 0A E3 6F AB CF 6F A9 FF 9F B4 46 30 4D CE 5F 27 CE 7A 98 7C 9F 7F 57 DB F2 3E F8 FB 29 75 BC F7 79 63 D5 37 F8 47 BA 56 E0 38 F9 C6 A9 AF 94 2C A4 2F D3 E4 40 1D 95 30 D9 4C 54 86 30 79 6F 06 8C 56 75 83 17 0C CF 90 EB 16 8E D3 71 E6 1F C7 86 3A 7B BC 36 6A 37 E5 AA 42 69 8D 20 1A D7 2E 44 A5 AC EF F0 B5 D3 E7 B3 EA 99 9C E9 60 13 F5 11 FF 85 6D BC D4 AD 7D B8 85 0A 7E 82 AB 93 9B 74 51 72 83 53 66 EB 86 B6 01 87 56 C0 CA 3E F4 42 7D AD 0D A0 C9 78 2A E2 58 B7 39 57 19 40 53 00 29 7D D2 09 20 9D 55 C2 00 6E DF BA 02 B4 B8 21 95 54 46 74 75 E1 49 93 64 78 66 8B 57 0A 58 16 9E 88 E3 9D 6B A5 69 A3 8F C6 6E C1 09 FE 7E 30 30 17 9E EC 7B AB A2 F8 34 72 0F 02 C4 08 4A 85 C9 D5 5C BC 47 81 16 D1 4A 00 38 8A 1F D2 44 33 E6 C0 23 F8 E5 8D B3 D8 21 14 BD 7D CC 04 8D D8 95 F9 55 F2 76 45 C9 A7 68 95 27 6E CE 4A A9 76 B5 8E 26 9F EB CE 85 EB 13 F7 4B 13 60 85 AD 6B 88 FB 04 36 D8 80 BF 8C 8E 83 EF 24 60 C5 E7 DA 45 F7 8D EF EA C9 F5 79 BC C2 E9 6F 01 BD B6 24 A0 ED C3 47 65 50 24 04 88 31 49 8C BD E4 A2 96 91 BF 56 80 6E 1A 5D 8C E7 52 F5 24 91 1C 1D 49 6A 71 AB 29 29 04 69 D0 AF 01 2D 0F AD 57 80 86 8B AE 53 09 0D 35 02 D9 40 E0 C4 70 E9 B3 49 87 DF 32 A1 83 32 97 12 08 8D C1 AE DD 90 21 BC 4B 0A 1C 5C 55 8A 1F 38 24 1A AA BE 8A 9F 9D AF 69 5C 8F F7 16 F5 DB 8A 02 B5 F3 54 93 37 6A AB 61 9D 9B D3 97 11 67 A0 CC 7A 63 58 3F A2 26 E4 1B 73 F7 15 0D AD 56 BF 6D A2 67 0B C0 82 F1 28 A0 E6 67 79 14 97 54 46 61 A2 F4 21 3D 56 5B F8 99 A0 6C 5C 7C 1E 83 5B 7D 07 4C 52 15 DB E4 3E 43 32 33 9C 21 75 0D 7D 3A 32 6D 26 5E 94 7E D9 01 F4 75 E5 01 6D 96 8F F3 B5 1F FA 9B FA FF F7 07 0B 20 4D 40 13 1D 3A 88 A2 7B D0 E7 FA 97 FE 82 71 55 E5 00 1D EF B0 12 7A C5 7E A4 63 49 97 65 25 4A 0F 27 E6 1D 47 08 4F 0F 6C ED E8 F2 BA FB 9C A5 07 70 65 D4 86 B8 28 CF D9 73 80 2E 68 7D C7 BC 9A 22 86 1C 68 D9 4E 92 32 4C FE A2 AD 50 89 92 CF 68 69 B9 49 8D D7 21 19 E1 F1 79 07 80 F7 65 56 B8 9A 0F EA 86 95 80 51 F2 2E C7 DD 18 90 B3 00 AA 0C AF 6C 13 4C 73 70 26 58 6E 92 A4 EF D6 F7 B1 4D FB F0 6F 2A 16 54 86 0B B4 53 8F D9 3D FC 19 DF EB B8 65 F9 7E 2D F5 C1 72 DC C1 F6 F4 EC 2B 54 00 35 41 95 F4 C6 F2 2A 87 89 8E 1D CC 0C 36 1A FC 5B 75 A9 2E F1 24 FC 88 89 37 C9 83 CC 99 56 4B 1F 2F 27 6B 3C 40 87 4E 00 02 55 98 78 DB F5 BC 9F 5A 0A 99 2F FD 1D 27 4C CC B3 BE 2F F9 DD 8E AE 15 13 0C D7 64 00 9D 9C B1 E7 54 8E 66 AE 45 25 8D 78 6C 74 39 86 F8 7B A6 20 A5 A7 66 99 F1 3F 84 74 56 06 F4 5A 0F EC CC 61 79 84 CB 14 37 38 6B 47 7E 13 D7 A0 1B D5 85 3D 25 5C 13 A1 D9 5A 82 D7 95 1A 98 9B B5 4A F2 85 CC FD D7 3C 9F 3C F4 71 5C 7B BF 0E EF 33 75 77 65 6E 2C 40 8A 63 40 EF A0 4E 7F C7 0F 73 51 4F 7D 2F 2E D5 B7 8E 5C 57 CD EE E4 88 F1 DD 3E 66 21 FA AD 84 BE B3 65 B0 A1 59 C0 02 52 40 74 A1 28 F9 5F 5E 2A D2 49 6B 7C A8 0E A9 5D 4F 2E D7 0F ED 82 8C CA B1 D3 14 86 89 92 FF 24 7D 91 08 40 B2 BF D6 B5 B8 24 2F 2F 89 0B 8E 9F 24 8B B7 D4 B5 9A F0 F7 00 67 B2 EB 5A D5 32 C0 53 65 B7 F5 02 D0 F2 3B 2C FA 28 DE 66 8B FA D0 98 1E 9F B9 4E B7 6E CA 1A 3D 36 B8 9F 25 36 F0 23 36 03 54 2C 61 36 D2 31 CF 2F 8C 16 F3 B9 43 75 8C 1B A9 76 21 5C 9C D2 FA D3 39 6C 4C EB D5 94 23 CA A7 17 0B 93 2C 2E 46 B5 3A 31 C2 EA 50 4E C8 28 DD 5B A8 13 22 08 85 E0 96 0D 46 09 76 BC FA 2B 4A F0 99 49 89 ED 28 4D 49 DC 76 6A 66 18 CB B1 34 48 96 E9 13 EC 22 7D 06 D6 35 4B 0B 75 81 F1 87 F4 2B 26 CE 20 7F 37 DC 66 A2 96 F8 80 DE 41 7E 5E F1 15 63 5B 48 73 2A 72 A3 AC 5E 3C 50 E8 91 54 7D 74 84 25 4D 98 9C 44 3A 71 29 B5 43 8C C2 A1 37 D1 6A 20 6E 3B 49 9C DD F3 80 CE AA 10 37 18 20 C8 CA E7 01 46 83 06 6A 1C 54 3D 92 A8 4A 72 09 90 45 4A 92 7E 0E 5A 83 96 68 C6 07 1F DF 46 3A B6 09 2C 35 11 56 1E FD B6 61 13 8E 99 13 73 38 0F 5D 6E EC 5C 5A C4 05 BA 72 80 50 8B 37 9A F0 BE 39 B6 91 D0 DB B4 14 FF 76 B1 F7 C2 AC 60 67 18 6E 3C 57 A2 5D 6F EE 97 B6 01 70 A4 72 3E DC 26 98 A5 65 5C 59 E6 A2 95 02 CF 6A 83 95 BE AC 12 7C A0 A7 46 21 96 56 DE E7 31 AF 26 74 67 FE EF 11 76 D7 0D 73 7D 69 E3 6B 9D 11 80 76 85 C4 51 46 B5 62 A9 EA D0 6B 1D 6F 11 84 89 E5 A6 9C ED 00 CC 55 61 C6 8C 04 F7 12 30 08 D8 4F 92 9A 09 BF 33 B8 E8 4D 55 35 F5 6C 79 EC C7 F5 B3 DE 4C FE EC A6 DE 26 D9 4F 09 24 59 1D A2 54 EE 65 C8 F7 78 18 40 4B C9 89 9B 0A A4 BD 5B 91 E0 B4 36 80 56 33 CD EA BC 93 36 4F AF E4 43 86 AE 03 87 B8 2C 5B 70 43 61 E6 7B C5 5D B4 54 49 4B 04 56 BC 65 C6 59 6A A8 18 7A 3A E8 87 9F 4B BB 16 25 BB E6 22 C7 40 7A 82 AE 75 46 01 5A 24 E3 A2 7D 28 39 C0 AB 2D A2 56 BB EC B5 C2 88 C7 6A 08 5F 34 BF 64 C9 39 8F DE 0E 81 0A 90 AA 48 25 6C 88 41 B7 95 69 0E 0D 9B 36 17 C6 9B 89 DE 0A F2 57 2E 6E 80 8A 54 23 8E A4 8D 37 B6 A6 BA 9A 89 79 9C F3 CE 1C 3E BF 70 85 F2 12 5A 00 7D 5D 67 80 56 2B 94 07 68 D6 8F 9E 71 0A 96 BF 6B 0A 86 52 AD 89 3E A5 41 55 02 D0 D6 48 71 A2 81 05 F1 FD D2 D7 E6 E5 4C DE 6C 0D E0 EC 32 3E 13 00 ED F1 4E 4E B4 0F 8E 40 F4 38 19 79 D6 F5 35 C0 A5 8D 57 32 A5 B5 F0 99 99 DA 7F 2F A3 C9 0C 46 24 F9 AA 47 F5 A4 C6 BB 6B 08 DC 2C 79 FB 47 C5 B6 F8 A8 07 54 22 75 B9 80 56 E7 15 C1 D2 52 42 23 9F B3 61 FD C7 AC D6 9E 3D BD 80 C6 05 C1 52 14 3D BA 23 4F 87 9C CC 05 58 2B 72 CB 94 22 85 B5 D2 1C 82 42 FD 79 E9 81 EC E1 30 4B F4 C5 DD 47 1E A7 0D D0 4E 7D 94 25 FB 93 3E 29 0F CF 33 62 E9 5C EF 63 55 42 19 90 51 FC DB CD CF E1 8C 17 52 BC 98 A9 78 2B 81 9B FC C4 02 6E 9D D7 09 B5 80 8F 7B B1 23 08 AE B5 80 1E 66 F7 A0 84 CC 5B AB 1C 27 38 6E 3E 09 B4 AC 9E 5E 40 F3 E0 7C D0 66 81 80 97 AB 79 CD 3D CF 6D EB 09 39 A9 B3 66 A2 93 42 4C 4A CB F2 38 F6 0C A0 73 46 D5 88 13 EC 50 C6 9C 48 5C 70 D8 61 0C 46 C9 3F 35 99 98 AD 27 29 68 00 30 0C 11 B0 C1 33 97 A4 09 5B BD D5 12 D5 28 A9 C2 F3 94 6C 37 01 95 3C A0 5D 0F CC 3B ED 73 D6 E3 21 74 DA E6 3A 74 0F 00 CD 71 7E 9D CD E1 70 1C 7A CE CB DD ED 80 76 F5 E7 CF 58 2E 80 5A BE 0B 7D BB D3 08 E8 4E CE E3 27 17 3C EB 90 A9 C6 09 64 94 74 6C 82 43 49 53 1F 35 0C 37 F8 8B C9 F0 D3 46 9C 97 69 23 2A 85 12 60 61 7C 95 43 21 D6 12 5B 7B 32 EA C9 91 8E 87 43 07 5E E2 B8 89 40 70 B9 F2 97 67 4A 1C 3F 61 6B 8A 1B 2C B4 00 B4 C3 16 6C 65 14 7A 6E 3B 21 C8 C0 65 27 31 75 84 60 7B B5 1C 4F 0F 39 A9 33 E9 8C E0 0F C8 3F F5 54 F4 E7 AB A7 AE 3F AB 07 22 9C 5E 71 97 F1 3B 60 7C 40 4B AE A4 04 32 4C AE 65 49 5B 04 D7 CA 63 A5 5F 98 44 BA F0 15 3A 34 FC A0 66 A7 05 85 EC 36 AC 4A C2 87 00 59 48 1A 40 5D A4 BE 41 15 31 2F 65 A2 FB 38 5C 4B D3 97 7B EE 3F 06 DE 45 C5 F6 8C 1C 53 5D 0F FC CE BE FB F1 63 E6 19 DB E2 46 57 7B 80 86 BE 6E 8E E3 F8 DD 05 B4 61 7C 66 4E 42 CB 18 E8 63 8A FB ED 12 67 D9 B1 4B 4A EF 00 D6 A9 51 D8 4B 75 E3 5C E7 0D 04 93 E6 35 1A DD DF 5B F1 24 A7 A0 48 16 D0 8D 37 74 74 8C 9C D4 A7 3C C5 ED 4E 84 F4 79 36 06 57 D8 D2 66 59 63 D0 DE F7 D1 C6 C3 90 2D 54 8F 71 96 48 21 FC D9 42 63 B0 8C B6 0D 26 91 58 0C 3C 17 48 A1 12 80 22 ED 31 59 F9 B5 7C 2F 31 DF F1 9B 15 C6 1D 0F DA A6 C2 17 4A 21 32 CA 6E C5 2D F4 59 8F 3F D7 72 9C C0 13 31 E5 DD 28 DA FB 40 F1 C0 61 49 42 66 86 49 C5 32 6C AD 5A F7 52 3A 53 EA 4A B2 36 F2 81 95 71 22 39 15 BA ED BA 94 A0 F2 50 39 FC FB 2D CB B0 8B EF B5 3A D7 14 F4 67 F0 79 29 EC DC 98 C7 F5 44 D4 B2 DD 97 AC CA A5 60 E1 01 E0 37 F8 BC B1 ED 80 DA 07 FB 76 E4 9D 89 D7 99 F1 B2 C5 27 9F 09 C2 65 7D 85 F7 61 F2 2F E3 25 F6 CD 0B A6 84 C3 FB 9B BB 60 93 EB 69 9C 58 9A 7E 36 77 1D F8 8E BD 2A 52 4F A5 98 67 82 15 8B E8 A3 0D E7 F5 7B 9A F6 20 9C 1A 8A 2B A4 73 98 7A 31 E2 14 A9 4F D7 11 87 03 5E 14 F7 AD 66 48 85 A3 95 56 4D 4A A3 CF 93 FA F8 A8 FA 7C 2D 4D 16 1C 2F 17 AA 74 8D 03 24 7D 76 27 35 5B 4F 80 6E EB 72 74 2B 9D 25 63 9B 25 CF A9 53 90 CE AE 4E 7E A9 0E 54 6C 20 70 51 A9 03 92 9E 3B BD 5E 4F 9F A7 DF 68 1B 70 80 89 88 F5 57 A5 54 1E 5B 1B 24 71 F4 68 BD 34 27 57 35 3D 46 56 42 73 E9 85 CF 11 AF 83 A9 A2 B7 D1 2A C5 93 F2 15 CC AD 8E D7 18 AE 34 F8 D8 0C 0E DF DF 0F 49 0C 00 4A 29 37 76 C7 7D 8A 62 17 48 1F C3 A4 45 82 80 BC A1 CC 96 9B 38 CF C3 05 9E 09 91 9B C0 01 C7 0B 54 D3 09 3D 5E BB 34 75 62 23 11 A1 04 7F 4C B1 BD CF F9 4D B6 9F D0 15 B6 9E D2 AB C1 0D D9 81 7B 23 6F 60 74 CD 5B 0A DD 3F 65 24 0B 31 C7 10 FA 8E 3F AA 8E 7B 19 A7 EC C7 1F 21 8F 0A B8 02 C2 E1 E0 41 D9 49 A4 26 FC 06 FD 10 DB E2 A1 83 10 03 4B D9 7D ED 44 47 46 1A BD B4 F3 69 C3 DD C0 CB 85 50 CD B4 7B E9 EC 46 48 FF 9E B8 C8 C4 39 D1 BC 13 48 AF FE 51 BF 0F 38 9C 14 2C A9 E0 A9 60 DF 3C 18 67 F9 EF 6F 74 C3 BD 6A 45 A9 EB CC 7C 1E AF 09 43 15 28 32 36 8D 84 4E 8E B1 DC 64 A5 CB CF 19 39 80 42 E4 96 1B B3 51 7D FF 7D 9A 24 74 7D 44 3A FA BA 65 EC B9 63 E4 E8 B3 88 55 C0 05 2A FA 39 83 9B 9F 25 65 E4 1F 27 2E BA F5 A6 C6 8B 7F 5D 97 9B B1 93 C9 61 EA 18 EA 63 B2 4A 75 84 E3 72 DD 44 FC 1D 6F 7B A7 26 22 F1 C5 E3 87 8A 5C 45 57 9B 0B 74 25 5A D9 57 A4 B9 75 A6 85 0F 9B 2D B4 28 D6 34 CF C8 5D B6 7E 9D BB 8D 21 7C 3F 6A 8C AE 52 9E 02 2F 92 26 EF 5C F1 FD B8 53 F3 6C E8 0C 10 65 F9 D7 D3 0B 89 9F 0D 9F 30 3A BD CE 59 49 3A B7 E3 3B F9 1D 49 BA 30 A6 B0 6F DB 15 4D 83 D9 4A DB B3 4C F0 83 0D BC E6 99 45 B6 E4 C3 61 9C D1 02 F2 BE 56 4F A0 E3 82 02 00 69 C6 59 22 4F B1 4A A6 24 2F 32 4C 5A D6 D0 76 5F F3 A7 D4 A8 7A 7A 0A 8D 31 B9 CE 68 62 3C 41 13 06 94 50 78 41 8C C1 9A 11 46 F0 06 C9 D8 41 05 32 E3 95 F2 78 41 97 87 7D 00 15 C3 44 3B C1 ED 49 2E E0 71 74 C7 57 6F 4B C6 72 3A 9A 1F 04 38 DF A5 52 10 03 EF 39 93 8D 5C 66 99 F6 0A A7 A7 D6 71 9F CD C4 C8 96 CC CA FE 4E 59 09 AB 98 44 5E 1E D0 CE EB 9B E3 13 F4 84 DC 61 74 B3 5E 45 06 7B E3 F9 C9 47 41 31 66 B0 25 8C 6B CD 09 FD 4B B5 29 B7 48 4E BB F1 C0 F3 82 0D C1 46 5F 2D 9F AD BE 68 1F D2 53 73 19 E9 AD 56 AF A2 A0 99 FA 4E 6A 84 B7 DF BE B6 1B C7 DA 90 51 4E D2 CB 9B CE 74 50 86 94 5C 6C 5B 50 EB 07 04 FD 8C A4 16 B9 99 CE A7 D9 45 5D CD 42 F3 77 B6 A7 6B CC EF 3C 73 2F A1 6B 31 03 DE 06 88 A2 EB 23 71 13 FA 96 9D 94 2F B4 8E 6C 75 0B C8 AE EB 40 17 80 99 B2 4F F4 F5 22 E9 42 40 18 D8 C4 04 66 31 DE 69 EF A1 1B B5 C9 99 24 85 CF BF 93 E8 70 AB 6D 3D B7 5B 97 63 D7 E9 3E 39 F0 7B 06 E2 5A AD 67 E5 1D DE 9D 50 21 7B 1B 24 69 0F 66 0A 12 A8 A5 D4 9F 90 A7 75 CE 01 E9 48 A7 EE B4 67 57 B3 37 B3 8F 5C 7B 05 78 DC CF 72 54 A7 4B 75 38 7A 97 F1 6B 77 C4 0D 29 04 5E AD 0D 18 3A B8 FF B6 C0 EA 76 EC BA 19 EB 26 2A 03 65 F6 A6 0F FB 2E 95 F8 13 1D 81 BA 59 21 94 4E 7A 99 20 84 80 99 B3 96 1F B0 01 00 F2 6B 5E D3 43 55 A3 C7 CD 70 1E 56 EA 09 B8 CD EA FB 3A 9B 06 1E 21 32 B6 56 BA 7E E4 5A 50 D5 93 EE C6 E5 95 BE 8E 32 15 7C 69 F7 C9 0C A8 F7 E0 C0 8E D9 62 24 94 5E 15 DF 6F 92 41 E5 1D 85 42 6B 9D 89 F5 E2 AC F1 36 5F 5D FB 0B 26 51 82 58 70 F1 3F 50 A5 52 B2 E2 57 F8 95 5D AB D7 57 4C 61 B0 B9 6A E7 16 6D 60 6D 37 E5 0E DC AC 82 3D 01 16 77 85 A0 E2 29 60 A6 79 A1 D9 07 C9 4D D5 DB 68 E7 74 A8 51 DA 37 3B 34 4A BC 0D B7 1C 01 AB 1E 4F 79 EA 47 55 0C 7D 2A 83 ED 64 50 20 39 92 25 B5 10 CE 1F 32 29 55 46 87 DD 0D 92 23 5B F4 A5 4E D1 B9 4D FE B5 29 30 C3 0F 3E B3 C1 9C B7 0D 38 12 76 0A 51 37 D9 90 4E 9D 34 B3 0A CC 3D 05 35 17 18 D9 A0 FD A0 5B 35 80 36 93 0F D0 84 92 83 69 7C 99 BD 26 03 B9 0F 1F 59 CA 36 A5 6A 5C 4B B4 BB 0C 93 EE C5 53 09 BF D6 7A CC AA B7 64 F5 76 B0 53 53 E5 67 3E 19 5D 0C 1C CE 13 33 FC 08 27 22 E4 B9 6D C6 66 75 61 C4 58 CB 39 5B 56 0C EE AC 30 F9 13 AE 1E 3A 6A 39 BA AC DF 7F BA BC 7B 71 26 36 87 C1 57 9E 99 57 B5 AE 40 2D CB 3C 67 1A 5F 69 EB C3 19 5E EA 24 BD 63 1B 11 22 31 C4 72 52 A6 EC 2B 8F 0B A4 11 A4 2E 67 83 3F EC 30 C2 E4 85 9A 9B 83 BE F8 1D F9 55 22 EB D3 1C AB DE 36 55 35 D7 10 73 22 4A 20 CC D0 FB B4 47 6C 16 08 01 8D FE 46 71 98 31 4A 0D 2A 8C 1E 95 6C 20 83 33 B9 FC 4A 92 C8 75 53 30 45 47 1F 1B 7A 12 A5 AF CB 19 B3 65 F5 D6 AA ED E5 CD AD C3 00 5F 35 CA 15 90 6B 4F B2 A9 75 F1 74 71 F5 11 8F 80 EA 98 9D 49 64 13 10 5F E0 62 C3 BE 48 2C 38 F8 ED AF 24 36 15 42 EE E0 13 43 75 01 3B 8B 2B 30 7D D7 4E 14 9D 3E E5 BE 6F 10 E1 ED DC 0A 10 30 5F 01 0C 32 48 75 64 8B 8B 9F FA 80 23 F6 AD 5E F7 50 B5 62 23 C5 05 06 F2 D3 98 1D F7 03 93 1A 2F 00 74 53 E5 59 2D 79 81 4B 84 A1 4C AC 52 1F C0 D5 E0 32 B0 1B D8 3D 38 6C B7 37 6F E6 D2 C7 E0 9A 67 EB 29 73 D8 B8 0D 03 97 38 55 D3 1E 17 14 32 BC 5C 87 DD 4F A5 48 1B 55 05 4A 3E 6B EA A5 55 C0 AE 5A A1 11 93 65 60 71 95 A4 DB F8 E5 9D CB 7D 70 9A AE D5 05 43 46 12 7D DC FD DD E1 D3 52 09 5C F0 79 33 0C AE 7C 1D 6A 97 35 B8 82 22 70 E4 9D 89 6F 27 57 58 18 DF 62 DF 05 B8 A8 7A 17 60 D5 5A E9 D6 19 60 81 0C D4 17 9F 4C 29 F0 FC B2 F9 67 39 01 73 B8 80 79 27 A0 1E 96 74 FA 1F 91 FB 0D BC 59 E8 EA 42 23 F4 80 DC A6 EA 3E 0A A6 20 A7 0E D5 DE C3 F8 42 4A 02 8D E2 77 1B AE 6F 25 A1 AB D6 D6 D0 6A F6 B2 18 18 86 E0 E3 82 9F 0B EA 23 6A 00 D7 93 F7 52 7A 3C 33 ED DE A3 6B 14 0F 92 21 49 B5 84 0B 00 87 E3 97 05 22 74 74 74 78 65 10 9C 90 6E 25 7C 65 18 FE 14 B6 FF 07 B1 27 9A 11 83 CD AA 05 00 00 00 00 49 45 4E 44 AE 42 60 82 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 21 23 EB 61 C1 02 00 00 A5 05 00 00 18 00 00 00 78 6C 2F 64 72 61 77 69 6E 67 73 2F 64 72 61 77 69 6E 67 32 2E 78 6D 6C 9C 94 5D 6F 9B 30 14 86 EF 27 ED 3F 20 DF 53 0C 71 48 40 21 55 3E 60 AB D4 6D D5 D4 5D 4F 8E 31 0D 12 D8 C8 76 93 54 53 FF FB 8E 0D 24 8A D6 4A DB AE 72 E2 73 7C 3E 9E F3 9A C5 ED A9 6D BC 03 57 BA 96 22 43 E1 0D 46 1E 17 4C 96 B5 78 CA D0 8F C7 C2 9F 23 4F 1B 2A 4A DA 48 C1 33 F4 C2 35 BA 5D 7E FC B0 38 95 2A 3D EA AD F2 20 81 D0 29 FC CD D0 DE 98 2E 0D 02 CD F6 BC A5 FA 46 76 5C 80 B7 92 AA A5 06 FE AA A7 A0 54 F4 08 A9 DB 26 88 30 8E 03 DD 29 4E 4B BD E7 DC 6C 7B 0F 1A F2 D1 FF C8 D6 D2 5A A0 A5 EB CC 1C E5 86 37 CD 4A B0 BD 54 1E 2F 6B B3 D2 19 82 09 EC E9 10 53 29 D9 F6 D1 4C 36 4B BC 08 EC 48 D6 74 19 C0 F8 56 55 CB 88 CC E2 E9 C5 67 8F 9C 5B C9 E3 32 EC AF 58 73 3C B3 FE 30 9A CC A3 E9 D9 E7 AE B8 DC 97 82 46 5E 0A 4F DE 2E 0C 75 A3 F7 0A 5F 92 5F 17 0E 27 71 FC 56 E1 B1 5C 57 B3 FE 82 38 3C D4 EC 41 0D 4D 7C 3D 3C 28 AF 2E 33 44 90 27 68 0B 5B 66 A6 C1 F8 27 E0 32 75 F9 19 36 B4 13 42 45 C8 2B B9 66 B0 E6 C7 17 2D 85 77 27 8C A2 10 01 34 69 CA 4F E6 5E 9B C1 F2 9E 55 9D A1 5F 45 11 AD A7 79 41 FC 02 2C 9F E0 35 F1 D7 39 49 FC 02 F8 E4 D1 AC D8 44 93 F8 D5 DE 0E E3 94 81 0E 0C 48 F0 AE 1C F7 1F C6 7F 28 A0 AD 99 92 5A 56 E6 86 C9 36 90 55 55 33 3E 2A 0A F4 14 92 C0 29 C0 8D F2 6B 1D E7 AB 28 5A E5 FE 0C E7 89 4F 08 F4 91 6C 8B A9 9F 24 AB 24 CF E7 C9 04 4F 67 AF 28 58 2E 02 D7 FD F8 EB A6 E8 95 60 B1 5C 08 59 5E 36 DA 8A E4 0A DF AE A9 BB A2 6E 40 36 34 B5 F6 30 C0 5F 3D 87 7E 86 AD 64 CF 2D 17 A6 7F 13 8A 37 0E 85 DE D7 9D 46 9E 4A 79 BB E3 B0 1C 75 57 86 EF B2 8E E6 2B 8C 93 68 ED 6F A6 78 03 AC 67 B9 BF 4A C8 CC 0E 3F 23 98 CC C3 4D B8 E9 59 93 F4 59 F3 7B C9 68 B3 ED EA 33 6C F2 CF B0 F1 00 FB 40 9B 0C E1 F7 40 F6 48 2C 1A AD D8 77 CE 0C 20 04 DB 28 6E D8 DE 9A 15 90 1B CE 61 11 67 87 C3 7C 21 6B A1 EB 0E 54 BA 3B 7E 91 25 08 94 3E 1B E9 60 9C 2A 05 8F 98 A6 40 D2 3B 65 A8 7F AE C8 7B C9 D0 24 24 93 68 6A 1B 73 0B F6 98 75 C3 D2 09 86 6F 1C 83 80 59 0C EF C5 05 40 E5 31 4F A7 B4 F9 C4 65 EB 59 03 A0 43 CB AE 0E 3D 80 32 7A B5 8C 21 36 B1 90 76 F5 AE 46 23 AE 0E 20 A7 3D 71 83 D8 D6 07 F3 FC 04 59 53 C3 CA B7 D4 D0 51 55 57 DF AC 21 DC 7E 61 97 BF 01 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 39 31 B5 91 DB 00 00 00 D0 01 00 00 23 00 00 00 78 6C 2F 77 6F 72 6B 73 68 65 65 74 73 2F 5F 72 65 6C 73 2F 73 68 65 65 74 31 2E 78 6D 6C 2E 72 65 6C 73 AC 91 CD 6A C3 30 0C 80 EF 83 BE 83 D1 BD 76 D2 C3 18 A3 4E 2F 63 D0 EB DA 3D 80 67 2B 89 59 22 1B 4B 5B D7 B7 9F 77 28 2C A5 B0 CB 6E FA 41 9F 3E A1 ED EE 6B 9E D4 27 16 8E 89 2C B4 BA 01 85 E4 53 88 34 58 78 3D 3E AF 1F 40 B1 38 0A 6E 4A 84 16 CE C8 B0 EB 56 77 DB 17 9C 9C D4 21 1E 63 66 55 29 C4 16 46 91 FC 68 0C FB 11 67 C7 3A 65 A4 DA E9 53 99 9D D4 B4 0C 26 3B FF EE 06 34 9B A6 B9 37 E5 37 03 BA 05 53 ED 83 85 B2 0F 1B 50 C7 73 AE 9B FF 66 A7 BE 8F 1E 9F 92 FF 98 91 E4 C6 0A 13 8A 3B D5 CB 2A D2 95 01 C5 82 D6 97 1A 5F 82 56 57 65 30 B7 6D DA FF B4 C9 25 92 60 39 A0 48 95 E2 85 D5 55 CF 5C E5 AD 7E 8B F4 23 69 16 7F E8 BE 01 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 3E 74 50 E3 DB 00 00 00 D0 01 00 00 23 00 00 00 78 6C 2F 77 6F 72 6B 73 68 65 65 74 73 2F 5F 72 65 6C 73 2F 73 68 65 65 74 32 2E 78 6D 6C 2E 72 65 6C 73 AC 91 CD 6A C3 30 0C 80 EF 83 BE 83 D1 BD 76 9A C3 18 A3 4E 2F 63 D0 EB DA 3D 80 67 2B 89 59 22 1B 4B 5B D7 B7 9F 77 28 2C A5 B0 CB 6E FA 41 9F 3E A1 ED EE 6B 9E D4 27 16 8E 89 2C 6C 74 03 0A C9 A7 10 69 B0 F0 7A 7C 5E 3F 80 62 71 14 DC 94 08 2D 9C 91 61 D7 AD EE B6 2F 38 39 A9 43 3C C6 CC AA 52 88 2D 8C 22 F9 D1 18 F6 23 CE 8E 75 CA 48 B5 D3 A7 32 3B A9 69 19 4C 76 FE DD 0D 68 DA A6 B9 37 E5 37 03 BA 05 53 ED 83 85 B2 0F 2D A8 E3 39 D7 CD 7F B3 53 DF 47 8F 4F C9 7F CC 48 72 63 85 09 C5 9D EA 65 15 E9 CA 80 62 41 EB 4B 8D 2F 41 AB AB 32 98 DB 36 9B FF B4 C9 25 92 60 39 A0 48 95 E2 85 D5 55 CF 5C E5 AD 7E 8B F4 23 69 16 7F E8 BE 01 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 2F 2C F3 C8 BE 00 00 00 24 01 00 00 23 00 00 00 78 6C 2F 64 72 61 77 69 6E 67 73 2F 5F 72 65 6C 73 2F 64 72 61 77 69 6E 67 31 2E 78 6D 6C 2E 72 65 6C 73 84 8F 41 6A 03 31 0C 45 F7 85 DE C1 68 5F 6B A6 8B 50 CA 78 B2 29 81 6C 4B 72 00 61 6B 3C A6 63 D9 D8 4E 48 6E 5F 43 37 0D 14 BA D4 FF FC F7 D0 B4 BF C5 4D 5D B9 D4 90 C4 C0 A8 07 50 2C 36 B9 20 DE C0 F9 74 78 79 03 55 1B 89 A3 2D 09 1B B8 73 85 FD FC FC 34 7D F2 46 AD 8F EA 1A 72 55 9D 22 D5 C0 DA 5A 7E 47 AC 76 E5 48 55 A7 CC D2 9B 25 95 48 AD 9F C5 63 26 FB 45 9E F1 75 18 76 58 7E 33 60 7E 60 AA A3 33 50 8E 6E 04 75 BA E7 6E FE 9F 9D 96 25 58 FE 48 F6 12 59 DA 1F 0A 0C B1 BB 3B 90 8A E7 66 40 6B 8C EC 02 FD E4 A3 CE E2 01 E7 09 1F 7E 9B BF 01 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 2F 2C F3 C8 BE 00 00 00 24 01 00 00 23 00 00 00 78 6C 2F 64 72 61 77 69 6E 67 73 2F 5F 72 65 6C 73 2F 64 72 61 77 69 6E 67 32 2E 78 6D 6C 2E 72 65 6C 73 84 8F 41 6A 03 31 0C 45 F7 85 DE C1 68 5F 6B A6 8B 50 CA 78 B2 29 81 6C 4B 72 00 61 6B 3C A6 63 D9 D8 4E 48 6E 5F 43 37 0D 14 BA D4 FF FC F7 D0 B4 BF C5 4D 5D B9 D4 90 C4 C0 A8 07 50 2C 36 B9 20 DE C0 F9 74 78 79 03 55 1B 89 A3 2D 09 1B B8 73 85 FD FC FC 34 7D F2 46 AD 8F EA 1A 72 55 9D 22 D5 C0 DA 5A 7E 47 AC 76 E5 48 55 A7 CC D2 9B 25 95 48 AD 9F C5 63 26 FB 45 9E F1 75 18 76 58 7E 33 60 7E 60 AA A3 33 50 8E 6E 04 75 BA E7 6E FE 9F 9D 96 25 58 FE 48 F6 12 59 DA 1F 0A 0C B1 BB 3B 90 8A E7 66 40 6B 8C EC 02 FD E4 A3 CE E2 01 E7 09 1F 7E 9B BF 01 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 9D E3 B8 9A 26 08 00 00 28 21 00 00 27 00 00 00 78 6C 2F 70 72 69 6E 74 65 72 53 65 74 74 69 6E 67 73 2F 70 72 69 6E 74 65 72 53 65 74 74 69 6E 67 73 31 2E 62 69 6E EC 57 7F 70 14 D5 1D 7F 01 AC 26 52 0B 06 C7 94 10 8B FC 30 38 0C F9 6D 4C D4 A1 DE DD EE E5 EE 66 EF 76 73 BB 47 A8 96 1E 9B BB 77 C9 C2 DE EE 76 7F 84 44 C6 40 1D 18 10 14 11 95 1F 0D 8A 8C 0E 74 46 01 09 1D AD 50 C1 46 5B 22 45 19 32 D0 96 11 02 6D 51 03 96 C6 A6 24 E0 50 48 BF FB 36 B9 49 02 C3 B4 CE 38 FE 73 7B 73 BB EF 7D DF F7 BD EF 77 3F EF F3 FD BC 3B 0E F9 11 07 9F 00 2A 42 25 E8 FF BF 32 C6 A0 EF 9D 44 CC 64 4F FF D2 5B 47 A1 0C 74 3E 4B BD 2D 0E CF 1F A0 B9 A3 EC FE DC 51 A3 E1 CE 20 8C 4C F8 60 A4 7F 83 18 23 A7 64 0C 18 EC E7 28 F8 4E 82 46 3F 5C 23 FD 28 7F 28 32 05 CD C8 8A 8F ED CA CE 7F 38 F7 E8 CD 42 0B 43 06 ED 8C 27 92 B5 ED D5 6F 81 18 A3 D1 98 D4 7D 0C 69 0F 5D 2B 03 FC BF C9 35 7C 9E 1D 2B 7D 7D 97 08 0C DD C5 19 59 08 F1 41 21 60 E7 33 0E E5 64 F9 A0 46 26 A3 08 52 90 84 1A 08 8F 0D 24 22 19 6C 1C 70 5A 02 BB 49 EE 75 C4 E2 01 C6 4F 46 E5 F0 9D 01 DE E5 A8 00 95 C1 B7 18 DD 0F AB F9 15 CD 32 DD 92 82 BC 6C 38 C8 B3 91 B0 87 46 61 9A A7 18 06 45 14 49 C7 86 DD F2 71 94 1A 8B F8 F9 88 1F 09 BA 85 11 CD 73 61 BF C0 5B 9A A6 EA 26 8E 3B 46 AF 68 98 61 6C A0 8A 5A 4D 43 D0 50 65 CB 94 54 05 95 17 15 C5 35 09 71 BA A4 98 D5 96 28 4B 66 53 95 AE 5A 1A E2 AA C9 33 5A 5C 81 28 BD 49 90 92 18 AA 9E 17 4D 4B 17 C9 C4 22 E4 96 C5 D8 42 BF B2 90 D7 74 2C C6 51 1C 27 44 4B 36 21 1D 8F 2A AB 3A 9F 14 75 13 B9 2C 53 4D C2 84 D8 30 73 94 78 B0 9A BD 90 11 A5 E3 75 D8 A3 2A A6 AE CA 28 A4 EA 49 51 BE 89 B3 4F 94 13 A6 AA 60 F0 F4 24 E3 C3 1D 05 DC 68 46 43 D8 32 75 51 AE D2 C5 26 C3 49 91 55 E4 A6 1B 38 A6 56 A2 B0 29 4A 23 62 92 A5 C2 55 6E 92 E8 8D 62 C1 FA 5A BD 14 33 FE A7 78 29 E7 9B C6 4C 79 DD 2C 2E 57 AF 9A EA F0 A0 65 B3 9C 2C 87 02 1F 75 FC 6E 1A CF 71 19 11 8C D5 25 AC 98 CE 16 73 6C 58 08 BB FC 02 80 E7 98 C3 2A 8C E0 E2 8A 22 E4 15 65 03 23 CA D2 64 DC 88 42 6C 88 06 1F C2 21 56 71 AB 66 3D 2F C5 B1 11 14 15 20 14 60 EF 38 73 A2 86 75 5E 7A 02 23 86 16 04 3A 4C B6 44 31 AC 24 D6 3D 96 01 34 21 0E 0E 5B 83 38 2E 89 42 93 86 11 C7 B8 FC 21 04 2F 28 43 64 C4 86 60 56 28 62 60 4A 4A 24 BC 92 6E 98 9C 08 E4 A9 57 A5 18 46 42 38 42 D2 00 0B DD 18 C3 0E BB BC 92 8C ED 48 1C 5D 5C 59 6E 27 39 6C D4 AF 98 58 4F 88 30 99 AF 57 17 0D 1F BB CE 99 51 17 D1 8A 4D B9 E1 7E 8E 79 0E D6 AF 1B 81 B4 1B B0 EE 57 0C AC 93 92 01 18 2D D3 29 E8 A1 30 7B 5D 1E 9A 62 6B 42 88 B3 94 58 BD A4 D4 01 DF 80 E0 BC 29 6A 72 AA E7 E3 06 FB 2C D4 B5 01 65 ED A0 4A C7 54 45 4D AA 71 3C D0 B7 89 EB 32 48 69 0E 58 5C 32 BC A3 02 E8 31 D8 84 96 CF AE 55 67 2E AC 69 17 A9 CF AA 05 A1 C1 66 94 F7 53 D1 92 F2 D2 A8 1B 9E 0F 14 97 45 7D A4 FF 00 E2 3C 5E 55 8E A7 52 49 BD 04 A9 6E C4 FB AE 33 8C 78 11 DF B0 E9 44 AE 38 5D D5 6C 15 B2 F4 18 A6 44 53 44 F5 5A AC 5E 86 2D 2A 88 89 B5 00 24 A3 49 3C 96 71 8C E0 46 E0 08 70 74 15 AD 88 B5 B0 9D B6 A8 80 0F 2D 1A 4D 0E F3 59 05 BA 41 1F 88 61 48 04 9D F2 71 5E FF C0 76 47 29 BF 9B 37 25 33 56 1F 15 40 E1 04 6C 98 88 92 0C 7B 19 7B 27 53 3C 13 74 8C 1B 24 BC 88 53 35 D0 3E 22 A3 03 E5 14 B4 B1 F5 B0 0C 1B 8E 06 59 8A B6 A9 A8 EA C4 58 52 66 6B A9 20 00 DC 92 99 14 35 83 07 80 6D 94 04 C1 1E 77 8C 36 29 28 C6 A1 32 C5 44 39 0F 43 58 18 60 68 25 A6 12 4C 23 82 B7 02 BC 02 6A AD 2B 16 53 2D 85 2C 01 7D D6 ED F2 78 84 28 79 86 84 68 55 98 8D 70 21 57 D0 A6 B8 5B D5 15 56 01 DC EC 77 75 B3 14 29 25 48 8B 93 45 20 8E CD 76 B8 81 32 73 71 A2 83 7C 4C 04 D4 1C C1 E4 7D 61 A9 AE DE B4 35 97 10 8A 6C 15 22 B6 21 7D 92 8E BB 09 72 62 81 C1 3A 54 33 0A B8 03 6C 2A CD 1A 95 83 43 69 00 A6 B9 41 C6 AE 32 A8 CA B8 BD 8D 16 EC 62 79 41 63 D2 56 D4 08 15 B4 77 37 28 6A A4 0E 07 47 E3 C9 81 71 C2 3F 47 45 78 40 4E C6 76 69 41 D4 C1 A5 87 8C B3 F1 F8 88 41 7B 9E 14 03 10 60 82 BD EF A9 54 F9 1A 6F 8D CD 0F C2 96 B0 AB 86 D7 54 95 54 12 49 38 25 78 94 73 60 CD 32 C0 82 1C 0E 03 2D 13 B6 60 F0 BE 81 41 68 79 75 38 9B 46 54 B2 53 A0 3E 37 1C 80 37 1A 09 84 E1 3C 33 E0 18 02 CA 19 08 7A 3E CA 6B C9 32 B4 7C 14 C8 81 09 AA 08 58 D9 3D 36 91 20 CF 90 6A 5E 3F 00 7B 0B 7C AD 92 55 C3 40 83 F9 70 BC 4B 96 EA 94 24 88 F4 A0 B4 31 BC 23 6D 8E DE 3A 68 12 6D B4 D5 DF 91 6C 47 1B 53 0A 3C B0 98 57 D5 19 95 70 83 B6 A0 1E B1 A8 A4 7F 53 7E 9B 08 74 65 C3 2F BA 08 E7 9B 0E FF 01 1A 37 5F D8 F4 F1 82 EC EA 09 07 8F 04 A9 55 B9 2F 3D 75 FB C2 E9 59 77 9D 9A B0 9C 7B EC F1 17 A5 92 DD BF DF 55 9C 29 14 17 4C DB E4 9B F6 50 89 FF DE A7 6E 9F B7 F2 58 D5 57 E3 2F 4E 6F BD 98 71 72 D5 A7 1D F9 DA F6 FE 87 DF 7D 3A BB 34 7B E1 D6 3F 4C EB 3F AD FD E6 C1 2F 5E 69 79 F7 6C F3 66 43 69 A8 A3 5B DF DE B3 7B 72 A0 2C 73 EA CB 6C DE BA 6D 82 AB 65 5D 16 E5 6A E9 3A 47 29 4F CF 0C 3E B2 F3 F2 9D 2B D6 BD 17 EB 7C A6 FD B9 71 7B FB 8F FE E5 9E 0F AA C7 5B 77 E1 1C 71 C3 BA 15 3D 3F 7D 64 E5 CC E7 EE B8 DC 5C B2 5F BB 96 B9 E9 67 1D E3 0E 1E DB 37 7B EB 29 E5 D2 A1 ED 1F AD 0C 2D CB 79 F1 F9 43 8F FE F1 CB 2F AF 9C CB BD 2F 83 0E F7 1E 2C 12 73 F3 F2 77 AC 9E B5 7E C2 B8 BA 13 3D 8F BE 7D 47 C5 82 92 FD 6F 8A D7 DE 19 2F 2E EE 61 EE 79 ED 4D FF 4F 76 3E FB 04 17 1F 9B 53 51 73 6C FE 3F 27 AC AF B8 58 69 4E FA EC B7 C5 1D BD A5 DB 9A 03 0F 35 CD AF D9 95 6C BB 20 76 94 E6 9D 6B 9D B6 E1 5F C2 D9 5B FE 54 B6 29 F4 F9 5B A1 D0 CC BC 17 8E FF 7C EF F1 B5 1B C6 BB DB 0B 97 B4 E5 6C 5C FB 81 67 C5 82 1F 9E CD F9 FE 0B F8 F5 33 CF 5E 7D FF EE 07 AB 0B 37 E6 36 5F FD 3A 53 6E 9F D6 B3 BD 9F 6D 13 9E B9 7A 61 F2 DD BD BB 13 13 D7 2E 78 EB 31 59 72 7D 31 E5 97 53 97 31 25 95 71 F9 B3 F7 EB 2E F7 DC B9 F8 C7 AB 9F 8C 2E 57 72 F7 1D F8 7A 4B E3 AA 9A 8A 5F 85 F3 FF FE E9 E9 D7 AE 9E 5C DB 97 F7 EB 2D 3F EA D8 D9 3D FB 93 FB E8 93 73 4E EC 28 DC 3B EB CF 5D 4B 0E 3C 5E 39 F6 5C E7 9E 65 DD F3 DA F0 C6 FD 57 36 7E F8 7C 67 4E A9 1E 09 FC E2 C0 A5 33 A7 9E 3C FC 4A 5F 7F E5 E6 7F 58 87 13 A3 7B 57 57 1C DD 80 FD 05 BF 2B 38 FF E1 81 D5 AD 13 BB E6 CD 3D BB F5 D2 AB 85 BD E7 76 BD BA 24 73 FF 5F 67 77 D6 4D 5D 13 1D 7B BE E1 3F 47 02 D7 2A 0F B7 6E 6B 99 92 78 6F E9 BC C5 B7 FE 3B 3A A9 77 7E 7B FB E5 16 66 29 53 D2 B2 E6 A5 8F 3E 7F A7 75 79 EE 9C E2 58 65 8B F1 C6 16 BD BB 63 5F 5F 4B E1 E9 F5 75 DD 3B 3A 27 75 71 57 BE EA BD B0 BC BB EF 93 B6 CE 43 07 B3 77 1C 59 B3 67 BF 78 A2 29 FF 6F 6F 64 DE BB 6C 9C FB 36 FB D6 F7 31 7B FC FE E6 33 E9 5A 4E 23 90 46 20 8D 40 1A 81 34 02 69 04 D2 08 A4 11 48 23 90 46 20 8D 40 1A 81 34 02 69 04 D2 08 A4 11 F8 B6 10 F8 2F 00 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 46 5D 29 6D B3 01 00 00 D5 12 00 00 27 00 00 00 78 6C 2F 70 72 69 6E 74 65 72 53 65 74 74 69 6E 67 73 2F 70 72 69 6E 74 65 72 53 65 74 74 69 6E 67 73 32 2E 62 69 6E EC 58 DD 4A 02 41 14 FE 5C 0D 45 58 24 E8 01 7C 81 A2 34 BC 08 82 4C D7 3F FC 59 DC AD 0C 16 42 B4 40 88 84 52 21 F0 21 7A 83 DE C7 67 E8 09 BA EA 2E BA A8 ED DB 51 B3 6C A5 D4 AC 0D F7 80 B3 73 CE 9C 99 3D F3 CD D9 39 1F 66 A0 22 8C 3C AA B8 C2 29 2E 91 63 DB A2 45 65 BF C9 67 01 DB D8 44 04 EB 83 5E 94 B6 0B 6A 75 B6 96 78 7C 6B F2 1D 9E 56 43 8F 90 24 78 70 1F 6C 06 EA 7C 7A 50 91 38 CA D6 B2 2E 58 26 BC 40 42 AF 37 7C 73 82 BB 69 73 4F 0D B1 CB E9 E3 C9 70 DE 39 3A 02 9D 06 6A C4 6B 3A 09 DA B8 5B 08 8D CB C2 B1 5A F4 59 CC BC BE 1F A6 29 32 4A E4 CB 50 03 46 3D 27 85 BE A7 E6 F2 61 4D D1 C3 E9 83 AC A6 C7 CB 7A AE B4 BF BB 25 DB 1D B3 63 11 77 03 FB BF 08 BC CF 3F A5 98 5C CA EC 1B BF 2B A7 D5 BD 6F C7 1F 18 DC C4 9F 6F DF 5B 04 66 4A 92 FE 4A 0F 2B 7E DB EA D7 BF E9 26 2F 9C F2 DA D7 81 AF E6 7D 3F D4 04 76 60 E0 08 59 14 91 44 89 3D 8D BA 86 63 FE 74 28 AC F6 51 56 7D CB A2 72 B4 44 86 60 D0 AF 4C FF 43 8E 96 85 77 05 31 B2 03 83 9E 06 32 F4 B3 D6 89 09 B6 10 C7 09 FA A3 09 A4 90 C6 06 B5 02 D7 70 C5 45 E0 37 10 48 91 9F D5 C8 63 9B 64 7C D7 64 AC 49 F2 B6 33 DA DA E4 70 2D 32 DD 9F 12 19 F2 9F 1F A8 6F 2C 02 D3 7C 31 9F 05 93 72 46 AE 2D 2F A3 75 BF 75 17 01 17 01 87 22 10 F9 10 57 88 DA CD E8 A6 8A 8B 7F 0A AA AC 16 B3 49 17 DD B9 B6 3D EF FC 25 CF BA 57 00 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 EF 0F A5 62 ED 00 00 00 1E 02 00 00 10 00 00 00 78 6C 2F 63 61 6C 63 43 68 61 69 6E 2E 78 6D 6C 74 91 DD 4A C4 30 10 85 EF 05 DF 21 CC BD 9B A6 B8 8B 48 D3 85 5A BC C8 B5 FB 00 21 8D DB 42 7E 4A 12 44 DF DE 41 B7 E9 9A AE 37 81 9C CC 7C 73 E6 A4 39 7E 5A 43 3E 74 88 93 77 1C D8 AE 02 A2 9D F2 C3 E4 CE 1C 4E 6F AF 0F 4F 40 62 92 6E 90 C6 3B CD E1 4B 47 38 B6 F7 77 8D 92 46 BD 8C 72 72 04 09 2E 72 18 53 9A 9F 29 8D 6A D4 56 C6 9D 9F B5 C3 97 77 1F AC 4C 78 0D 67 1A E7 A0 E5 10 47 AD 93 35 B4 AE AA 03 B5 08 80 B6 51 24 70 E8 59 0D 64 E2 80 A7 41 2B 40 B3 FE 78 D1 57 E5 B0 51 D0 E7 4F EF 5A 83 AB DC A0 75 79 CA 52 D9 B1 A5 32 2B F5 C2 FF EB 44 B0 72 8A 60 A5 13 C1 4A B7 E2 C6 C4 92 D3 6D 38 FD FE 97 C3 CA 34 2E 3A A6 84 A1 AF 29 89 FD C6 49 26 2C 7B 89 7F 7A FB EB 5E 9A 7F B6 FD 06 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 DF DA DB F2 6B 01 00 00 AC 02 00 00 11 00 08 01 64 6F 63 50 72 6F 70 73 2F 63 6F 72 65 2E 78 6D 6C 20 A2 04 01 28 A0 00 01 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 7C 92 51 4B C3 30 14 85 DF 05 FF 43 C9 7B 9B A4 65 EA 42 57 65 CA 9E 14 04 2B 13 DF 42 72 B7 05 9B A4 24 D1 D9 7F 6F D6 75 75 4E F1 31 9C 73 3E CE 3D A4 BC FE D4 4D F2 01 CE 2B 6B 66 88 66 04 25 60 84 95 CA AC 67 E8 B9 5E A4 57 28 F1 81 1B C9 1B 6B 60 86 3A F0 E8 BA 3A 3F 2B 45 CB 84 75 F0 E8 6C 0B 2E 28 F0 49 24 19 CF 44 3B 43 9B 10 5A 86 B1 17 1B D0 DC 67 D1 61 A2 B8 B2 4E F3 10 9F 6E 8D 5B 2E DE F8 1A 70 4E C8 05 D6 10 B8 E4 81 E3 1D 30 6D 47 22 1A 90 52 8C C8 F6 DD 35 3D 40 0A 0C 0D 68 30 C1 63 9A 51 FC ED 0D E0 B4 FF 33 D0 2B 47 4E AD 42 D7 C6 9B 86 BA C7 6C 29 F6 E2 E8 FE F4 6A 34 6E B7 DB 6C 5B F4 35 62 7F 8A 5F 1E EE 9F FA 53 53 65 76 5B 09 40 55 29 05 13 0E 78 B0 AE 5A AA A6 51 5C 67 73 E0 BE 81 EE 26 74 DE 9A 4C 58 5D E2 23 D7 6E D1 86 FB F0 10 C7 5F 29 90 F3 AE 9A BB 8E 9B 64 88 95 F8 B7 E1 90 79 74 CA 04 90 55 4E E8 34 25 45 4A A7 35 A5 2C 27 8C 4E 5F C7 DC C1 14 AB F5 4B EC FB 81 4C E2 6D 6C BF C4 41 59 16 B7 77 F5 02 7D F3 2E EB 08 9B 4C 18 21 91 77 92 DF DD BA 07 EA A1 FA BF C4 9C A6 64 92 E6 93 9A 16 AC 18 1A 9E 02 AA BE F4 CF FF 55 7D 01 00 00 FF FF 03 00 50 4B 03 04 14 00 06 00 08 00 00 00 21 00 46 15 30 C4 89 01 00 00 22 03 00 00 10 00 08 01 64 6F 63 50 72 6F 70 73 2F 61 70 70 2E 78 6D 6C 20 A2 04 01 28 A0 00 01 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 9C 92 41 4F DC 30 10 85 EF 95 FA 1F 22 DF 59 87 6D 85 AA 95 63 04 0B 88 43 51 57 DA 85 9E 8D 33 D9 58 24 B6 E5 19 A2 5D 7E 7D 27 89 08 D9 72 A8 D4 DB CC BC D1 F3 E7 67 AB CB 43 DB 64 1D 24 74 C1 17 E2 7C 91 8B 0C BC 0D A5 F3 FB 42 3C EE EE CE 7E 88 0C C9 F8 D2 34 C1 43 21 8E 80 E2 52 7F FD A2 36 29 44 48 E4 00 33 B6 F0 58 88 9A 28 AE A4 44 5B 43 6B 70 C1 B2 67 A5 0A A9 35 C4 6D DA CB 50 55 CE C2 4D B0 AF 2D 78 92 CB 3C BF 90 70 20 F0 25 94 67 71 32 14 A3 E3 AA A3 FF 35 2D 83 ED F9 F0 69 77 8C 0C AC D5 55 8C 8D B3 86 F8 96 FA C1 D9 14 30 54 94 DD 1E 2C 34 4A CE 45 C5 74 5B B0 AF C9 D1 51 E7 4A CE 5B B5 B5 A6 81 35 1B EB CA 34 08 4A 7E 0C D4 3D 98 3E B4 8D 71 09 B5 EA 68 D5 81 A5 90 32 74 6F 1C DB 52 64 CF 06 A1 C7 29 44 67 92 33 9E 18 AB 5F 1B 9B A1 6E 22 52 D2 BF 43 7A C1 1A 80 50 49 5E 18 87 43 39 DF 9D D7 EE BB 5E 0E 0B 5C 9C 2E F6 06 23 08 0B A7 88 3B 47 0D E0 AF 6A 63 12 FD 8B 78 60 18 79 47 9C AB 39 DA 04 79 FD 09 78 C8 80 8F FE EB B0 75 68 A3 F1 47 16 A6 EA A7 F3 2F F8 18 77 E1 C6 10 BC E7 7B 3A 54 DB DA 24 28 F9 49 A6 FC A7 81 BA E7 68 53 D3 9B AC 6B E3 F7 50 BE EF 7C 16 FA DF F0 34 7E 79 7D 7E B1 C8 BF E5 FC D0 B3 99 92 1F 9F 5B FF 01 00 00 FF FF 03 00 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 64 31 2F CD A4 01 00 00 2E 07 00 00 13 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 00 5B 43 6F 6E 74 65 6E 74 5F 54 79 70 65 73 5D 2E 78 6D 6C 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 B5 55 30 23 F4 00 00 00 4C 02 00 00 0B 00 00 00 00 00 00 00 00 00 00 00 00 00 DD 03 00 00 5F 72 65 6C 73 2F 2E 72 65 6C 73 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 85 90 AC 37 8C 03 00 00 AA 08 00 00 0F 00 00 00 00 00 00 00 00 00 00 00 00 00 02 07 00 00 78 6C 2F 77 6F 72 6B 62 6F 6F 6B 2E 78 6D 6C 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 FE 69 EA 57 0A 01 00 00 CC 03 00 00 1A 00 00 00 00 00 00 00 00 00 00 00 00 00 BB 0A 00 00 78 6C 2F 5F 72 65 6C 73 2F 77 6F 72 6B 62 6F 6F 6B 2E 78 6D 6C 2E 72 65 6C 73 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 2D CC 2D 95 A9 1B 00 00 7D 97 00 00 18 00 00 00 00 00 00 00 00 00 00 00 00 00 05 0D 00 00 78 6C 2F 77 6F 72 6B 73 68 65 65 74 73 2F 73 68 65 65 74 31 2E 78 6D 6C 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 26 2B A3 14 87 1B 00 00 9E 9C 00 00 18 00 00 00 00 00 00 00 00 00 00 00 00 00 E4 28 00 00 78 6C 2F 77 6F 72 6B 73 68 65 65 74 73 2F 73 68 65 65 74 32 2E 78 6D 6C 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 C1 17 10 BE 4E 07 00 00 C6 20 00 00 13 00 00 00 00 00 00 00 00 00 00 00 00 00 A1 44 00 00 78 6C 2F 74 68 65 6D 65 2F 74 68 65 6D 65 31 2E 78 6D 6C 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 9E AC D5 76 F3 0C 00 00 ED D1 00 00 0D 00 00 00 00 00 00 00 00 00 00 00 00 00 20 4C 00 00 78 6C 2F 73 74 79 6C 65 73 2E 78 6D 6C 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 B4 56 B5 DE 31 0B 00 00 9E 26 00 00 14 00 00 00 00 00 00 00 00 00 00 00 00 00 3E 59 00 00 78 6C 2F 73 68 61 72 65 64 53 74 72 69 6E 67 73 2E 78 6D 6C 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 3A 3A C6 E7 C2 02 00 00 A5 05 00 00 18 00 00 00 00 00 00 00 00 00 00 00 00 00 A1 64 00 00 78 6C 2F 64 72 61 77 69 6E 67 73 2F 64 72 61 77 69 6E 67 31 2E 78 6D 6C 50 4B 01 02 2D 00 0A 00 00 00 00 00 00 00 21 00 CA 02 BC 85 BF 19 00 00 BF 19 00 00 13 00 00 00 00 00 00 00 00 00 00 00 00 00 99 67 00 00 78 6C 2F 6D 65 64 69 61 2F 69 6D 61 67 65 31 2E 70 6E 67 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 21 23 EB 61 C1 02 00 00 A5 05 00 00 18 00 00 00 00 00 00 00 00 00 00 00 00 00 89 81 00 00 78 6C 2F 64 72 61 77 69 6E 67 73 2F 64 72 61 77 69 6E 67 32 2E 78 6D 6C 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 39 31 B5 91 DB 00 00 00 D0 01 00 00 23 00 00 00 00 00 00 00 00 00 00 00 00 00 80 84 00 00 78 6C 2F 77 6F 72 6B 73 68 65 65 74 73 2F 5F 72 65 6C 73 2F 73 68 65 65 74 31 2E 78 6D 6C 2E 72 65 6C 73 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 3E 74 50 E3 DB 00 00 00 D0 01 00 00 23 00 00 00 00 00 00 00 00 00 00 00 00 00 9C 85 00 00 78 6C 2F 77 6F 72 6B 73 68 65 65 74 73 2F 5F 72 65 6C 73 2F 73 68 65 65 74 32 2E 78 6D 6C 2E 72 65 6C 73 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 2F 2C F3 C8 BE 00 00 00 24 01 00 00 23 00 00 00 00 00 00 00 00 00 00 00 00 00 B8 86 00 00 78 6C 2F 64 72 61 77 69 6E 67 73 2F 5F 72 65 6C 73 2F 64 72 61 77 69 6E 67 31 2E 78 6D 6C 2E 72 65 6C 73 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 2F 2C F3 C8 BE 00 00 00 24 01 00 00 23 00 00 00 00 00 00 00 00 00 00 00 00 00 B7 87 00 00 78 6C 2F 64 72 61 77 69 6E 67 73 2F 5F 72 65 6C 73 2F 64 72 61 77 69 6E 67 32 2E 78 6D 6C 2E 72 65 6C 73 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 9D E3 B8 9A 26 08 00 00 28 21 00 00 27 00 00 00 00 00 00 00 00 00 00 00 00 00 B6 88 00 00 78 6C 2F 70 72 69 6E 74 65 72 53 65 74 74 69 6E 67 73 2F 70 72 69 6E 74 65 72 53 65 74 74 69 6E 67 73 31 2E 62 69 6E 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 46 5D 29 6D B3 01 00 00 D5 12 00 00 27 00 00 00 00 00 00 00 00 00 00 00 00 00 21 91 00 00 78 6C 2F 70 72 69 6E 74 65 72 53 65 74 74 69 6E 67 73 2F 70 72 69 6E 74 65 72 53 65 74 74 69 6E 67 73 32 2E 62 69 6E 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 EF 0F A5 62 ED 00 00 00 1E 02 00 00 10 00 00 00 00 00 00 00 00 00 00 00 00 00 19 93 00 00 78 6C 2F 63 61 6C 63 43 68 61 69 6E 2E 78 6D 6C 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 DF DA DB F2 6B 01 00 00 AC 02 00 00 11 00 00 00 00 00 00 00 00 00 00 00 00 00 34 94 00 00 64 6F 63 50 72 6F 70 73 2F 63 6F 72 65 2E 78 6D 6C 50 4B 01 02 2D 00 14 00 06 00 08 00 00 00 21 00 46 15 30 C4 89 01 00 00 22 03 00 00 10 00 00 00 00 00 00 00 00 00 00 00 00 00 D6 96 00 00 64 6F 63 50 72 6F 70 73 2F 61 70 70 2E 78 6D 6C 50 4B 05 06 00 00 00 00 15 00 15 00 BF 05 00 00 95 99 00 00 00 00','</template>']
        with open('resources.txt', 'w') as f:
            for item in resourcedefault:
                f.write("%s\n" % item)
#Main Program
if __name__ == 'Main':
    with open('vers', 'r') as f:
        new_version=f.readlines()
        f.close()
        os.remove('vers')
    isgood = False
    console = Console()
    try:
        el1 = Elockouts()
        el2 = Elockouts()
        el3 = Elockouts()
        el4 = Elockouts()
        el5 = Elockouts()

        ol1 = Olockouts()
        ol2 = Olockouts()
        ol3 = Olockouts()
        ol4 = Olockouts()
        ol5 = Olockouts()
        ol6 = Olockouts()

        es1 = eStops()
        es2 = eStops()

        rl1 = Rlockouts()
        rl2 = Rlockouts()
        rl3 = Rlockouts()
        rl4 = Rlockouts()
        rl5 = Rlockouts()
        rl6 = Rlockouts()
        rl7 = Rlockouts()
        with open('version', 'rb') as fpv:
                version = pickle.load(fpv)
                fpv.close()
        console.print(Panel("[cyan]Lockout Card Builder\n\n\n[white]Put any images for the cards in a subfolder layout show below\nAll images will need to be named properly for the program to utilize them\nAn example would be [Panel 2-3-5 2.jpg]\nthe very last number in the name will tell the program what spot to place the image.\n\n[cyan]v"+version+"[white]"),justify="center")
        console.print("Runing system Checks", style="yellow",justify="center")
        console.print("Checking for required folders", style="yellow",justify="center")
     
        #Check for folders and files.
        savepath = r"new/"
        imagepath = r"images/"
        dir = savepath
        if check_exists(dir):
            console.print("Directory /new/ found!",style="green",justify="center")
        else:
            console.print("Directory /new/ is being created new cards will be in there",style="bright_white",justify="center")
            os.mkdir(savepath)
            if check_exists(dir):
                console.print("Directory /new/ found!",style="green",justify="center")
            else:
                console.print("RROR CANNOT CREATE DIRECTORY /new/ Check permissions and try again",style="red",justify="center")
        if check_exists(imagepath):
            console.print("Directory /images/ found!",style="green",justify="center")
        else:
            console.print("Directory /images/ being created.",style="bright_white",justify="center")
            os.mkdir(imagepath)
            if not check_exists(imagepath):
                console.print("ERROR CANNOT CREATE DIRECTORY /images/ Check permissions and try again",style="red",justify="center")
        console.print("Loading resources...",style="yellow",justify="center")
        #load resources text and add to lists
        try:
            with open('resources.txt') as f:
                pass
        except IOError:
            generate_resources()
        with open('resources.txt') as f:
            prRecordingMode = False
            vRecordingMode = False
            oRecordingMode = False
            proRecordingMode = False
            tmpRecordingMode = False
            timgRecordingMode = False
            for line in f:
                if not prRecordingMode:
                    if line.startswith('<panelrooms>'):
                        prRecordingMode = True
                elif line.startswith('</panelrooms>'):
                    prRecordingMode = False
                else:
                    prlist.append(line)
                    prlist = [x.strip('\n') for x in prlist]
                    prlist = [x.strip('\t') for x in prlist]

                if not vRecordingMode:
                    if line.startswith('<volts>'):
                        vRecordingMode = True
                elif line.startswith('</volts>'):
                    vRecordingMode = False
                else:
                    vlist.append(line)
                    vlist = [x.strip('\n') for x in vlist]
                    vlist = [x.strip('\t') for x in vlist]

                if not oRecordingMode:
                    if line.startswith('<otherlockouts>'):
                        oRecordingMode = True
                elif line.startswith('</otherlockouts>'):
                    oRecordingMode = False
                else:
                    olist.append(line)
                    olist = [x.strip('\n') for x in olist]
                    olist = [x.strip('\t') for x in olist]

                if not proRecordingMode:
                    if line.startswith('<ssProcedure>'):
                        proRecordingMode = True
                elif line.startswith('</ssProcedure>'):
                    proRecordingMode = False
                else:
                    prolist.append(line)
                    prolist = [x.strip('\n') for x in prolist]
                    prolist = [x.strip('\t') for x in prolist]
                if not tmpRecordingMode:
                    if line.startswith('<template>'):
                        tmpRecordingMode = True
                elif line.startswith('</template>'):
                    tmpRecordingMode = False
                else:
                    tempsave.append(line)
                    tempsave = [x.strip('\n') for x in tempsave]
                    tempsave = [x.strip('\t') for x in tempsave]
                if not timgRecordingMode:
                    if line.startswith('<timage>'):
                        timgRecordingMode = True
                elif line.startswith('</timage>'):
                    timgRecordingMode = False
                else:
                    timage.append(line)
                    timage = [x.strip('\n') for x in timage]
                    timage = [x.strip('\t') for x in timage]
        if os.path.isfile('template.xlsx'):
            pass
        else:
            f = open('template.xlsx', 'wb')
            for x in tempsave:
                f.write(
                binascii.unhexlify(''.join(x.split()))
                )
            f.close()
        console.print("Found Records...",style="yellow",justify="center")

        table = Table(title="Found Records")
        table.add_column("Panel Rooms", justify="center", style="cyan", no_wrap=True)
        table.add_column("Voltages", style="cyan", justify="center")
        table.add_column("Lockout Types", justify="center", style="cyan")
        table.add_row(str(len(prlist)), str(len(vlist)), str(len(olist)))
        console.print(table,justify="center")

        template_file = "template.xlsx"
        wb = load_workbook(template_file)
        ws = wb['A'] #Getting the sheet named as 'A'
        # Folder in which my source excel sheets are present
        # To get the list of excel files
    except:
        pass
    while isgood is False:
        if Confirm.ask("Would you like to reload your previous file?"):
            with open('crash.bak', 'rb') as fp:
                crashsave = pickle.load(fp)
            prnsave=[]
            equipnumber=crashsave[0]
            equipname=crashsave[1]
            sdtype=crashsave[2]
            hz=crashsave[3]
            prnsave.append(crashsave[4])
            elsave=crashsave[5]
            el1.svolts=crashsave[6]
            el1.bucket=crashsave[7]
            el1.sproced=crashsave[8]
            el2.svolts=crashsave[9]
            el2.bucket=crashsave[10]
            el2.sproced=crashsave[11]
            el3.svolts=crashsave[12]
            el3.bucket=crashsave[13]
            el3.sproced=crashsave[14]
            el4.svolts=crashsave[15]
            el4.bucket=crashsave[16]
            el4.sproced=crashsave[17]
            el5.svolts=crashsave[18]
            el5.bucket=crashsave[19]
            el5.sproced=crashsave[20]
            olsave=crashsave[21]
            ol1.sType=crashsave[22]
            ol1.cType=crashsave[23]
            ol1.name=crashsave[24]
            ol1.proced=crashsave[25]
            ol1.sproced=crashsave[26]
            ol2.sType=crashsave[27]
            ol2.cType=crashsave[28]
            ol2.name=crashsave[29]
            ol2.proced=crashsave[30]
            ol2.sproced=crashsave[31]
            ol3.sType=crashsave[32]
            ol3.cType=crashsave[33]
            ol3.name=crashsave[34]
            ol3.proced=crashsave[35]
            ol3.sproced=crashsave[36]
            ol4.sType=crashsave[37]
            ol4.cType=crashsave[38]
            ol4.name=crashsave[39]
            ol4.proced=crashsave[40]
            ol4.sproced=crashsave[41]
            ol5.sType=crashsave[42]
            ol5.cType=crashsave[43]
            ol5.name=crashsave[44]
            ol5.proced=crashsave[45]
            ol5.sproced=crashsave[46]
            ol6.sType=crashsave[47]
            ol6.cType=crashsave[48]
            ol6.name=crashsave[49]
            ol6.proced=crashsave[50]
            ol6.sproced=crashsave[51]
            essave=crashsave[52]
            es1.location=crashsave[53]
            es2.location=crashsave[54]
            rlsave=crashsave[55]
            rl1.location=crashsave[56]
            rl2.location=crashsave[57]
            rl3.location=crashsave[58]
            rl4.location=crashsave[59]
            rl5.location=crashsave[60]
            rl6.location=crashsave[61]
            rl7.location=crashsave[62]
            prnsave.append(crashsave[63])
            hz1=crashsave[64]
        else:
            pass
        console.print("Enter Equipment number to begin. Example: 000-000000",style="white",justify="center")
        equipnumber = Prompt.ask("Equipment Number",default=equipnumber)
        if os.path.isfile('bak/equip/'+equipnumber+'.bak'):
            if Confirm.ask("We found a Backup of ("+equipnumber+") Would you like to load it?"):
                with open('bak/equip/'+equipnumber+'.bak', 'rb') as fp:
                    crashsave = pickle.load(fp)
                prnsave=[]
                equipnumber=crashsave[0]
                equipname=crashsave[1]
                sdtype=crashsave[2]
                hz=crashsave[3]
                prnsave.append(crashsave[4])
                elsave=crashsave[5]
                el1.svolts=crashsave[6]
                el1.bucket=crashsave[7]
                el1.sproced=crashsave[8]
                el2.svolts=crashsave[9]
                el2.bucket=crashsave[10]
                el2.sproced=crashsave[11]
                el3.svolts=crashsave[12]
                el3.bucket=crashsave[13]
                el3.sproced=crashsave[14]
                el4.svolts=crashsave[15]
                el4.bucket=crashsave[16]
                el4.sproced=crashsave[17]
                el5.svolts=crashsave[18]
                el5.bucket=crashsave[19]
                el5.sproced=crashsave[20]
                olsave=crashsave[21]
                ol1.sType=crashsave[22]
                ol1.cType=crashsave[23]
                ol1.name=crashsave[24]
                ol1.proced=crashsave[25]
                ol1.sproced=crashsave[26]
                ol2.sType=crashsave[27]
                ol2.cType=crashsave[28]
                ol2.name=crashsave[29]
                ol2.proced=crashsave[30]
                ol2.sproced=crashsave[31]
                ol3.sType=crashsave[32]
                ol3.cType=crashsave[33]
                ol3.name=crashsave[34]
                ol3.proced=crashsave[35]
                ol3.sproced=crashsave[36]
                ol4.sType=crashsave[37]
                ol4.cType=crashsave[38]
                ol4.name=crashsave[39]
                ol4.proced=crashsave[40]
                ol4.sproced=crashsave[41]
                ol5.sType=crashsave[42]
                ol5.cType=crashsave[43]
                ol5.name=crashsave[44]
                ol5.proced=crashsave[45]
                ol5.sproced=crashsave[46]
                ol6.sType=crashsave[47]
                ol6.cType=crashsave[48]
                ol6.name=crashsave[49]
                ol6.proced=crashsave[50]
                ol6.sproced=crashsave[51]
                essave=crashsave[52]
                es1.location=crashsave[53]
                es2.location=crashsave[54]
                rlsave=crashsave[55]
                rl1.location=crashsave[56]
                rl2.location=crashsave[57]
                rl3.location=crashsave[58]
                rl4.location=crashsave[59]
                rl5.location=crashsave[60]
                rl6.location=crashsave[61]
                rl7.location=crashsave[62]
                prnsave.append(crashsave[63])
                hz1=crashsave[64]
            else:
                pass

        console.print("Checking for images."+equipnumber,style="yellow",justify="center")
        try:
            imgfolder = os.listdir(r'images/' + equipnumber + '/')
            console.print("Checking Photo Dimensions",style="yellow",justify="center")
            for file in imgfolder:
                corrected = imagepath+equipnumber+'/'+file
                resize_Image(corrected)
        except IOError as e:
            console.print("No images Found.\nGet off your @$$ and take some!",style="red",justify="center")
        equipname = Prompt.ask("\nEquipment Name",default=equipname)
        panelroom = select_PR(prnumber,prnsave)
        prnsave=panelroom
        el = elecLockouts(elsave)
        elsave=el
        if el != 0:
            if el == 1:
                el1.gen_elect()
            if el == 2:
                el1.gen_elect()
                el2.gen_elect()
            if el == 3:
                el1.gen_elect()
                el2.gen_elect()
                el3.gen_elect()
            if el == 4:
                el1.gen_elect()
                el2.gen_elect()
                el3.gen_elect()
                el4.gen_elect()
            if el == 5:
                el1.gen_elect()
                el2.gen_elect()
                el3.gen_elect()
                el4.gen_elect()
                el5.gen_elect()
        ol = otherLockouts(olsave)
        olsave=ol
        if ol != 0:
            if ol == 1:
                ol1.gen_other()
            if ol == 2:
                ol1.gen_other()
                ol2.gen_other()
            if ol == 3:
                ol1.gen_other()
                ol2.gen_other()
                ol3.gen_other()
            if ol == 4:
                ol1.gen_other()
                ol2.gen_other()
                ol3.gen_other()
                ol4.gen_other()
            if ol == 5:
                ol1.gen_other()
                ol2.gen_other()
                ol3.gen_other()
                ol4.gen_other()
                ol5.gen_other()
            if ol == 6:
                ol1.gen_other()
                ol2.gen_other()
                ol3.gen_other()
                ol4.gen_other()
                ol5.gen_other()
                ol6.gen_other()
        es = eStop_count(essave)
        essave=es
        if es != 0:
            if es == 1:
                console.print("Where is the First eStop?",style="red")
                es1.gen_eStops()
            if es == 2:
                console.print("Where is the First eStop?",style="red")
                es1.gen_eStops()
                console.print("Where is the Second eStop?",style="red")
                es2.gen_eStops()
        rl = remoteLockouts(rlsave)
        rlsave=rl
        if rl != 0:
            if rl == 1:
                console.print("Where is the 1st Remote Lockout?",style="bright_white", justify="left")
                rl1.gen_remoteLockouts()
            if rl == 2:
                console.print("Where is the 1st Remote Lockout?",style="bright_white", justify="left")
                rl1.gen_remoteLockouts()
                console.print("Where is the 2nd Remote Lockout?",style="bright_white", justify="left")
                rl2.gen_remoteLockouts()
            if rl == 3:
                console.print("Where is the 1st Remote Lockout?",style="bright_white", justify="left")
                rl1.gen_remoteLockouts()
                console.print("Where is the 2nd Remote Lockout?",style="bright_white", justify="left")
                rl2.gen_remoteLockouts()
                console.print("Where is the 3rd Remote Lockout?",style="bright_white", justify="left")
                rl3.gen_remoteLockouts()
            if rl == 4:
                console.print("Where is the 1st Remote Lockout?",style="bright_white", justify="left")
                rl1.gen_remoteLockouts()
                console.print("Where is the 2nd Remote Lockout?",style="bright_white", justify="left")
                rl2.gen_remoteLockouts()
                console.print("Where is the 3rd Remote Lockout?",style="bright_white", justify="left")
                rl3.gen_remoteLockouts()
                console.print("Where is the 4th Remote Lockout?",style="bright_white", justify="left")
                rl4.gen_remoteLockouts()
            if rl == 5:
                console.print("Where is the 1st Remote Lockout?",style="bright_white", justify="left")
                rl1.gen_remoteLockouts()
                console.print("Where is the 2nd Remote Lockout?",style="bright_white", justify="left")
                rl2.gen_remoteLockouts()
                console.print("Where is the 3rd Remote Lockout?",style="bright_white", justify="left")
                rl3.gen_remoteLockouts()
                console.print("Where is the 4th Remote Lockout?",style="bright_white", justify="left")
                rl4.gen_remoteLockouts()
                console.print("Where is the 5th Remote Lockout?",style="bright_white", justify="left")
                rl5.gen_remoteLockouts()
            if rl == 6:
                console.print("Where is the 1st Remote Lockout?",style="bright_white", justify="left")
                rl1.gen_remoteLockouts()
                console.print("Where is the 2nd Remote Lockout?",style="bright_white", justify="left")
                rl2.gen_remoteLockouts()
                console.print("Where is the 3rd Remote Lockout?",style="bright_white", justify="left")
                rl3.gen_remoteLockouts()
                console.print("Where is the 4th Remote Lockout?",style="bright_white", justify="left")
                rl4.gen_remoteLockouts()
                console.print("Where is the 5th Remote Lockout?",style="bright_white", justify="left")
                rl5.gen_remoteLockouts()
                console.print("Where is the 6th Remote Lockout?",style="bright_white", justify="left")
                rl6.gen_remoteLockouts()
            if rl == 7:
                console.print("Where is the 1st Remote Lockout?",style="bright_white", justify="left")
                rl1.gen_remoteLockouts()
                console.print("Where is the 2nd Remote Lockout?",style="bright_white", justify="left")
                rl2.gen_remoteLockouts()
                console.print("Where is the 3rd Remote Lockout?",style="bright_white", justify="left")
                rl3.gen_remoteLockouts()
                console.print("Where is the 4th Remote Lockout?",style="bright_white", justify="left")
                rl4.gen_remoteLockouts()
                console.print("Where is the 5th Remote Lockout?",style="bright_white", justify="left")
                rl5.gen_remoteLockouts()
                console.print("Where is the 6th Remote Lockout?",style="bright_white", justify="left")
                rl6.gen_remoteLockouts()
                console.print("Where is the 7th Remote Lockout?",style="bright_white", justify="left")
                rl7.gen_remoteLockouts()
        sp = shutdownprocedure(sdtype)
        sdtype=sp
        hz = hazard(hz)

        #images
        try:
            imagefn = glob.glob(r'images/' + equipnumber + '/* 1.jpg')
            if imagefn:
                imagesingle = imagefn.pop()
                img = Image.open(imagesingle)
                img3 = openpyxl.drawing.image.Image(img)
                Tempname = imagesingle.strip('.jpg')
                Tempname = Tempname.strip('.JPG')
                tempname = Tempname.strip('1')
                Tempname = Tempname.strip('images/' + equipnumber + '/')
                Tempname = Tempname.strip('\\')
                Iname1=Tempname
                ws['A37'] = Iname1
                ws.add_image(img3,'A26')
            else:
                pass
        except IOError as e:
            console.print("No 1st Image",style="red",justify="center") #Does not exist OR no read permissions
        try:
            imagefn = glob.glob(r'images/' + equipnumber + '/* 2.jpg')
            if imagefn:
                imagesingle = imagefn.pop()
                img = Image.open(imagesingle)
                img3 = openpyxl.drawing.image.Image(img)
                Tempname = imagesingle.strip('.jpg')
                Tempname = Tempname.strip('.JPG')
                tempname = Tempname.strip('2')
                Tempname = Tempname.strip('images/' + equipnumber + '/')
                Tempname = Tempname.strip('\\')
                Iname2 = Tempname
                ws['F37'] = Iname2
                ws.add_image(img3,'F26')
            else:
                pass
        except IOError as e:
            console.print("No 2nd Image",style="red",justify="center") #Does not exist OR no read permissions
        try:
            imagefn = glob.glob(r'images/' + equipnumber + '/* 3.jpg')

            if imagefn:
                imagesingle = imagefn.pop()
                img = Image.open(imagesingle)
                img3 = openpyxl.drawing.image.Image(img)
                Tempname = imagesingle.strip('.jpg')
                Tempname = Tempname.strip('.JPG')
                tempname = Tempname.strip('3')
                Tempname = Tempname.strip('images/' + equipnumber + '/')
                Tempname = Tempname.strip('\\')
                Iname3 = Tempname
                ws['K37'] = Iname3
                ws.add_image(img3,'K26')
            else:
                pass
        except IOError as e:
            console.print("No 3rd Image",style="red",justify="center") #Does not exist OR no read permissions
        try:
            imagefn = glob.glob(r'images/' + equipnumber + '/* 4.jpg')

            if imagefn:
                imagesingle = imagefn.pop()
                img = Image.open(imagesingle)
                img3 = openpyxl.drawing.image.Image(img)
                Tempname = imagesingle.strip('.jpg')
                Tempname = Tempname.strip('.JPG')
                tempname = Tempname.strip('4')
                Tempname = Tempname.strip('images/' + equipnumber + '/')
                Tempname = Tempname.strip('\\')
                Iname4 = Tempname
                ws['A50'] = Iname4
                ws.add_image(img3,'A38')
            else:
                pass
        except IOError as e:
            console.print("No 4th Image",style="red",justify="center") #Does not exist OR no read permissions
        try:
            imagefn = glob.glob(r'images/' + equipnumber + '/* 5.jpg')

            if imagefn:
                imagesingle = imagefn.pop()
                img = Image.open(imagesingle)
                img3 = openpyxl.drawing.image.Image(img)
                Tempname = imagesingle.strip('.jpg')
                Tempname = Tempname.strip('.JPG')
                tempname = Tempname.strip('5')
                Tempname = Tempname.strip('images/' + equipnumber + '/')
                Tempname = Tempname.strip('\\')
                Iname5 = Tempname
                ws['F50'] = Iname5
                ws.add_image(img3,'F38')
            else:
                pass
        except IOError as e:
            console.print("No 5th Image",style="red",justify="center") #Does not exist OR no read permissions
        try:
            imagefn = glob.glob(r'images/' + equipnumber + '/* 6.jpg')

            if imagefn:
                imagesingle = imagefn.pop()
                img = Image.open(imagesingle)
                img3 = openpyxl.drawing.image.Image(img)
                Tempname = imagesingle.strip('.jpg')
                Tempname = Tempname.strip('.JPG')
                tempname = Tempname.strip('6')
                Tempname = Tempname.strip('images/' + equipnumber + '/')
                Tempname = Tempname.strip('\\')
                Iname6 = Tempname
                ws['K50'] = Iname6
                ws.add_image(img3,'K38')
            else:
                pass
        except IOError as e:
            console.print("No 6th Image",style="red",justify="center") #Does not exist OR no read permissions

    #save preveous to file in case of crash.
        console.print("Saving Inputes to backup file.",style="red bold blink", justify="center")
        if os.path.exists("crash.bak"):
          os.remove("crash.bak")
        else:
          print("")

        enbak=equipnumber
        eqnbak=equipname
        sdtypesave=sdtype
        hzsave=hz
        p12 = prnsave[0]
        p13 = panelroom[1]
        #prnsavebak=prnsave
        #el objects variables
        elsavebak=elsave
        el1v=el1.svolts
        el1b=el1.bucket
        el1p=el1.sproced
        el2v=el2.svolts
        el2b=el2.bucket
        el2p=el2.sproced
        el3v=el3.svolts
        el3b=el3.bucket
        el3p=el3.sproced
        el4v=el4.svolts
        el4b=el4.bucket
        el4p=el4.sproced
        el5v=el5.svolts
        el5b=el5.bucket
        el5p=el5.sproced
        #ol Object variable
        ols=olsave
        ol1t=ol1.sType
        ol1c=ol1.cType
        ol1n=ol1.name
        ol1p=ol1.proced
        ol1sp=ol1.sproced
        ol2t=ol2.sType
        ol2c=ol2.cType
        ol2n=ol2.name
        ol2p=ol2.proced
        ol2sp=ol2.sproced
        ol3t=ol3.sType
        ol3c=ol3.cType
        ol3n=ol3.name
        ol3p=ol3.proced
        ol3sp=ol3.sproced
        ol4t=ol4.sType
        ol4c=ol4.cType
        ol4n=ol4.name
        ol4p=ol4.proced
        ol4sp=ol4.sproced
        ol5t=ol5.sType
        ol5c=ol5.cType
        ol5n=ol5.name
        ol5p=ol5.proced
        ol5sp=ol5.sproced
        ol6t=ol6.sType
        ol6c=ol6.cType
        ol6n=ol6.name
        ol6p=ol6.proced
        ol6sp=ol6.sproced
        #eStop saves
        essavebak=essave
        es1l=es1.location
        es2l=es2.location
        #save Remote lockout
        rlsaveback=elsave
        rl1l=rl1.location
        rl2l=rl2.location
        rl3l=rl3.location
        rl4l=rl4.location
        rl5l=rl5.location
        rl6l=rl6.location
        rl7l=rl7.location

        crashsave.append(enbak)
        crashsave.append(eqnbak)
        crashsave.append(sdtypesave)
        crashsave.append(hzsave)
        crashsave.append(p12)
        crashsave.append(elsavebak)
        crashsave.append(el1v)
        crashsave.append(el1b)
        crashsave.append(el1p)
        crashsave.append(el2v)
        crashsave.append(el2b)
        crashsave.append(el2p)
        crashsave.append(el3v)
        crashsave.append(el3b)
        crashsave.append(el3p)
        crashsave.append(el4v)
        crashsave.append(el4b)
        crashsave.append(el4p)
        crashsave.append(el5v)
        crashsave.append(el5b)
        crashsave.append(el5p)
        crashsave.append(ols)
        crashsave.append(ol1t)
        crashsave.append(ol1c)
        crashsave.append(ol1n)
        crashsave.append(ol1p)
        crashsave.append(ol1sp)
        crashsave.append(ol2t)
        crashsave.append(ol2c)
        crashsave.append(ol2n)
        crashsave.append(ol2p)
        crashsave.append(ol2sp)
        crashsave.append(ol3t)
        crashsave.append(ol3c)
        crashsave.append(ol3n)
        crashsave.append(ol3p)
        crashsave.append(ol3sp)
        crashsave.append(ol4t)
        crashsave.append(ol4c)
        crashsave.append(ol4n)
        crashsave.append(ol4p)
        crashsave.append(ol4sp)
        crashsave.append(ol5t)
        crashsave.append(ol5c)
        crashsave.append(ol5n)
        crashsave.append(ol5p)
        crashsave.append(ol5sp)
        crashsave.append(ol6t)
        crashsave.append(ol6c)
        crashsave.append(ol6n)
        crashsave.append(ol6p)
        crashsave.append(ol6sp)
        crashsave.append(essavebak)
        crashsave.append(es1l)
        crashsave.append(es2l)
        crashsave.append(rlsaveback)
        crashsave.append(rl1l)
        crashsave.append(rl2l)
        crashsave.append(rl3l)
        crashsave.append(rl4l)
        crashsave.append(rl5l)
        crashsave.append(rl6l)
        crashsave.append(rl7l)
        crashsave.append(p13)
        crashsave.append(hz1)
        with open('crash.bak', 'wb') as fp:

            pickle.dump(crashsave, fp)
        if not os.path.exists('bak/equip/'):
                os.makedirs('bak/equip/')
        with open('bak/equip/'+equipnumber+'.bak', 'wb') as fp:
            pickle.dump(crashsave, fp)

    #card demo table
        if hz is None:
            hz = "No Hazardous Energy Procedure"
        else:
            pass
        console.print("\n")
        console.print(Panel(equipnumber+" | "+equipname+"\nElectric Lockouts: "+str(el)+ " | Other Lockouts: "+str(ol) +"\nShutdown Procedure: "+sp+" | "+hz),justify="center")
        table = Table()

        table.add_column("Volts", style="yellow", no_wrap=True)
        table.add_column("Bucket", style="yellow", no_wrap=True)
        table.add_column("Procedure", style="yellow", no_wrap=True)
        table.add_column("", style="yellow", no_wrap=True)
        table.add_column("Type", style="green", no_wrap=True)
        table.add_column("Device", style="green", no_wrap=True)
        table.add_column("Procedure", style="green", no_wrap=True)

        try:
            table.add_row(el1.volts,el1.bucket,el1.proced,"",ol1.Type,ol1.name,ol1.proced)
            table.add_row(el2.volts,el2.bucket,el2.proced,"",ol2.Type,ol2.name,ol2.proced)
            table.add_row(el3.volts,el3.bucket,el3.proced,"",ol3.Type,ol3.name,ol3.proced)
            table.add_row(el4.volts,el4.bucket,el4.proced,"",ol4.Type,ol4.name,ol4.proced)
            table.add_row(el5.volts,el5.bucket,el5.proced,"",ol5.Type,ol5.name,ol5.proced)
            table.add_row(""       ,""        ,""        ,"",ol6.Type,ol6.name,ol6.proced)
        except:
            pass

        table3 = Table()

        table3.add_column("eStops: "+str(es), style="bright_white", no_wrap=True)
        table3.add_column("Remote Lockouts: "+str(rl), style="bright_white")
        table3.add_column("Images: ", style="bright_white")
        try:
            table3.add_row(str(es1.location),str(rl1.location),str(Iname1))
            table3.add_row(es2.location,str(rl2.location),str(Iname2))
            table3.add_row("",str(rl3.location),str(Iname3))
            table3.add_row("",str(rl4.location),str(Iname4))
            table3.add_row("",str(rl5.location),str(Iname5))
            table3.add_row("",str(rl6.location),str(Iname6))
            table3.add_row("",str(rl6.location),"")
        except:
            pass

        console.print(table)
        console.print(table3)
        if hz == "No Hazardous Energy Procedure":
            hz = None
        if Confirm.ask("Is this info correct?"):
            isgood = True
            pass
        else:
            isgood = False
            continue
        #construct card
        ws['D56'] = equipnumber
        ws['J56'] = equipname
        panelroom2=panelroom[0]
        ws['A10'] = panelroom2
        #Electric Lockouts
        if el != 0:
            if el == 1:
                ws['A14'] = el1.volts
                ws['C14'] = el1.bucket
                ws['G14'] = el1.proced
            if el == 2:
                ws['A14'] = el1.volts
                ws['C14'] = el1.bucket
                ws['G14'] = el1.proced
                ws['A16'] = el2.volts
                ws['C16'] = el2.bucket
                ws['G16'] = el2.proced
            if el == 3:
                ws['A14'] = el1.volts
                ws['C14'] = el1.bucket
                ws['G14'] = el1.proced
                ws['A16'] = el2.volts
                ws['C16'] = el2.bucket
                ws['G16'] = el2.proced
                ws['A18'] = el3.volts
                ws['C18'] = el3.bucket
                ws['G18'] = el3.proced
            if el == 4:
                ws['A14'] = el1.volts
                ws['C14'] = el1.bucket
                ws['G14'] = el1.proced
                ws['A16'] = el2.volts
                ws['C16'] = el2.bucket
                ws['G16'] = el2.proced
                ws['A18'] = el3.volts
                ws['C18'] = el3.bucket
                ws['G18'] = el3.proced
                ws['A20'] = el4.volts
                ws['C20'] = el4.bucket
                ws['G20'] = el4.proced
            if el == 5:
                ws['A14'] = el1.volts
                ws['C14'] = el1.bucket
                ws['G14'] = el1.proced
                ws['A16'] = el2.volts
                ws['C16'] = el2.bucket
                ws['G16'] = el2.proced
                ws['A18'] = el3.volts
                ws['C18'] = el3.bucket
                ws['G18'] = el3.proced
                ws['A20'] = el4.volts
                ws['C20'] = el4.bucket
                ws['G20'] = el4.proced
                ws['A22'] = el5.volts
                ws['C22'] = el5.bucket
                ws['G22'] = el5.proced
        #other lockouts.
        if ol != 0:
            if ol == 1:
                ws['I12'] = ol1.Type
                ws['K12'] = ol1.name
                ws['O12'] = ol1.proced
            if ol == 2:
                ws['I12'] = ol1.Type
                ws['K12'] = ol1.name
                ws['O12'] = ol1.proced
                ws['I14'] = ol2.Type
                ws['K14'] = ol2.name
                ws['O14'] = ol2.proced
            if ol == 3:
                ws['I12'] = ol1.Type
                ws['K12'] = ol1.name
                ws['O12'] = ol2.proced
                ws['I14'] = ol2.Type
                ws['K14'] = ol2.name
                ws['O14'] = ol2.proced
                ws['I16'] = ol3.Type
                ws['K16'] = ol3.name
                ws['O16'] = ol3.proced
            if ol == 4:
                ws['I12'] = ol1.Type
                ws['K12'] = ol1.name
                ws['O12'] = ol2.proced
                ws['I14'] = ol2.Type
                ws['K14'] = ol2.name
                ws['O14'] = ol2.proced
                ws['I16'] = ol3.Type
                ws['K16'] = ol3.name
                ws['O16'] = ol3.proced
                ws['I18'] = ol4.Type
                ws['K18'] = ol4.name
                ws['O18'] = ol4.procedd
            if ol == 5:
                ws['I12'] = ol1.Type
                ws['K12'] = ol1.name
                ws['O12'] = ol2.proced
                ws['I14'] = ol2.Type
                ws['K14'] = ol2.name
                ws['O14'] = ol2.proced
                ws['I16'] = ol3.Type
                ws['K16'] = ol3.name
                ws['O16'] = ol3.proced
                ws['I18'] = ol4.Type
                ws['K18'] = ol4.name
                ws['O18'] = ol4.proced
                ws['I20'] = ol5.Type
                ws['K20'] = ol5.name
                ws['O20'] = ol5.proced
            if ol == 6:
                ws['I12'] = ol1.Type
                ws['K12'] = ol1.name
                ws['O12'] = ol2.proced
                ws['I14'] = ol2.Type
                ws['K14'] = ol2.name
                ws['O14'] = ol2.proced
                ws['I16'] = ol3.Type
                ws['K16'] = ol3.name
                ws['O16'] = ol3.proced
                ws['I18'] = ol4.Type
                ws['K18'] = ol4.name
                ws['O18'] = ol4.proced
                ws['I20'] = ol5.Type
                ws['K20'] = ol5.name
                ws['O20'] = ol5.proced
                ws['I22'] = ol6.Type
                ws['K22'] = ol6.name
                ws['O22'] = ol6.proced

        #More information like estops remote lockouts

        #eStops
        if es != 0:
            if es == 1:
                ws['R46'] = es1.location
            if es == 2:
                ws['R46'] = es1.location
                ws['R47'] = es2.location
        #remote Lockout
        if rl != 0:
            if rl == 1:
                ws['R50'] = rl1.location
            if rl == 2:
                ws['R50'] = rl1.location
                ws['R51'] = rl2.location
            if rl == 3:
                ws['R50'] = rl1.location
                ws['R51'] = rl2.location
                ws['R52'] = rl3.location
            if rl == 4:
                ws['R50'] = rl1.location
                ws['R51'] = rl2.location
                ws['R52'] = rl3.location
                ws['R53'] = rl4.location
            if rl == 5:
                ws['R50'] = rl1.location
                ws['R51'] = rl2.location
                ws['R52'] = rl3.location
                ws['R53'] = rl4.location
                ws['R54'] = rl5.location
            if rl == 6:
                ws['R50'] = rl1.location
                ws['R51'] = rl2.location
                ws['R52'] = rl3.location
                ws['R53'] = rl4.location
                ws['R54'] = rl5.location
                ws['R55'] = rl6.location
            if rl == 7:
                ws['R50'] = rl1.location
                ws['R51'] = rl2.location
                ws['R52'] = rl3.location
                ws['R53'] = rl4.location
                ws['R54'] = rl5.location
                ws['R55'] = rl6.location
                ws['R56'] = rl7.location
        #Shutdown Procedure
        ws['R05'] = sp
        #Hazardous Energy
        if hz is None:
            pass
        else:
            ws['B51'] = "SOP# " + str(hz)
        #create Date
        ws['J54'] = date.today()

        output_file = str(equipnumber) + ".xlsx"
        console.print("SAVING CARD!",justify="center", style="red blink bold")
        if os.path.isfile(savepath+output_file):
           os.remove(savepath+output_file)
           wb.save(savepath+output_file)
        else:
            wb.save(savepath+output_file)

